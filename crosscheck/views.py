import glob
import os
import io
import json
import re
import time
import duckdb
import warnings
import openpyxl
import pandas as pd
import threading
import calendar
from copy import copy
from datetime import datetime
from django.conf import settings
from django.shortcuts import render, redirect
from django.http import FileResponse, HttpRequest, HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt, csrf_protect
from django.core.files.storage import FileSystemStorage
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from docxtpl import DocxTemplate
from openpyxl.cell.text import InlineFont
from openpyxl.cell.rich_text import TextBlock, CellRichText

# --- GLOBAL DUCKDB CONNECTION ---
_GLOBAL_DUCKDB_CONN = None
_DB_LOCK = threading.Lock()

# --- Helpers ---

def get_db_connection():
    global _GLOBAL_DUCKDB_CONN
    
    with _DB_LOCK:
        if _GLOBAL_DUCKDB_CONN is None:
            db_path = str(os.path.join(settings.BASE_DIR, 'datawarehouse.duckdb'))
            _GLOBAL_DUCKDB_CONN = duckdb.connect(db_path)
            
            _GLOBAL_DUCKDB_CONN.execute("""
                CREATE TABLE IF NOT EXISTS sessions (
                    ovatr VARCHAR PRIMARY KEY,
                    company_name VARCHAR,
                    tin VARCHAR,
                    status VARCHAR,
                    total_rows INTEGER,
                    match_rate DOUBLE,
                    created_at TIMESTAMP,
                    last_modified TIMESTAMP
                )
            """)
            
    return _GLOBAL_DUCKDB_CONN.cursor()

def update_session_metadata(con, ovatr, company_name=None, tin=None, status=None, total_rows=None, match_rate=None):
    if not ovatr: return
    now = datetime.now()
    
    exists = con.execute("SELECT 1 FROM sessions WHERE ovatr = ?", [ovatr]).fetchone()
    
    if not exists:
        con.execute("""
            INSERT INTO sessions (ovatr, company_name, tin, status, total_rows, match_rate, created_at, last_modified)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, [ovatr, company_name or 'Unknown', tin or '', status or 'Processing', 0, 0.0, now, now])
    else:
        updates = ["last_modified = ?"]
        params = [now]
        if company_name: 
            updates.append("company_name = ?"); params.append(company_name)
        if tin:
            updates.append("tin = ?"); params.append(tin)
        if status:
            updates.append("status = ?"); params.append(status)
        if total_rows is not None:
            updates.append("total_rows = ?"); params.append(total_rows)
        if match_rate is not None:
            updates.append("match_rate = ?"); params.append(match_rate)
            
        params.append(ovatr)
        query = f"UPDATE sessions SET {', '.join(updates)} WHERE ovatr = ?"
        con.execute(query, params)

def clean_currency(val):
    s = str(val).strip()
    if s.lower() in ['nan', 'none', '', 'nat', '-']:
        return 0.0
    clean_s = re.sub(r'[^\d.-]', '', s)
    if '(' in s and ')' in s:
        clean_s = '-' + re.sub(r'[^\d.]', '', s)
    try:
        return float(clean_s)
    except ValueError:
        return 0.0

def clean_invoice_text(val):    
    if pd.isna(val) or not val:
        return ""
    
    s = str(val).strip()
    if s.endswith('.0'):
        s = s[:-2]
        
    return re.sub(r'[^a-zA-Z0-9]', '', s)

def cleanup_old_files():
    directories = [
        os.path.join(settings.MEDIA_ROOT, 'temp_uploads'),
        os.path.join(settings.MEDIA_ROOT, 'temp_reports')
    ]
    current_time = time.time()
    for folder in directories:
        if not os.path.exists(folder): continue
        for f in glob.glob(os.path.join(folder, '*')):
            try:
                if current_time - os.path.getctime(f) > 86400: os.remove(f)
            except: pass

def to_excel_date(date_val):
    if not date_val or pd.isna(date_val): 
        return None
    
    # If DuckDB already returned a native Python date, return it directly
    from datetime import date, datetime
    if isinstance(date_val, (datetime, date, pd.Timestamp)): 
        return date_val

    for fmt in ('%d-%m-%Y', '%Y-%m-%d', '%d/%m/%Y', '%Y/%m/%d'):
        try:
            return datetime.strptime(str(date_val).strip(), fmt)
        except ValueError:
            continue
    return date_val

# --- Views ---

def new_crosscheck(request):
    return render(request, 'crosscheck/new.html')

def processing_view(request):
    try:
        # Get the active session code
        code = request.GET.get('ovatr_code') or request.session.get('ovatr_code', '')
        
        # Keep the session sticky in Django memory
        if code:
            request.session['ovatr_code'] = code
            
        # Allow the page to load normally!
        return render(request, 'crosscheck/processing.html', {'ovatr_code': code})
        
    except Exception as e:
        import traceback
        from django.http import HttpResponse
        return HttpResponse(f"<h1>Crash in Processing Page</h1><pre>{traceback.format_exc()}</pre>", status=500)

def results_view(request):
    try:
        # Get code from URL
        code = request.GET.get('ovatr_code')
        if code:
            # If found in URL, save it to Django's memory immediately
            request.session['ovatr_code'] = code
        else:
            # If URL is empty (sidebar click), revive it from Django's memory
            code = request.session.get('ovatr_code', '')
            
        context = {'ovatr_code': code}
        return render(request, 'crosscheck/results.html', context)
    except Exception as e:
        import traceback
        from django.http import HttpResponse
        return HttpResponse(f"<h1>Crash in Results Page</h1><pre>{traceback.format_exc()}</pre>", status=500)

def history_view(request):
    return render(request, 'crosscheck/history.html')

def report_view(request):
    try:
        code = request.GET.get('ovatr_code')
        if code:
            request.session['ovatr_code'] = code
        else:
            code = request.session.get('ovatr_code', '')
            
        context = {'ovatr_code': code}
        return render(request, 'crosscheck/report.html', context)
    except Exception as e:
        import traceback
        from django.http import HttpResponse
        return HttpResponse(f"<h1>Crash in Report Page</h1><pre>{traceback.format_exc()}</pre>", status=500)

# --- API: History Data ---

def get_history_api(request):
    try:
        query = request.GET.get('q', '').strip()
        conn = get_db_connection()
        
        sql = """
            SELECT s.ovatr, s.company_name, s.status, s.total_rows, s.match_rate, s.last_modified, c."vatin" as tin
            FROM sessions s
            LEFT JOIN company_info c ON s.ovatr = c."ovatr"
        """
        params = []
        where_clauses = []

        if query:
            where_clauses.append("(s.ovatr ILIKE ? OR s.company_name ILIKE ? OR c.\"vatin\" ILIKE ?)")
            params = [f'%{query}%', f'%{query}%', f'%{query}%']
            
        if where_clauses: sql += " WHERE " + " AND ".join(where_clauses)
        sql += " ORDER BY s.last_modified DESC LIMIT 50"
        
        rows = conn.execute(sql, params).fetchall()
        conn.close()
        
        data = []
        for r in rows:
            last_mod = r[5]
            time_ago = "Just now"
            if last_mod:
                diff = datetime.now() - last_mod
                if diff.days > 0: time_ago = f"{diff.days} days ago"
                elif diff.seconds > 3600: time_ago = f"{diff.seconds // 3600} hours ago"
                elif diff.seconds > 60: time_ago = f"{diff.seconds // 60} mins ago"

            data.append({
                'ovatr': r[0], 'company_name': r[1], 'status': r[2],
                'total_rows': r[3], 'match_rate': round(r[4], 1),
                'last_modified': last_mod.strftime('%Y-%m-%d %H:%M') if last_mod else '',
                'tin': r[6] or 'N/A', 'time_ago': time_ago
            })
        return JsonResponse({'status': 'success', 'data': data})
    except Exception as e: return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

# --- Upload & Save APIs ---

@csrf_exempt
def upload_init(request):
    cleanup_old_files()
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        fs = FileSystemStorage()
        clean_name = fs.get_available_name(file.name)
        filename = fs.save(os.path.join("temp", clean_name), file)
        uploaded_file_path = fs.path(filename)

        try:
            try:
                df = pd.read_excel(uploaded_file_path, sheet_name='COMPANY INFO', header=None)
            except:
                df = pd.read_excel(uploaded_file_path, sheet_name=0, header=None)
            
            data_map = {
                'company_name_kh': '', 'company_name_en': '', 'file_barcode': '',
                'old_vatin': '', 'vatin': '', 'enterprise_id': '',
                'registered_entity': '', 'reg_date': '', 'success_date': '', 
                'taxpayer_type': '', 'status': '', 'tax_year': '',
                'address_main': '', 'address_office': '', 'phone': '', 'email': '',
                'employee_count': '', 'total_salary': '', 'property_type': '',
                'rent_per_month': '', 'enterprise_form': '', 'add_ent_form': '', 'signage': '',
                'h_date': '', 'h_real_12m': '', 'h_real_3m': '', 'h_est_12m': '', 'h_est_3m': '',
                
                'i_auditor_names': '', 'i_request_date': '', 'i_contact_person': '', 'i_contact_position': '',
                'i_audit_timeline': '', 'i_amount_requested': '', 'i_moc_number': '', 'i_moc_date': '',
                'i_patent_date': '', 'i_vat_cert_date': '', 'i_patent_amount': '',
                
                'business_activities': [], 'enterprise_accounts': [], 'related_institutions': []
            }

            def get_val_safe(val):
                s = str(val).strip() if pd.notna(val) else ""
                if s.endswith(".0"): s = s[:-2]
                return s

            def get_col(row, idx):
                return get_val_safe(row[idx]) if idx < len(row) else ""

            def extract_val_smart(row):
                c0 = get_col(row, 0)
                val = ""
                if '៖' in c0:
                    parts = c0.split('៖')
                    if len(parts) > 1: val = parts[1].strip()
                if not val: val = get_col(row, 1)
                return val.replace('"', '').replace("'", "")

            def clean_khmer_only(text):
                if pd.isna(text): return ""
                text = str(text)
                cleaned = re.sub(r'[A-Za-z]', '', text)
                return " ".join(cleaned.split())

            current_section = None 
            header_found = False
            estimate_header_index = None

            for index, row in df.iterrows():
                cell_0 = get_col(row, 0)
                
                if "ការប៉ាន់ស្មានផលរបរ" in cell_0:
                    estimate_header_index = index
                
                if "ឈ្មោះសហគ្រាសជាអក្សរខ្មែរ" in cell_0: data_map['company_name_kh'] = extract_val_smart(row)
                elif "ឈ្មោះសហគ្រាសជាអក្សរឡាតាំង" in cell_0: data_map['company_name_en'] = extract_val_smart(row)
                elif "លេខបារកូដឯកសារ" in cell_0: data_map['file_barcode'] = extract_val_smart(row)
                elif "លេខអត្តសញ្ញាណកម្មចាស់" in cell_0: data_map['old_vatin'] = extract_val_smart(row)
                elif "លេខអត្តសញ្ញាណកម្ម" in cell_0: data_map['vatin'] = extract_val_smart(row)
                elif "លេខកាតសម្គាល់សហគ្រាស" in cell_0: data_map['enterprise_id'] = extract_val_smart(row)
                elif "ចុះបញ្ជីនៅ" in cell_0: data_map['registered_entity'] = extract_val_smart(row)
                elif "កាលបរិច្ឆេទចុះបញ្ជី" in cell_0: data_map['reg_date'] = extract_val_smart(row)
                elif "កាលបរិច្ឆេទជោគជ័យ" in cell_0: data_map['success_date'] = extract_val_smart(row)
                elif "ប្រភេទអ្នកជាប់ពន្ធ" in cell_0: data_map['taxpayer_type'] = extract_val_smart(row)
                elif "ស្ថានភាព" in cell_0: data_map['status'] = extract_val_smart(row)
                elif "ទ្រង់ទ្រាយសហគ្រាស" in cell_0 and "បន្ថែម" not in cell_0: data_map['enterprise_form'] = extract_val_smart(row)
                elif "ទ្រង់ទ្រាយសហគ្រាសបន្ថែម" in cell_0: data_map['add_ent_form'] = extract_val_smart(row)
                elif "ឆ្នាំជាប់ពន្ធ" in cell_0: data_map['tax_year'] = extract_val_smart(row)
                elif "អាសយដ្ឋានអាជីវកម្មគោលដេីម" in cell_0: data_map['address_main'] = extract_val_smart(row)
                elif "អាសយដ្ឋានទីចាត់ការ" in cell_0: data_map['address_office'] = extract_val_smart(row)
                elif "លេខទូរសព្ទ" in cell_0: data_map['phone'] = extract_val_smart(row)
                elif "សារអេឡិចត្រូនិក" in cell_0: data_map['email'] = extract_val_smart(row)
                elif "អចលនទ្រព្យ" in cell_0: data_map['property_type'] = extract_val_smart(row)
                elif "ផ្លាកយីហោ" in cell_0: data_map['signage'] = extract_val_smart(row)
                elif "ថ្លៃឈ្នួល/១ខែ" in cell_0: data_map['rent_per_month'] = extract_val_smart(row)
                elif "ចំនួននិយោជិក" in cell_0: data_map['employee_count'] = extract_val_smart(row)
                elif "ប្រាក់ខែសរុប" in cell_0: data_map['total_salary'] = extract_val_smart(row)

                if "សកម្មភាពអាជីវកម្ម" in cell_0:
                    current_section = 'business_activities'; header_found = False; continue
                elif "គណនីសហគ្រាស" in cell_0:
                    current_section = 'enterprise_accounts'; header_found = False; continue
                elif "ស្ថាប័នពាក់ព័ន្ធ" in cell_0:
                    current_section = 'related_institutions'; header_found = False; continue
                
                if current_section:
                    is_empty = all(pd.isna(val) or str(val).strip() == "" for val in row[:3])
                    if is_empty:
                        if header_found: current_section = None
                        continue

                    if not header_found:
                        row_str = str(row.values).lower()
                        if "ល.រ" in row_str or "no" in row_str or "code" in row_str: header_found = True
                        continue

                    if current_section == 'business_activities':
                        data_map['business_activities'].append({
                            'no': get_col(row, 1), 'code': get_col(row, 2),
                            'name': clean_khmer_only(get_col(row, 3)), 'desc': clean_khmer_only(get_col(row, 4)),
                            'type': get_col(row, 5)
                        })
                    elif current_section == 'enterprise_accounts':
                        data_map['enterprise_accounts'].append({
                            'no': get_col(row, 1), 'bank': get_col(row, 2), 'number': get_col(row, 3), 
                            'account_name': get_col(row, 4), 'currency': get_col(row, 5), 'type': get_col(row, 6)
                        })
                    elif current_section == 'related_institutions':
                        data_map['related_institutions'].append({
                            'no': get_col(row, 1), 'name': get_col(row, 2), 
                            'ref': get_col(row, 3), 'date': get_col(row, 4)
                        })

            if estimate_header_index is not None:
                i = estimate_header_index
                def get_cell(r, c):
                    if r < len(df) and c < len(df.columns): return get_val_safe(df.iat[r, c])
                    return ""
                data_map['h_date'] = get_cell(i + 2, 2)
                data_map['h_real_12m'] = get_cell(i + 4, 2)
                data_map['h_real_3m'] = get_cell(i + 4, 3)
                data_map['h_est_12m'] = get_cell(i + 5, 2)
                data_map['h_est_3m'] = get_cell(i + 5, 3)

            return JsonResponse({'status': 'success', 'data': data_map, 'temp_path': filename})

        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})
            
    return JsonResponse({'status': 'error', 'message': 'No file provided'})

@csrf_exempt
def save_company_info(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            clean_data = {
                k.lower(): (json.dumps(v, ensure_ascii=False) if isinstance(v, (list, dict)) else v)
                for k, v in data.items()
            }
            
            if 'ovatr' not in clean_data or not clean_data['ovatr']:
                return JsonResponse({'status': 'error', 'message': 'Missing Critical Field: OVATR'}, status=400)

            ovatr = clean_data['ovatr']
            comp_name = clean_data.get('company_name_kh') or clean_data.get('company_name_en') or 'Unknown Company'

            request.session['ovatr_code'] = ovatr

            con = get_db_connection()
            
            table_check = con.execute("SELECT 1 FROM information_schema.tables WHERE table_name = 'company_info'").fetchone()
            
            if not table_check:
                columns_schema = [f'"{k}" VARCHAR PRIMARY KEY' if k == 'ovatr' else f'"{k}" VARCHAR' for k in clean_data.keys()]
                con.execute(f"CREATE TABLE company_info ({', '.join(columns_schema)})")
            else:
                existing_cols_res = con.execute("DESCRIBE company_info").fetchall()
                existing_cols = [row[0].lower() for row in existing_cols_res]
                
                for key in clean_data.keys():
                    if key.lower() not in existing_cols:
                        con.execute(f'ALTER TABLE company_info ADD COLUMN "{key}" VARCHAR')

            columns = [f'"{k}"' for k in clean_data.keys()]
            placeholders = ['?'] * len(clean_data)
            values = list(clean_data.values())
            
            con.execute(f"INSERT OR REPLACE INTO company_info ({', '.join(columns)}) VALUES ({', '.join(placeholders)})", values)
            
            update_session_metadata(con, ovatr, company_name=comp_name, status="Processing")

            con.close()
            return JsonResponse({'status': 'success', 'message': 'Company Info saved successfully'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    return JsonResponse({'status': 'error', 'message': 'Invalid method'}, status=405)

@csrf_exempt
def save_taxpaid(request):
    if request.method == 'POST':
        try:
            body = json.loads(request.body)
            ovatr_val = body.get('ovatr') or body.get('OVATR')

            fs = FileSystemStorage()
            full_path = fs.path(body['temp_path'])
            try:
                df = pd.read_excel(full_path, sheet_name='TAXPAID', header=None)
            except ValueError:
                return JsonResponse({'status': 'error', 'message': 'Sheet "TAXPAID" not found'}, status=400)

            extracted_rows = []
            current_year = None
            
            def clean_money(val):
                s = str(val).strip().replace(',', '')
                if s in ['-', '', 'nan', 'None']: return 0.0
                try: return float(s)
                except ValueError: return 0.0

            for idx, row in df.iterrows():
                row_vals = [str(x).strip() for x in row.values]
                col0 = row_vals[0] if len(row_vals) > 0 else ""
                col1 = row_vals[1] if len(row_vals) > 1 else ""

                if "ព័ត៌មានលម្អិតប្រចាំឆ្នាំ" in col0:
                    found_year = None
                    if col1.isdigit(): found_year = col1
                    elif re.search(r'\d{4}', col0): found_year = re.search(r'\d{4}', col0).group()
                    if found_year: current_year = found_year
                    continue

                if "មករា" in str(row_vals): continue

                if current_year and len(row_vals) > 15:
                    description = row_vals[2]
                    if not description or description.lower() in ['nan', 'close', ''] or description == "ឆ្នាំបង់ពន្ធ": continue

                    extracted_rows.append({
                        'ovatr': ovatr_val, 'tax_year': current_year, 'description': description,
                        'jan': clean_money(row.values[3]), 'feb': clean_money(row.values[4]), 'mar': clean_money(row.values[5]),
                        'apr': clean_money(row.values[6]), 'may': clean_money(row.values[7]), 'jun': clean_money(row.values[8]),
                        'jul': clean_money(row.values[9]), 'aug': clean_money(row.values[10]), 'sep': clean_money(row.values[11]),
                        'oct': clean_money(row.values[12]), 'nov': clean_money(row.values[13]), 'dec': clean_money(row.values[14]),
                        'total': clean_money(row.values[15]),
                    })

            if extracted_rows:
                con = get_db_connection()
                con.execute("CREATE TABLE IF NOT EXISTS tax_paid (ovatr VARCHAR, tax_year VARCHAR, description VARCHAR, jan DOUBLE, feb DOUBLE, mar DOUBLE, apr DOUBLE, may DOUBLE, jun DOUBLE, jul DOUBLE, aug DOUBLE, sep DOUBLE, oct DOUBLE, nov DOUBLE, dec DOUBLE, total DOUBLE, PRIMARY KEY (ovatr, tax_year, description))")
                con.execute("DELETE FROM tax_paid WHERE ovatr = ?", [ovatr_val])
                con.executemany("INSERT INTO tax_paid VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)", [list(d.values()) for d in extracted_rows])
                con.close()
                return JsonResponse({'status': 'success', 'message': f'Saved {len(extracted_rows)} records for TaxPaid.'})
            return JsonResponse({'status': 'warning', 'message': 'No valid tax data found in TAXPAID sheet.'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    return JsonResponse({'status': 'error', 'message': 'Invalid Method'}, status=405)

@csrf_exempt
def save_purchase(request):
    if request.method == 'POST':
        try:
            body = json.loads(request.body)
            ovatr_val = body.get('ovatr') or body.get('OVATR')

            fs = FileSystemStorage()
            try:
                df = pd.read_excel(fs.path(body['temp_path']), sheet_name='PURCHASE', header=None)
            except ValueError:
                return JsonResponse({'status': 'error', 'message': 'Sheet "PURCHASE" not found'}, status=400)

            df = df.iloc[3:]
            if len(df.columns) < 17:
                return JsonResponse({'status': 'error', 'message': f'Format Mismatch: Expected 17 columns (A-Q), found {len(df.columns)}.'})

            target_cols = [
                'excel_no', 'date', 'invoice_no', 'type', 'supplier_tin', 'supplier_name', 
                'total_amount', 'exclude_vat', 'non_vat_purchase', 'vat_0', 
                'purchase', 'import', 'non_creditable_vat', 'purchase_state_charge', 'import_state_charge', 
                'description', 'status'
            ]
            df = df.iloc[:, :17]; df.columns = target_cols
            df = df[df['date'].notna()]
            df['no'] = range(1, len(df) + 1); df['no'] = df['no'].astype(str)

            for col in ['total_amount', 'exclude_vat', 'non_vat_purchase', 'vat_0', 'purchase', 'import', 'non_creditable_vat', 'purchase_state_charge', 'import_state_charge']:
                df[col] = df[col].apply(clean_currency)

            df['ovatr'] = ovatr_val
            df['user_status'] = None
            df['comment'] = ''  # New Comment Support
            
            con = get_db_connection()
            
            con.execute("""
                CREATE TABLE IF NOT EXISTS purchase (
                    ovatr VARCHAR, no VARCHAR, date VARCHAR, invoice_no VARCHAR, type VARCHAR, 
                    supplier_tin VARCHAR, supplier_name VARCHAR, total_amount DOUBLE, 
                    exclude_vat DOUBLE, non_vat_purchase DOUBLE, vat_0 DOUBLE, purchase DOUBLE, 
                    import DOUBLE, non_creditable_vat DOUBLE, purchase_state_charge DOUBLE, 
                    import_state_charge DOUBLE, description VARCHAR, status VARCHAR, 
                    user_status VARCHAR, comment VARCHAR,
                    PRIMARY KEY (ovatr, no)
                )
            """)
            
            try: con.execute("ALTER TABLE purchase ADD COLUMN user_status VARCHAR")
            except: pass
            try: con.execute("ALTER TABLE purchase ADD COLUMN comment VARCHAR DEFAULT ''")
            except: pass
            try: con.execute("ALTER TABLE purchase ADD COLUMN approve_amount FLOAT DEFAULT 0.0")
            except: pass
            try: con.execute("ALTER TABLE purchase ADD COLUMN annex2_note VARCHAR DEFAULT ''")
            except: pass

            con.execute("DELETE FROM purchase WHERE ovatr = ?", [ovatr_val])
            con.register('df_purchase', df)
            
            con.execute("""
                INSERT INTO purchase (
                    ovatr, no, date, invoice_no, type, supplier_tin, supplier_name, 
                    total_amount, exclude_vat, non_vat_purchase, vat_0, purchase, 
                    import, non_creditable_vat, purchase_state_charge, import_state_charge, 
                    description, status, user_status, comment
                )
                SELECT 
                    ovatr, no, date, invoice_no, type, supplier_tin, supplier_name, 
                    total_amount, exclude_vat, non_vat_purchase, vat_0, purchase, 
                    import, non_creditable_vat, purchase_state_charge, import_state_charge, 
                    description, status, user_status, comment 
                FROM df_purchase
            """)
            con.close()
            return JsonResponse({'status': 'success', 'message': f'Saved {len(df)} Purchase Invoices.'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    return JsonResponse({'status': 'error', 'message': 'Invalid Method'}, status=405)

@csrf_exempt
def save_sale(request):
    if request.method == 'POST':
        try:
            body = json.loads(request.body)
            ovatr_val = body.get('ovatr') or body.get('OVATR')

            fs = FileSystemStorage()
            try:
                df = pd.read_excel(fs.path(body['temp_path']), sheet_name='SALE', header=None)
            except ValueError:
                return JsonResponse({'status': 'error', 'message': 'Sheet "SALE" not found'}, status=400)

            df = df.iloc[3:]
            if len(df.columns) < 23:
                 return JsonResponse({'status': 'error', 'message': f'Format Mismatch: Expected 23+ columns (A-W), found {len(df.columns)}'})

            target_cols = [
                'excel_no', 'date', 'invoice_no', 'credit_note_no', 'buyer_type', 'tax_registration_id', 
                'buyer_name', 'total_invoice_amount', 'amount_exclude_vat', 'non_vat_sales', 
                'vat_zero_rate', 'vat_local_sale', 'vat_export', 'vat_local_sale_state_burden', 
                'vat_withheld_by_national_treasury', 'plt', 'special_tax_on_goods', 
                'special_tax_on_services', 'accommodation_tax', 'income_tax_redemption_rate', 
                'notes', 'description', 'tax_declaration_status'
            ]
            df = df.iloc[:, :23]; df.columns = target_cols
            df = df[df['date'].notna()]
            df['no'] = range(1, len(df) + 1); df['no'] = df['no'].astype(str)

            numeric_cols = [
                'total_invoice_amount', 'amount_exclude_vat', 'non_vat_sales', 'vat_zero_rate', 
                'vat_local_sale', 'vat_export', 'vat_local_sale_state_burden', 
                'vat_withheld_by_national_treasury', 'plt', 'special_tax_on_goods', 
                'special_tax_on_services', 'accommodation_tax', 'income_tax_redemption_rate'
            ]
            for col in numeric_cols:
                df[col] = df[col].apply(clean_currency)

            df['ovatr'] = ovatr_val
            
            con = get_db_connection()
            con.execute("""
                CREATE TABLE IF NOT EXISTS sale (
                    ovatr VARCHAR, no VARCHAR, date VARCHAR, invoice_no VARCHAR, credit_note_no VARCHAR,
                    buyer_type VARCHAR, tax_registration_id VARCHAR, buyer_name VARCHAR,
                    total_invoice_amount DOUBLE, amount_exclude_vat DOUBLE, non_vat_sales DOUBLE,
                    vat_zero_rate DOUBLE, vat_local_sale DOUBLE, vat_export DOUBLE,
                    vat_local_sale_state_burden DOUBLE, vat_withheld_by_national_treasury DOUBLE, plt DOUBLE,
                    special_tax_on_goods DOUBLE, special_tax_on_services DOUBLE, accommodation_tax DOUBLE,
                    income_tax_redemption_rate DOUBLE, notes VARCHAR, description VARCHAR,
                    tax_declaration_status VARCHAR, PRIMARY KEY (ovatr, no)
                )
            """)
            con.execute("DELETE FROM sale WHERE ovatr = ?", [ovatr_val])
            con.register('df_sale', df)
            con.execute("""
                INSERT INTO sale 
                SELECT 
                    ovatr, no, date, invoice_no, credit_note_no, buyer_type, 
                    tax_registration_id, buyer_name, total_invoice_amount, 
                    amount_exclude_vat, non_vat_sales, vat_zero_rate, 
                    vat_local_sale, vat_export, vat_local_sale_state_burden, 
                    vat_withheld_by_national_treasury, plt, special_tax_on_goods, 
                    special_tax_on_services, accommodation_tax, 
                    income_tax_redemption_rate, notes, description, 
                    tax_declaration_status
                FROM df_sale
            """)
            con.close()
            return JsonResponse({'status': 'success', 'message': f'Saved {len(df)} Sale Invoices.'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    return JsonResponse({'status': 'error', 'message': 'Invalid Method'}, status=405)

@csrf_exempt
def save_reverse_charge(request):
    if request.method == 'POST':
        try:
            body = json.loads(request.body)
            ovatr_val = body.get('ovatr') or body.get('OVATR')

            fs = FileSystemStorage()
            try:
                try: df = pd.read_excel(fs.path(body['temp_path']), sheet_name='REVERSE_CHARGE', header=None)
                except: df = pd.read_excel(fs.path(body['temp_path']), sheet_name='REVERSE CHARGE', header=None)
            except ValueError:
                return JsonResponse({'status': 'error', 'message': 'Sheet "REVERSE_CHARGE" not found'}, status=400)

            df = df.iloc[3:]
            if len(df.columns) < 14:
                 return JsonResponse({'status': 'error', 'message': f'Format Mismatch: Expected 14+ columns, found {len(df.columns)}'})

            target_cols = [
                'excel_no', 'date', 'invoice_no', 'supplier_non_resident', 'supplier_tin', 
                'supplier_name', 'address', 'email', 'non_vat_supply', 'exclude_vat', 
                'vat', 'description', 'status', 'declaration_status'
            ]
            df = df.iloc[:, :14]; df.columns = target_cols
            df = df[df['date'].notna()]
            df['no'] = range(1, len(df) + 1); df['no'] = df['no'].astype(str)

            for col in ['non_vat_supply', 'exclude_vat', 'vat']:
                df[col] = df[col].apply(clean_currency)

            df['ovatr'] = ovatr_val
            
            con = get_db_connection()
            con.execute("""
                CREATE TABLE IF NOT EXISTS reverse_charge (
                    ovatr VARCHAR, no VARCHAR, date VARCHAR, invoice_no VARCHAR, 
                    supplier_non_resident VARCHAR, supplier_tin VARCHAR, supplier_name VARCHAR, 
                    address VARCHAR, email VARCHAR, non_vat_supply DOUBLE, exclude_vat DOUBLE, 
                    vat DOUBLE, description VARCHAR, status VARCHAR, declaration_status VARCHAR, 
                    PRIMARY KEY (ovatr, no)
                )
            """)
            con.execute("DELETE FROM reverse_charge WHERE ovatr = ?", [ovatr_val])
            con.register('df_rc', df)
            con.execute("""
                INSERT INTO reverse_charge 
                SELECT 
                    ovatr, no, date, invoice_no, supplier_non_resident, 
                    supplier_tin, supplier_name, address, email, 
                    non_vat_supply, exclude_vat, vat, description, 
                    status, declaration_status 
                FROM df_rc
            """)
            con.close()
            return JsonResponse({'status': 'success', 'message': f'Saved {len(df)} Reverse Charge Records.'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    return JsonResponse({'status': 'error', 'message': 'Invalid Method'}, status=405)

# --- Analytics & Reporting ---

@csrf_exempt
def update_result_row(request):
    """
    SMART UPDATE: 
    Updates the DuckDB tables and triggers an Excel rebuild.
    Includes Comment and User Status saving.
    """
    def format_db_date(val):
        if not val or str(val).strip().lower() in ['nan', 'none', 'nat', '']: 
            return None 
        v = str(val).strip()
        if re.match(r'^\d{4}-\d{2}-\d{2}', v): return v[:10]
        for fmt in ('%d-%m-%Y', '%d/%m/%Y', '%m/%d/%Y', '%m-%d-%Y'):
            try: return datetime.strptime(v, fmt).strftime('%Y-%m-%d')
            except ValueError: pass
        return None

    if request.method == 'POST':
        con = None
        try:
            body = json.loads(request.body)
            ovatr = body.get('ovatr')
            row_no = str(body.get('no', '')).strip() 
            updates = body.get('updates', {})
            history_data = body.get('history', {}) 
            
            con = get_db_connection()
            
            try: con.execute("ALTER TABLE purchase ADD COLUMN user_status VARCHAR")
            except: pass
            try: con.execute("ALTER TABLE purchase ADD COLUMN comment VARCHAR DEFAULT ''")
            except: pass
            
            # --- 1. Map Purchase Table Updates ---
            db_updates = {}
            if 'p_desc' in updates: db_updates['description'] = str(updates['p_desc'])
            if 'p_supp' in updates: db_updates['supplier_name'] = str(updates['p_supp'])
            if 'p_tin' in updates: db_updates['supplier_tin'] = str(updates['p_tin'])
            if 'p_inv' in updates: db_updates['invoice_no'] = str(updates['p_inv'])
            if 'p_date' in updates: db_updates['date'] = format_db_date(updates['p_date']) 
            if 'p_comment' in updates: db_updates['comment'] = str(updates['p_comment']) 
            
            if 'user_status' in updates: 
                db_updates['user_status'] = str(updates['user_status']) if updates['user_status'] else None
            
            if 'p_amt' in updates:
                try: amt = clean_currency(updates['p_amt'])
                except: amt = updates['p_amt']
                if body.get('type', 'local') == 'import': db_updates['"import"'] = amt
                else: db_updates['purchase'] = amt

            # --- 2. Map Tax Declaration Updates (All 22 Fields) ---
            d_data = updates.get('d_data', {})
            d_field_map = {
                'date': 'date', 'invoice_no': 'invoice_number', 'credit_no': 'credit_notification_letter_number',
                'buyer_type': 'buyer_type', 'tin': 'tax_registration_id', 'name': 'buyer_name',
                'total_amt': 'total_invoice_amount', 'excl_vat': 'amount_exclude_vat', 'non_vat': 'non_vat_sales',
                'vat_0': 'vat_zero_rate', 'vat_local': 'vat_local_sale', 'vat_export': 'vat_export',
                'state_burden': 'vat_local_sale_state_burden', 'withheld': 'vat_withheld_by_national_treasury',
                'plt': 'plt', 'spec_goods': 'special_tax_on_goods', 'spec_serv': 'special_tax_on_services',
                'accom': 'accommodation_tax', 'redemption': 'income_tax_redemption_rate', 'notes': 'notes',
                'desc': 'description', 'dec_status': 'tax_declaration_status'
            }
            numeric_fields = {
                'total_amt', 'excl_vat', 'non_vat', 'vat_0', 'vat_local', 'vat_export',
                'state_burden', 'withheld', 'plt', 'spec_goods', 'spec_serv', 'accom', 'redemption'
            }

            d_updates = {}
            for f_key, db_col in d_field_map.items():
                if f_key in d_data:
                    val = d_data[f_key]
                    if f_key in numeric_fields:
                        try: val = clean_currency(val)
                        except: val = 0.0
                    elif f_key == 'date': val = format_db_date(val) 
                    else: val = str(val) if val is not None else ""
                    d_updates[db_col] = val

            orig_inv = updates.get('original_d_inv')
            orig_tin = updates.get('original_d_tin')

            # --- SMART HISTORY FALLBACK ---
            if (not orig_inv or str(orig_inv).strip() == '') and d_updates:
                try:
                    recent_inv_change = con.execute("SELECT new_value FROM change_history WHERE ovatr = ? AND row_no = ? AND (field = 'd_data.invoice_no' OR field = 'd_inv') ORDER BY timestamp DESC LIMIT 1", [ovatr, row_no]).fetchone()
                    if recent_inv_change and recent_inv_change[0]: orig_inv = recent_inv_change[0]
                        
                    recent_tin_change = con.execute("SELECT new_value FROM change_history WHERE ovatr = ? AND row_no = ? AND (field = 'd_data.tin' OR field = 'd_tin') ORDER BY timestamp DESC LIMIT 1", [ovatr, row_no]).fetchone()
                    if recent_tin_change and recent_tin_change[0]: orig_tin = recent_tin_change[0]
                        
                    if not orig_inv:
                        p_info = con.execute("SELECT invoice_no FROM purchase WHERE ovatr = ? AND CAST(no AS VARCHAR) = ?", [ovatr, row_no]).fetchone()
                        if p_info:
                            orig_inv = p_info[0]
                            if not orig_tin:
                                user_tin_row = con.execute("SELECT vatin FROM company_info WHERE ovatr = ?", [ovatr]).fetchone()
                                if user_tin_row: orig_tin = user_tin_row[0]
                except Exception as e:
                    print(f"History fallback error: {e}")

            con.execute("CREATE TABLE IF NOT EXISTS change_history (timestamp TIMESTAMP, ovatr VARCHAR, row_no VARCHAR, table_type VARCHAR, field VARCHAR, old_value VARCHAR, new_value VARCHAR)")
            current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            for field, vals in history_data.items():
                old_v = str(vals.get('old', ''))
                new_v = str(vals.get('new', ''))
                if old_v != new_v:
                    con.execute("INSERT INTO change_history VALUES (?, ?, ?, ?, ?, ?, ?)", [current_time, ovatr, row_no, body.get('type', 'local'), field, old_v, new_v])

            # --- EXECUTE PURCHASE UPDATE ---
            if db_updates:
                set_clause = ", ".join([f"{k} = ?" for k in db_updates.keys()])
                params = list(db_updates.values()) + [ovatr, row_no]
                con.execute(f"UPDATE purchase SET {set_clause} WHERE ovatr = ? AND CAST(no AS VARCHAR) = ?", params)

            # --- EXECUTE TAX DECLARATION UPDATE ---
            if orig_inv and d_updates:
                d_set_clause = [f"{k} = ?" for k in d_updates.keys()]
                d_params = list(d_updates.values())
                query_where = "WHERE regexp_replace(upper(invoice_number), '[^A-Z0-9]', '', 'g') = ?"
                d_params.append(clean_invoice_text(orig_inv))
                if orig_tin:
                    query_where += " AND regexp_replace(upper(tax_registration_id), '[^A-Z0-9]', '', 'g') = ?"
                    d_params.append(clean_invoice_text(orig_tin))
                con.execute(f"UPDATE tax_declaration SET {', '.join(d_set_clause)} {query_where}", d_params)

            try: update_session_metadata(con, ovatr)
            except Exception: pass

            con.commit()
            con.close()
            con = None
            
            try:
                mock_request = HttpRequest()
                mock_request.GET = {'ovatr_code': ovatr}
                gen_res = generate_annex_iii(mock_request)
                if gen_res.status_code != 200: return gen_res 
            except Exception as e: pass

            return JsonResponse({'status': 'success', 'message': 'Row updated'})
        except Exception as e:
            if con: con.close()
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    return JsonResponse({'status': 'error', 'message': 'Invalid Method'}, status=405)

def get_row_history(request):
    try:
        ovatr = request.GET.get('ovatr')
        row_no = request.GET.get('no')
        if not ovatr or not row_no:
            return JsonResponse({'status': 'error', 'message': 'Missing params'}, status=400)

        con = get_db_connection()
        try:
            con.execute("SELECT 1 FROM change_history LIMIT 1")
        except:
            con.close()
            return JsonResponse({'status': 'success', 'data': []})

        data = con.execute("""
            SELECT timestamp, field, old_value, new_value 
            FROM change_history 
            WHERE ovatr = ? AND row_no = ? 
            ORDER BY timestamp DESC
        """, [ovatr, row_no]).fetchall()
        con.close()

        history = []
        for row in data:
            history.append({
                'timestamp': str(row[0]),
                'field': row[1],
                'old_value': row[2],
                'new_value': row[3]
            })
            
        return JsonResponse({'status': 'success', 'data': history})
    except Exception as e:
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

def get_crosscheck_stats(request):
    ovatr_code = request.GET.get('ovatr_code') or request.session.get('ovatr_code')
    if not ovatr_code: 
        return JsonResponse({'status': 'error', 'message': 'Missing OVATR code'}, status=400)

    try:
        conn = get_db_connection()
        
        res_p = conn.execute("""
            SELECT 
                COUNT(CASE WHEN purchase > 0 THEN 1 END),
                COUNT(CASE WHEN "import" > 0 THEN 1 END)
            FROM purchase WHERE ovatr = ?
        """, [ovatr_code]).fetchone()
        
        count_local = res_p[0] if res_p else 0
        count_import = res_p[1] if res_p else 0
        total_rows = count_local + count_import
        
        res_d = conn.execute("""
            SELECT COUNT(DISTINCT d.id)
            FROM tax_declaration d
            JOIN purchase p ON 
                regexp_replace(upper(d.invoice_number), '[^A-Z0-9]', '', 'g') = regexp_replace(upper(p.invoice_no), '[^A-Z0-9]', '', 'g')
            JOIN company_info c ON p.ovatr = c.ovatr
            WHERE p.ovatr = ?
            AND regexp_replace(upper(d.tax_registration_id), '[^A-Z0-9]', '', 'g') = regexp_replace(upper(c.vatin), '[^A-Z0-9]', '', 'g')
            AND month(d.date) = month(COALESCE(try_cast(p.date as DATE), strptime(p.date, '%d-%m-%Y')))
            AND year(d.date) = year(COALESCE(try_cast(p.date as DATE), strptime(p.date, '%d-%m-%Y')))
        """, [ovatr_code]).fetchone()
        
        count_d = res_d[0] if res_d else 0
        
        match_rate = (count_d / total_rows * 100) if total_rows > 0 else 0.0
        update_session_metadata(conn, ovatr_code, total_rows=total_rows, match_rate=match_rate, status="Completed")
        
        conn.close()
        
        file_path = os.path.join(settings.MEDIA_ROOT, 'temp_reports', f"AnnexIII_{ovatr_code}.xlsx")
        
        return JsonResponse({
            'status': 'success',
            'total_rows': max(total_rows, count_d),
            'purchase_count': total_rows, 
            'local_count': count_local,
            'import_count': count_import,
            'declaration_count': count_d,
            'is_ready': os.path.exists(file_path)
        })
        
    except Exception as e: 
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

@csrf_exempt
def save_report_summary(request):
    """
    Receives JSON payload from the frontend containing the extracted 
    AnnexIII-Local Pur summary table and saves it to DuckDB.
    """
    if request.method == 'POST':
        con = get_db_connection()
        try:
            data = json.loads(request.body)
            summary_data = data.get('summary_data', [])
            
            if not summary_data:
                return JsonResponse({'status': 'error', 'message': 'No summary data provided'}, status=400)
            
            ovatr_code = summary_data[0].get('ovatr')
            if not ovatr_code:
                return JsonResponse({'status': 'error', 'message': 'Missing ovatr code in payload'}, status=400)
            
            # 1. Ensure the table exists
            con.execute("""
                CREATE TABLE IF NOT EXISTS report_summary (
                    ovatr VARCHAR,
                    description VARCHAR,
                    total_amount VARCHAR,
                    other VARCHAR
                )
            """)
            
            # 2. Clear old summary data for this session to prevent duplicates
            con.execute("DELETE FROM report_summary WHERE ovatr = ?", [ovatr_code])
            
            # 3. Prepare data for bulk insert
            insert_values = [
                (
                    str(row.get('ovatr', '')), 
                    str(row.get('description', '')), 
                    str(row.get('total_amount', '0')), 
                    str(row.get('other', ''))
                )
                for row in summary_data
            ]
            
            # 4. Bulk Insert
            con.executemany("""
                INSERT INTO report_summary (ovatr, description, total_amount, other) 
                VALUES (?, ?, ?, ?)
            """, insert_values)
            
            return JsonResponse({'status': 'success', 'message': f'Saved {len(insert_values)} rows to report_summary.'})
            
        except Exception as e:
            import traceback
            print(traceback.format_exc())
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
        finally:
            con.close()

    return JsonResponse({'status': 'error', 'message': 'Invalid request method'}, status=405)

@csrf_exempt
def run_processing_engine(request):
    import json
    import re
    import pandas as pd
    
    if request.method == 'POST':
        try:
            body = json.loads(request.body)
            ovatr_code = body.get('ovatr_code')
            if not ovatr_code:
                return JsonResponse({'status': 'error', 'message': 'Missing OVATR'}, status=400)

            conn = get_db_connection()
            
            new_cols = [
                ("matched_d_id", "VARCHAR"), ("sys_status", "VARCHAR"), 
                ("v_inv", "BOOLEAN"), ("v_tin", "BOOLEAN"), 
                ("v_date", "BOOLEAN"), ("v_diff", "DOUBLE")
            ]
            for col_name, col_type in new_cols:
                try: conn.execute(f"ALTER TABLE purchase ADD COLUMN {col_name} {col_type}")
                except: pass

            vatin_row = conn.execute("SELECT vatin FROM company_info WHERE ovatr = ?", [ovatr_code]).fetchone()
            user_vatin_clean = clean_invoice_text(vatin_row[0]).upper() if vatin_row else ""

            raw_decs = conn.execute("""
                SELECT id, date, invoice_number, tax_registration_id, vat_local_sale, vat_export
                FROM tax_declaration d
                JOIN company_info c ON regexp_replace(upper(CAST(d.tax_registration_id AS VARCHAR)), '[^A-Z0-9]', '', 'g') = regexp_replace(upper(CAST(c.vatin AS VARCHAR)), '[^A-Z0-9]', '', 'g')
                WHERE c.ovatr = ?
            """, [ovatr_code]).fetchall()
            
            dec_map = {clean_invoice_text(d[2]).upper(): d for d in raw_decs if clean_invoice_text(d[2])}

            purchases = conn.execute("SELECT no, invoice_no, date, purchase, \"import\" FROM purchase WHERE ovatr = ?", [ovatr_code]).fetchall()

            def check_date_match(v1, v2):
                try:
                    if pd.isna(v1) or pd.isna(v2) or str(v1).strip() == "" or str(v2).strip() == "": return False
                    dt1 = pd.to_datetime(v1, dayfirst=True, errors='coerce')
                    dt2 = pd.to_datetime(v2, dayfirst=True, errors='coerce')
                    if pd.isna(dt1) or pd.isna(dt2): return False
                    return dt1.month == dt2.month and dt1.year == dt2.year
                except: return False

            khmer_map = {
                'MATCHED': 'បានប្រកាស (អនុញ្ញាត)',
                'SHORTAGE': 'អនុញ្ញាត (អ្នកផ្គត់ផ្គង់ប្រកាសខ្វះ)',
                'NOT FOUND': 'ព្យួរទុក (មិនមានទិន្នន័យ)',
                'MISMATCH': 'ប្រកាសខុស (ព្យួរទុក)'
            }

            update_data = []
            for p in purchases:
                p_no = str(p[0])
                p_inv_clean = clean_invoice_text(p[1])
                p_amt = float(p[3] if p[3] else (p[4] if p[4] else 0.0))
                
                d_full = dec_map.get(p_inv_clean.upper() if p_inv_clean else "")
                
                if d_full:
                    d_id = str(d_full[0])
                    d_inv_clean = clean_invoice_text(d_full[2])
                    d_tin_clean = clean_invoice_text(d_full[3])
                    d_amt = float(d_full[4] if d_full[4] else (d_full[5] if d_full[5] else 0.0))
                    
                    v_inv = (p_inv_clean.upper() == d_inv_clean.upper()) and p_inv_clean != ""
                    v_tin = (d_tin_clean.upper() == user_vatin_clean)
                    v_date = check_date_match(p[2], d_full[1])
                    
                    v_diff = d_amt - p_amt 
                    
                    if v_inv and v_date and v_tin:
                        eng_status = 'SHORTAGE' if v_diff < -0.05 else 'MATCHED'
                    elif not v_inv and not v_date and not v_tin:
                        eng_status = 'NOT FOUND'
                    else:
                        eng_status = 'MISMATCH'
                        
                    sys_status = khmer_map[eng_status]
                else:
                    d_id = None
                    sys_status = khmer_map['NOT FOUND']
                    v_inv = v_tin = v_date = False
                    v_diff = 0.0 - p_amt 
                
                update_data.append([d_id, sys_status, v_inv, v_tin, v_date, v_diff, ovatr_code, p_no])

            conn.executemany("""
                UPDATE purchase 
                SET matched_d_id = ?, sys_status = ?, v_inv = ?, v_tin = ?, v_date = ?, v_diff = ?
                WHERE ovatr = ? AND CAST(no AS VARCHAR) = ?
            """, update_data)
            
            update_session_metadata(conn, ovatr_code, status="Completed")
            conn.close()

            return JsonResponse({'status': 'success', 'message': 'Processing complete. Results saved to database.'})
        except Exception as e:
            import traceback
            print(traceback.format_exc())
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    return JsonResponse({'status': 'error', 'message': 'Invalid Method'}, status=405)

def get_results_data(request):
    ovatr_code = request.GET.get('ovatr_code')
    table_type = request.GET.get('table_type', 'local')
    page = int(request.GET.get('page', 1))
    page_size = int(request.GET.get('page_size', 500))
    
    if not ovatr_code: 
        return JsonResponse({'status': 'error', 'message': 'Missing OVATR Code'}, status=400)

    try:
        conn = get_db_connection()
        
        hist = set()
        try:
            rows = conn.execute("SELECT DISTINCT row_no FROM change_history WHERE ovatr = ?", [ovatr_code]).fetchall()
            hist = {str(r[0]) for r in rows}
        except: pass

        amt_col = '"import"' if table_type == 'import' else 'purchase'

        # ADDED p.comment at index 8, shifted everything below by +1
        sql = f"""
            SELECT 
                p.no, p.description, p.supplier_name, p.supplier_tin, p.invoice_no, p.date, p.{amt_col}, p.user_status, p.comment,
                d.id, d.date, d.invoice_number, d.tax_registration_id, d.buyer_name, d.total_invoice_amount, 
                d.vat_local_sale, d.vat_export, 
                p.sys_status, p.v_inv, p.v_tin, p.v_date, p.v_diff,
                d.credit_notification_letter_number, d.buyer_type, d.amount_exclude_vat, d.non_vat_sales, 
                d.vat_zero_rate, d.vat_local_sale_state_burden, d.vat_withheld_by_national_treasury, 
                d.plt, d.special_tax_on_goods, d.special_tax_on_services, d.accommodation_tax, 
                d.income_tax_redemption_rate, d.notes, d.description as d_desc, d.tax_declaration_status
            FROM purchase p
            LEFT JOIN tax_declaration d ON p.matched_d_id = CAST(d.id AS VARCHAR)
            WHERE p.ovatr = ? AND p.{amt_col} > 0
            ORDER BY CAST(p.no AS INTEGER) ASC
        """
        
        db_rows = conn.execute(sql, [ovatr_code]).fetchall()
        conn.close()

        results = []
        stats = {'total': len(db_rows), 'matched': 0, 'not_found': 0, 'mismatch': 0, 'eff_counts': {}}
        
        def cl_dt(v):
            if pd.isna(v) or str(v).strip() == "" or v is None: return ""
            try: return pd.to_datetime(v, dayfirst=True).strftime('%d-%m-%Y')
            except: return str(v).split(' ')[0]

        khmer_matched = 'បានប្រកាស (អនុញ្ញាត)'
        khmer_shortage = 'អនុញ្ញាត (អ្នកផ្គត់ផ្គង់ប្រកាសខ្វះ)'
        khmer_not_found = 'ព្យួរទុក (មិនមានទិន្នន័យ)'

        for r in db_rows:
            # Shifted indices: sys_status is now 17
            sys_status = str(r[17]) if r[17] else khmer_not_found
            u_status = str(r[7]).strip() if r[7] and str(r[7]).strip().lower() not in ['none', 'null', 'nan', ''] else ""
            
            if sys_status in [khmer_matched, khmer_shortage]: stats['matched'] += 1
            elif sys_status == khmer_not_found: stats['not_found'] += 1
            else: stats['mismatch'] += 1

            eff_status = u_status if u_status else sys_status
            stats['eff_counts'][eff_status] = stats['eff_counts'].get(eff_status, 0) + 1

            d_data = {}
            if r[9]:
                d_data = {
                    'id': str(r[9]), 'date': cl_dt(r[10]), 'invoice_no': str(r[11]) if r[11] else "", 'credit_no': r[22],
                    'buyer_type': r[23], 'tin': str(r[12]) if r[12] else "", 'name': str(r[13]) if r[13] else "",
                    'total_amt': r[14], 'excl_vat': r[24], 'non_vat': r[25],
                    'vat_0': r[26], 'vat_local': r[15], 'vat_export': r[16],
                    'state_burden': r[27], 'withheld': r[28], 'plt': r[29],
                    'spec_goods': r[30], 'spec_serv': r[31], 'accom': r[32],
                    'redemption': r[33], 'notes': r[34], 'desc': r[35],
                    'dec_status': r[36]
                }

            results.append({
                'no': str(r[0]), 
                'has_history': str(r[0]) in hist, 
                'status': sys_status,
                'user_status': u_status,
                'p_comment': str(r[8]) if r[8] else "", # NEW COMMENT FIELD
                'p_inv_clean': clean_invoice_text(r[4]), 
                'd_inv_clean': clean_invoice_text(r[11]) if r[11] else "",
                'v_inv': bool(r[18]),
                'v_tin': bool(r[19]),
                'v_date': bool(r[20]),
                'v_diff': float(r[21]) if r[21] else 0.0,
                'p_desc': str(r[1]) if r[1] else "", 
                'p_supp': str(r[2]) if r[2] else "", 
                'p_tin': str(r[3]) if r[3] else "", 
                'p_inv': str(r[4]) if r[4] else "", 
                'p_date': cl_dt(r[5]), 
                'p_amt': float(r[6]) if r[6] else 0.0,
                'd_inv': str(r[11]) if r[11] else "", 
                'd_tin': d_data.get('tin', ''),
                'd_date': d_data.get('date', ''), 
                'd_name': d_data.get('name', ''),
                'd_amt': float(r[15] if table_type == 'local' else r[16]) if r[9] else 0.0, 
                'd_data': d_data
            })

        total_pages = (stats['total'] + page_size - 1) // page_size if page_size > 0 else 1
        start = (page - 1) * page_size
        end = start + page_size
        
        return JsonResponse({
            'status': 'success', 
            'data': results[start:end], 
            'stats': stats, 
            'pagination': {
                'current_page': page, 'total_pages': total_pages,
                'page_size': page_size, 'total_rows': stats['total']
            }
        })
    except Exception as e: 
        import traceback
        print(traceback.format_exc())
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

def generate_annex_iii(request):
    ovatr_code = request.GET.get('ovatr_code') or request.session.get('ovatr_code')
    if not ovatr_code: 
        return JsonResponse({'status': 'error', 'message': 'Session ID missing'}, status=400)
        
    return JsonResponse({'status': 'success', 'redirect_url': f"/crosscheck/results/?ovatr_code={ovatr_code}"})

def download_report(request):
    import math
    import re
    
    ovatr_code = request.GET.get('ovatr_code')
    if not ovatr_code:
        return JsonResponse({'status': 'error', 'message': 'Missing Session ID'}, status=400)
    
    conn = None
    try:
        template_path = os.path.join(settings.BASE_DIR, 'core', 'templates', 'static', 'CC - guide.xlsx')
        if not os.path.exists(template_path): 
            template_path = os.path.join(settings.BASE_DIR, 'core', 'static', 'CC - guide.xlsx')
        
        conn = get_db_connection()

        vatin_row = conn.execute("SELECT vatin FROM company_info WHERE ovatr = ?", [ovatr_code]).fetchone()
        user_vatin = vatin_row[0] if vatin_row else ""
        user_vatin_safe = user_vatin.replace('"', '""')

        # NEW COMMENT at index 8
        local_purchases = conn.execute("""
            SELECT description, supplier_name, supplier_tin, invoice_no, date, purchase, no, user_status, comment 
            FROM purchase WHERE ovatr = ? AND purchase > 0 ORDER BY CAST(no AS INTEGER) ASC
        """, [ovatr_code]).fetchall()

        import_purchases = conn.execute("""
            SELECT description, supplier_name, supplier_tin, invoice_no, date, "import", no, user_status, comment
            FROM purchase WHERE ovatr = ? AND "import" > 0 ORDER BY CAST(no AS INTEGER) ASC
        """, [ovatr_code]).fetchall()

        raw_decs = conn.execute("""
            SELECT 
                d.date, d.invoice_number, d.credit_notification_letter_number, d.buyer_type, 
                d.tax_registration_id, d.buyer_name, d.total_invoice_amount, d.amount_exclude_vat, 
                d.non_vat_sales, d.vat_zero_rate, d.vat_local_sale, d.vat_export, 
                d.vat_local_sale_state_burden, d.vat_withheld_by_national_treasury, d.plt, 
                d.special_tax_on_goods, d.special_tax_on_services, d.accommodation_tax, 
                d.income_tax_redemption_rate, d.notes, d.description, d.tax_declaration_status, 
                p.invoice_no
            FROM tax_declaration d
            JOIN purchase p ON 
                regexp_replace(upper(d.invoice_number), '[^A-Z0-9]', '', 'g') = regexp_replace(upper(p.invoice_no), '[^A-Z0-9]', '', 'g')
            JOIN company_info c ON p.ovatr = c.ovatr
            WHERE p.ovatr = ?
            AND regexp_replace(upper(d.tax_registration_id), '[^A-Z0-9]', '', 'g') = regexp_replace(upper(c.vatin), '[^A-Z0-9]', '', 'g')
            AND month(d.date) = month(COALESCE(try_cast(p.date as DATE), strptime(p.date, '%d-%m-%Y')))
            AND year(d.date) = year(COALESCE(try_cast(p.date as DATE), strptime(p.date, '%d-%m-%Y')))
        """, [ovatr_code]).fetchall()
        
        dec_map = {}
        for dec in raw_decs:
            p_inv_key = clean_invoice_text(dec[22])  
            if p_inv_key: dec_map[p_inv_key] = dec

        with warnings.catch_warnings():
            warnings.filterwarnings("ignore", category=UserWarning)
            wb = load_workbook(template_path)
        
        if 'AnnexIII-Import' not in wb.sheetnames and 'AnnexIII-Local Pur' in wb.sheetnames:
            target = wb.copy_worksheet(wb['AnnexIII-Local Pur'])
            target.title = 'AnnexIII-Import'

        def clean_num(val):
            if val is None or val == "": return 0.0
            try:
                f = float(val)
                if math.isnan(f) or math.isinf(f): return 0.0
                return f
            except: return 0.0

        align_center = Alignment(horizontal='center', vertical='center', wrap_text=False)

        def clean_text(val):
            if pd.isna(val) or val is None: return ""
            s = str(val).strip()
            if s.lower() in ['nan', 'none', 'null']: return ""
            return re.sub(r'[\x00-\x08\x0b-\x0c\x0e-\x1f]', '', s)

        def process_sheet(sheet_name, data_rows):
            if sheet_name not in wb.sheetnames: return
            ws = wb[sheet_name]
            
            if hasattr(ws, 'data_validations') and ws.data_validations.dataValidation:
                ws.data_validations.dataValidation = []
            
            start_row = 8
            for r in range(1, 15):
                if ws.cell(row=r, column=1).value and "ល.រ" in str(ws.cell(row=r, column=1).value):
                    start_row = r + 1; break
            
            if ws.max_row >= start_row:
                 ws.delete_rows(start_row, ws.max_row - start_row + 1)
            
            for i, p_row in enumerate(data_rows):
                r = start_row + i
                
                p_inv_val = clean_text(p_row[3])
                p_inv_clean = clean_invoice_text(p_inv_val)
                
                d_row = dec_map.get(p_inv_clean)
                d_inv_val = clean_text(d_row[1] if d_row else "")
                d_inv_clean = clean_invoice_text(d_inv_val)

                v_inv = (p_inv_clean == d_inv_clean) if (p_inv_clean and d_inv_clean) else False
                v_date = True if d_row else False 
                v_tin = (clean_invoice_text(p_row[2]) == clean_invoice_text(d_row[4])) if (d_row and d_row[4] and p_row[2]) else False
                
                i_val = clean_num(p_row[5])
                ag_val = clean_num(d_row[10] if d_row else 0)
                u_val = ag_val - i_val

                if v_inv and v_date and v_tin:
                    j_status = "អនុញ្ញាត (អ្នកផ្គត់ផ្គង់ប្រកាសខ្វះ)" if u_val < -0.05 else "បានប្រកាស (អនុញ្ញាត)"
                elif not v_inv and not v_date and not v_tin:
                    j_status = "ព្យួរទុក (មិនមានទិន្នន័យ)"
                else:
                    j_status = "ប្រកាសខុស (ព្យួរទុក)"

                ws.cell(row=r, column=1, value=clean_text(p_row[6])) 
                ws.cell(row=r, column=2, value=clean_text(p_row[0]))
                ws.cell(row=r, column=3, value=clean_text(p_row[1]))
                ws.cell(row=r, column=4, value=clean_text(p_row[2]))
                ws.cell(row=r, column=5, value=p_inv_val)
                # ws.cell(row=r, column=6, value=to_excel_date(p_row[4]))
                ws.cell(row=r, column=9, value=i_val)
                
                # Formula updated with shifted validation cells (Q, R, S and W Diff)
                status_formula = f'=IF(AND(Q{r}=TRUE, R{r}=TRUE, S{r}=TRUE), IF(W{r}<-0.05, "អនុញ្ញាត (អ្នកផ្គត់ផ្គង់ប្រកាសខ្វះ)", "បានប្រកាស (អនុញ្ញាត)"), IF(AND(Q{r}=FALSE, R{r}=FALSE, S{r}=FALSE), "ព្យួរទុក (មិនមានទិន្នន័យ)", "ប្រកាសខុស (ព្យួរទុក)"))'
                ws.cell(row=r, column=10, value=status_formula)
                ws.cell(row=r, column=11, value=clean_text(p_row[7]))
                
                # --- NEW COMMENT (L) and EVERYTHING SHIFTED +1 ---
                ws.cell(row=r, column=12, value=clean_text(p_row[8]))
                ws.cell(row=r, column=13, value=f"=AH{r}") # M (was 12: AG)
                ws.cell(row=r, column=14, value=f"=IF(W{r}<0,AH{r},I{r})") # N (was 13: V<0,AG,I) -> Wait, Diff is W(23) now
                ws.cell(row=r, column=15, value=f"=I{r}-M{r}") # O (was 14: I-L)
                ws.cell(row=r, column=16, value=None) # P (was 15: None)
                ws.cell(row=r, column=17, value=p_inv_clean) # Q (was 16: p_inv_clean)
                ws.cell(row=r, column=18, value=d_inv_clean) # R (was 17: d_inv_clean)
                ws.cell(row=r, column=19, value=f"=Q{r}=R{r}") # S (was 18: P=Q)
                ws.cell(row=r, column=20, value=f"=AND(MONTH(F{r})=MONTH(X{r}), YEAR(F{r})=YEAR(X{r}))") # T (was 19: ...W...)
                ws.cell(row=r, column=21, value=f'=AC{r}="{user_vatin_safe}"') # U (was 20: AA=user_vatin)
                ws.cell(row=r, column=22, value=f"=AH{r}-I{r}") # V (was 21: AG-I) => Wait, Diff is usually this one? 
                ws.cell(row=r, column=23, value=None)
                
                # w_date_val = str(d_row[0]).split()[0] if d_row and d_row[0] else ""
                # ws.cell(row=r, column=24, value=to_excel_date(w_date_val))  
                ws.cell(row=r, column=25, value=d_inv_val)  
                ws.cell(row=r, column=26, value=clean_text(d_row[2] if d_row else ""))  
                ws.cell(row=r, column=27, value=clean_text(d_row[3] if d_row else ""))  
                ws.cell(row=r, column=28, value=clean_text(d_row[4] if d_row else ""))  
                ws.cell(row=r, column=29, value=clean_text(d_row[5] if d_row else ""))  
                
                ws.cell(row=r, column=30, value=clean_num(d_row[6] if d_row else 0)) 
                ws.cell(row=r, column=31, value=clean_num(d_row[7] if d_row else 0)) 
                ws.cell(row=r, column=32, value=clean_num(d_row[8] if d_row else 0)) 
                ws.cell(row=r, column=33, value=clean_num(d_row[9] if d_row else 0)) 
                ws.cell(row=r, column=34, value=ag_val) # AH
                ws.cell(row=r, column=35, value=clean_num(d_row[11] if d_row else 0)) 
                ws.cell(row=r, column=36, value=clean_num(d_row[12] if d_row else 0)) 
                ws.cell(row=r, column=37, value=clean_num(d_row[13] if d_row else 0)) 
                ws.cell(row=r, column=38, value=clean_num(d_row[14] if d_row else 0)) 
                ws.cell(row=r, column=39, value=clean_num(d_row[15] if d_row else 0)) 
                ws.cell(row=r, column=40, value=clean_num(d_row[16] if d_row else 0)) 
                ws.cell(row=r, column=41, value=clean_num(d_row[17] if d_row else 0)) 
                ws.cell(row=r, column=42, value=clean_num(d_row[18] if d_row else 0)) 
                
                ws.cell(row=r, column=43, value=clean_text(d_row[19] if d_row else "")) 
                ws.cell(row=r, column=44, value=clean_text(d_row[20] if d_row else "")) 
                ws.cell(row=r, column=45, value=clean_text(d_row[21] if d_row else "")) 

                raw_date = p_row[4]
                dt_val = ""
                if raw_date and str(raw_date).lower() not in ['nan', 'nat', 'none', '']:
                    try:
                        # Extract exact Python Date object for true Excel sorting
                        dt_val = pd.to_datetime(raw_date).date()
                    except:
                        dt_val = str(raw_date).split()[0]
                
                dt_cell = ws.cell(row=r, column=6, value=dt_val)
                dt_cell.alignment = align_center
                dt_cell.number_format = 'DD-MM-YYYY'

                raw_d_date = d_row[0] if d_row else ""
                dt_d_val = ""
                if raw_d_date and str(raw_d_date).lower() not in ['nan', 'nat', 'none', '']:
                    try:
                        # Extract exact Python Date object for true Excel sorting
                        dt_d_val = pd.to_datetime(raw_d_date).date()
                    except:
                        dt_d_val = str(raw_d_date).split()[0]
                
                dt_d_cell = ws.cell(row=r, column=24, value=dt_d_val)
                dt_d_cell.alignment = align_center
                dt_d_cell.number_format = 'DD-MM-YYYY'

                format_cols = [9, 13, 14, 15, 23] + list(range(30, 43))
                for col_idx in format_cols:
                    ws.cell(row=r, column=col_idx).number_format = '#,###0'

        process_sheet('AnnexIII-Local Pur', local_purchases)
        process_sheet('AnnexIII-Import', import_purchases)

        save_dir = os.path.join(settings.MEDIA_ROOT, 'temp_reports')
        os.makedirs(save_dir, exist_ok=True)
        filename = f"AnnexIII_{ovatr_code}.xlsx"
        file_path = os.path.join(save_dir, filename)
        wb.save(file_path)

        return FileResponse(open(file_path, 'rb'), as_attachment=True, filename=filename)

    except Exception as e: 
        import traceback
        print(traceback.format_exc())
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    finally:
        if conn: 
            conn.close()

def get_report_data(request):
    try:
        ovatr = request.GET.get('ovatr_code')
        sheet = request.GET.get('sheet')
        
        if not ovatr or not sheet:
            return JsonResponse({'status': 'error', 'message': 'Missing parameters'})
            
        con = get_db_connection()
        
        # --- NEW: Harvest Global + Session Custom Statuses safely ---
        status_list = []
        try:
            rows = con.execute("SELECT name FROM user_status_config").fetchall()
            existing_names = [r[0] for r in rows]
            status_list = [{'name': name} for name in existing_names]
            
            if ovatr:
                p_stat = con.execute("SELECT DISTINCT user_status FROM purchase WHERE ovatr = ? AND user_status IS NOT NULL AND user_status != ''", [ovatr]).fetchall()
                for r in p_stat:
                    st = str(r[0]).strip()
                    if st and st.lower() not in ['none', 'null', 'nan'] and st not in existing_names:
                        existing_names.append(st)
                        status_list.append({'name': st})
        except Exception as e:
            print("Status fetch error:", e)
        # --------------------------------------------------------------
        
        data = []
        columns = []
        
        if sheet == 'company':
            row = con.execute("SELECT * FROM company_info WHERE ovatr = ?", [ovatr]).fetchone()
            if row:
                cols = [desc[0] for desc in con.description]
                for i, col_name in enumerate(cols):
                    if col_name == 'ovatr': continue
                    data.append({'key': col_name, 'value': row[i]})
                columns = [{'key': 'key', 'label': 'Field'}, {'key': 'value', 'label': 'Value'}]
                
        elif sheet == 'annex_1': 
            res = con.execute("SELECT no, description, invoice_no, supplier_name, supplier_tin, date, import_state_charge, user_status, sys_status FROM purchase WHERE ovatr = ? AND import_state_charge <> 0 ORDER BY CAST(no AS INTEGER)", [ovatr])
            cols = [desc[0] for desc in con.description]
            data = [dict(zip(cols, r)) for r in res.fetchall()]
            columns = [{'key': c, 'label': c.replace('_', ' ').title()} for c in cols]
            
        elif sheet == 'annex_2': 
            try: con.execute("ALTER TABLE purchase ADD COLUMN approve_amount DOUBLE DEFAULT 0.0")
            except: pass
            try: con.execute("ALTER TABLE purchase ADD COLUMN annex2_note VARCHAR DEFAULT ''")
            except: pass

            res = con.execute("""
                SELECT no, description, invoice_no, supplier_name, supplier_tin, date, 
                       import, approve_amount, (import - COALESCE(approve_amount, 0)) AS shortfall, 
                       annex2_note, user_status, sys_status 
                FROM purchase 
                WHERE ovatr = ? AND import <> 0 
                ORDER BY CAST(no AS INTEGER)
            """, [ovatr])
            
            cols = [desc[0] for desc in con.description]
            data = [dict(zip(cols, r)) for r in res.fetchall()]
            columns = [{'key': c, 'label': c.replace('_', ' ').title()} for c in cols]
            
        elif sheet == 'annex_3': 
            res = con.execute("SELECT no, description, date, invoice_no, supplier_name, supplier_tin, purchase as amount, user_status, sys_status FROM purchase WHERE ovatr = ? AND purchase > 0 ORDER BY CAST(no AS INTEGER)", [ovatr])
            cols = [desc[0] for desc in con.description]
            data = [dict(zip(cols, r)) for r in res.fetchall()]
            columns = [{'key': c, 'label': c.replace('_', ' ').title()} for c in cols]
            
        elif sheet == 'annex_4': 
            res = con.execute("SELECT no, description, invoice_no, buyer_name, tax_registration_id, date, vat_export FROM sale WHERE ovatr = ? AND vat_export <> 0 ORDER BY CAST(no AS INTEGER)", [ovatr])
            cols = [desc[0] for desc in con.description]
            data = [dict(zip(cols, r)) for r in res.fetchall()]
            columns = [{'key': c, 'label': c.replace('_', ' ').title()} for c in cols]
            
        elif sheet == 'annex_5': 
            res = con.execute("SELECT no, description, invoice_no, buyer_name, tax_registration_id, date, vat_local_sale FROM sale WHERE ovatr = ? AND vat_local_sale <> 0 ORDER BY CAST(no AS INTEGER)", [ovatr])
            cols = [desc[0] for desc in con.description]
            data = [dict(zip(cols, r)) for r in res.fetchall()]
            columns = [{'key': c, 'label': c.replace('_', ' ').title()} for c in cols]
            
        elif sheet == 'taxpaid':
            company_info = con.execute("SELECT i_request_date FROM company_info WHERE ovatr = ?", [ovatr]).fetchone()
            
            query = "SELECT * FROM tax_paid WHERE ovatr = ?"
            params = [ovatr]
            
            import re
            khmer_months = {'មករា': 1, 'កុម្ភៈ': 2, 'មីនា': 3, 'មេសា': 4, 'ឧសភា': 5, 'មិថុនា': 6, 'កក្កដា': 7, 'សីហា': 8, 'កញ្ញា': 9, 'តុលា': 10, 'វិច្ឆិកា': 11, 'ធ្នូ': 12}
            month_cols = {1: 'jan', 2: 'feb', 3: 'mar', 4: 'apr', 5: 'may', 6: 'jun', 7: 'jul', 8: 'aug', 9: 'sep', 10: 'oct', 11: 'nov', 12: 'dec'}
            
            start_m, start_y, end_m, end_y = None, None, None, None
            
            if company_info and company_info[0]:
                req_date_str = str(company_info[0]).strip()
                years_found = re.findall(r'\b(20\d{2})\b', req_date_str)
                months_found = []
                for k_month in khmer_months.keys():
                    occurrences = [m.start() for m in re.finditer(k_month, req_date_str)]
                    for idx in occurrences:
                        months_found.append((idx, khmer_months[k_month]))
                months_found.sort(key=lambda x: x[0])
                
                if len(years_found) >= 1:
                    start_y = int(years_found[0])
                    end_y = int(years_found[-1])
                if len(months_found) >= 1:
                    start_m = months_found[0][1]
                    end_m = months_found[-1][1]
                    
                if start_y and end_y:
                    query += " AND CAST(tax_year AS INTEGER) >= ? AND CAST(tax_year AS INTEGER) <= ?"
                    params.extend([start_y, end_y])
                        
            res = con.execute(query, params)
            cols = [desc[0] for desc in con.description]
            raw_data = [dict(zip(cols, r)) for r in res.fetchall()]
            
            for row in raw_data:
                try: t_year = int(row.get('tax_year', 0))
                except: t_year = 0
                    
                if start_y and end_y and t_year:
                    new_total = 0
                    for m_num, m_col in month_cols.items():
                        val = row.get(m_col)
                        val = float(val) if val is not None and str(val).strip() != '' else 0.0
                        
                        is_before_start = (t_year == start_y and start_m is not None and m_num < start_m)
                        is_after_end = (t_year == end_y and end_m is not None and m_num > end_m)
                        
                        if is_before_start or is_after_end:
                            row[m_col] = 0 
                        else:
                            new_total += val 
                            
                    row['total'] = new_total
                data.append(row)
                
            columns = [{'key': c, 'label': c.replace('_', ' ').title()} for c in cols]

        con.close()
        
        # INJECT STATUSES DIRECTLY IN THE JSON SO FRONTEND DROPDOWN NEVER FAILS
        return JsonResponse({'status': 'success', 'data': data, 'columns': columns, 'statuses': status_list})
        
    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

@csrf_exempt
def api_user_statuses(request):
    con = get_db_connection()
    con.execute("CREATE TABLE IF NOT EXISTS user_status_config (name VARCHAR PRIMARY KEY, summary VARCHAR, action VARCHAR)")
    try: con.execute("ALTER TABLE user_status_config ADD COLUMN color VARCHAR")
    except: pass
    
    try:
        con.execute("UPDATE user_status_config SET color = 'red' WHERE name = 'ព្យួរទុក (មិនមានទិន្នន័យ)' AND (color IS NULL OR color = 'gray')")
        con.execute("UPDATE user_status_config SET color = 'green' WHERE name = 'បានប្រកាស (អនុញ្ញាត)' AND (color IS NULL OR color = 'gray')")
        con.execute("UPDATE user_status_config SET color = 'orange' WHERE name = 'ប្រកាសខុស (ព្យួរទុក)' AND (color IS NULL OR color = 'gray')")
        con.execute("UPDATE user_status_config SET color = 'blue' WHERE name = 'អនុញ្ញាត (អ្នកផ្គត់ផ្គង់ប្រកាសខ្វះ)' AND (color IS NULL OR color = 'gray')")
        con.execute("UPDATE user_status_config SET color = 'orange' WHERE name = 'ព្យួរទុក (មិនមានឯកសារគាំទ្រ)' AND (color IS NULL OR color = 'gray')")
        con.execute("UPDATE user_status_config SET color = 'orange' WHERE name = 'ព្យួរទុក (ខុសវិធានវិក្កយបត្រអាករ)' AND (color IS NULL OR color = 'gray')")
        con.commit()
    except: pass

    if con.execute("SELECT COUNT(*) FROM user_status_config").fetchone()[0] == 0:
        con.executemany("INSERT INTO user_status_config (name, summary, action, color) VALUES (?, ?, ?, ?)", [
            ('ព្យួរទុក (មិនមានទិន្នន័យ)', 'ចំនួនប្រាក់អាករដែលមិនមានទិន្នន័យ', 'ព្យួរទុក', 'red'),
            ('បានប្រកាស (អនុញ្ញាត)', 'ចំនួនប្រាក់អាករដែលបានប្រកាស', 'គួរអនុញ្ញាត', 'green'),
            ('ប្រកាសខុស (ព្យួរទុក)', 'ចំនួនប្រាក់អាករដែលប្រកាសខុស', 'ព្យួរទុក', 'orange'),
            ('អនុញ្ញាត (អ្នកផ្គត់ផ្គង់ប្រកាសខ្វះ)', 'ចំនួនប្រាក់អាករដែលអ្នកផ្គត់ផ្គង់ប្រកាសខ្វះ', 'គួរអនុញ្ញាត', 'blue'),
            ('ព្យួរទុក (មិនមានឯកសារគាំទ្រ)', 'ចំនួនប្រាក់អាករដែលមិនមានឯកសារគាំទ្រ', 'ព្យួរទុក', 'orange'),
            ('ព្យួរទុក (ខុសវិធានវិក្កយបត្រអាករ)', 'ចំនួនប្រាក់អាករដែលខុសវិធានវិក្កយបត្រ', 'ព្យួរទុក', 'orange')
        ])
        con.commit()

    if request.method == 'GET':
        ovatr = request.GET.get('ovatr')
        rows = con.execute("SELECT name, summary, action, color FROM user_status_config").fetchall()
        data = [{'name': r[0], 'summary': r[1], 'action': r[2], 'color': r[3] if r[3] else 'gray'} for r in rows]
        
        if ovatr:
            try:
                p_stat = con.execute("SELECT DISTINCT user_status FROM purchase WHERE ovatr = ? AND user_status IS NOT NULL AND user_status != ''", [ovatr]).fetchall()
                existing_names = [d['name'] for d in data]
                for r in p_stat:
                    stat_name = str(r[0]).strip()
                    if stat_name and stat_name.lower() not in ['none', 'null', 'nan'] and stat_name not in existing_names:
                        data.append({'name': stat_name, 'summary': 'Custom', 'action': 'ព្យួរទុក', 'color': 'gray'})
                        existing_names.append(stat_name)
            except: pass
        con.close()
        return JsonResponse({'status': 'success', 'data': data})
    elif request.method == 'POST':
        try:
            body = json.loads(request.body)
            if body.get('type') == 'add':
                con.execute("INSERT OR REPLACE INTO user_status_config (name, summary, action, color) VALUES (?, ?, ?, ?)", [body.get('name'), body.get('summary'), body.get('action'), body.get('color', 'gray')])
            elif body.get('type') == 'delete':
                con.execute("DELETE FROM user_status_config WHERE name = ?", [body.get('name')])
            con.commit()
            con.close()
            return JsonResponse({'status': 'success'})
        except Exception as e:
            con.close()
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    con.close()
    return JsonResponse({'status': 'error', 'message': 'Invalid Method'}, status=405)

@csrf_exempt
def update_report_cell(request):
    if request.method == 'POST':
        try:
            body = json.loads(request.body)
            ovatr = body.get('ovatr')
            sheet = body.get('sheet')
            id_val = body.get('id_val')
            field = body.get('field')
            value = body.get('value')
            old_value = body.get('old_value')
            
            con = get_db_connection()
            
            table_map = {
                'company': {'table': 'company_info', 'pk': 'ovatr'},
                'annex_1': {'table': 'purchase', 'pk': 'no'},
                'annex_2': {'table': 'purchase', 'pk': 'no'},
                'annex_3': {'table': 'purchase', 'pk': 'no'},
                'annex_4': {'table': 'sale', 'pk': 'no'},
                'annex_5': {'table': 'sale', 'pk': 'no'},
                'taxpaid': {'table': 'tax_paid', 'pk': 'description'} 
            }
            
            config = table_map.get(sheet)
            if not config:
                return JsonResponse({'status': 'error', 'message': 'Invalid sheet'})
                
            table = config['table']
            pk_col = config['pk']
            
            # --- MODIFIED: Ensure approve_amount is stripped of formatting before DB save ---
            if field == 'approve_amount':
                value = clean_currency(value)
            # --------------------------------------------------------------------------------
            
            if sheet == 'company':
                db_field = id_val 
                query = f'UPDATE {table} SET "{db_field}" = ? WHERE ovatr = ?'
                params = [value, ovatr]
            elif sheet == 'taxpaid':
                query = f'UPDATE {table} SET "{field}" = ? WHERE ovatr = ? AND description = ?'
                params = [value, ovatr, id_val]
            else:
                query = f'UPDATE {table} SET "{field}" = ? WHERE ovatr = ? AND "{pk_col}" = ?'
                params = [value, ovatr, id_val]

            con.execute(query, params)
            
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            con.execute("INSERT INTO change_history VALUES (?, ?, ?, ?, ?, ?, ?)", 
                        [timestamp, ovatr, str(id_val), table, field, str(old_value), str(value)])
            
            update_session_metadata(con, ovatr)
            con.close()
            
            return JsonResponse({'status': 'success'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
        
def download_full_report(request):
    ovatr_code = request.GET.get('ovatr_code')
    if not ovatr_code: return JsonResponse({'status': 'error', 'message': 'Missing Session ID'}, status=400)
    
    con = get_db_connection()
    try:
        try: con.execute("ALTER TABLE purchase ADD COLUMN user_status VARCHAR")
        except: pass
        try: con.execute("ALTER TABLE purchase ADD COLUMN comment VARCHAR DEFAULT ''")
        except: pass
        try: con.execute("ALTER TABLE purchase ADD COLUMN approve_amount DOUBLE DEFAULT 0.0")
        except: pass
        try: con.execute("ALTER TABLE purchase ADD COLUMN annex2_note VARCHAR DEFAULT ''")
        except: pass
            
        row = con.execute("SELECT * FROM company_info WHERE ovatr = ?", [ovatr_code]).fetchone()
        if not row: return JsonResponse({'status': 'error', 'message': 'Company info not found'}, status=404)
        
        cols = [desc[0] for desc in con.description]
        company_data = dict(zip(cols, row))
        user_vatin_safe = str(company_data.get('vatin', '')).replace('"', '""')

        status_configs = []
        try: status_configs = con.execute("SELECT name, summary, action FROM user_status_config").fetchall()
        except: pass

        annex_i_rows = con.execute("SELECT description, invoice_no, date, import_state_charge FROM purchase WHERE ovatr = ? AND import_state_charge <> 0 ORDER BY CAST(no AS INTEGER) ASC", [ovatr_code]).fetchall()
        
        # MODIFIED: Fetch the new custom fields for Annex II
        annex_ii_rows = con.execute("SELECT description, supplier_name, invoice_no, date, \"import\", approve_amount, annex2_note FROM purchase WHERE ovatr = ? AND \"import\" <> 0 ORDER BY CAST(no AS INTEGER) ASC", [ovatr_code]).fetchall()
        
        annex_iii_local_purchases = con.execute("SELECT description, supplier_name, supplier_tin, invoice_no, date, purchase, status, user_status, comment FROM purchase WHERE ovatr = ? AND purchase > 0 ORDER BY CAST(no AS INTEGER) ASC", [ovatr_code]).fetchall()
        
        annex_iii_raw_decs = con.execute("""
            SELECT 
                d.date, d.invoice_number, d.credit_notification_letter_number, d.buyer_type, 
                d.tax_registration_id, d.buyer_name, d.total_invoice_amount, d.amount_exclude_vat, 
                d.non_vat_sales, d.vat_zero_rate, d.vat_local_sale, d.vat_export, 
                d.vat_local_sale_state_burden, d.vat_withheld_by_national_treasury, d.plt, 
                d.special_tax_on_goods, d.special_tax_on_services, d.accommodation_tax, 
                d.income_tax_redemption_rate, d.notes, d.description, d.tax_declaration_status,
                p.invoice_no
            FROM purchase p
            LEFT JOIN tax_declaration d ON regexp_replace(upper(d.invoice_number), '[^A-Z0-9]', '', 'g') = regexp_replace(upper(p.invoice_no), '[^A-Z0-9]', '', 'g')
            WHERE p.ovatr = ?
        """, [ovatr_code]).fetchall()
        
        def clean_invoice_text(val):
            if pd.isna(val) or val is None: return ""
            s = str(val).strip()
            if s.lower() in ['nan', 'none', 'null']: return ""
            return re.sub(r'[\x00-\x08\x0b-\x0c\x0e-\x1f]', '', s)
            
        def clean_currency(val):
            try: return float(str(val).replace(',', ''))
            except: return 0.0

        def get_last_9_digits(val):
            if pd.isna(val) or val is None: return ""
            digits = re.sub(r'\D', '', str(val))
            return digits[-9:] if len(digits) >= 9 else digits

        dec_map = {clean_invoice_text(d[22]): d for d in annex_iii_raw_decs if clean_invoice_text(d[22]) and d[1]}

        rc_rows = con.execute("SELECT description, invoice_no, date, vat FROM reverse_charge WHERE ovatr = ? ORDER BY CAST(no AS INTEGER) ASC", [ovatr_code]).fetchall()
        annex_iv_rows = con.execute("SELECT description, invoice_no, date, vat_export FROM sale WHERE ovatr = ? AND vat_export <> 0 ORDER BY CAST(no AS INTEGER) ASC", [ovatr_code]).fetchall()
        annex_v_rows = con.execute("SELECT description, invoice_no, date, vat_local_sale FROM sale WHERE ovatr = ? AND vat_local_sale <> 0 ORDER BY CAST(no AS INTEGER) ASC", [ovatr_code]).fetchall()
        taxpaid_raw = con.execute("SELECT * FROM tax_paid WHERE ovatr = ? ORDER BY tax_year ASC", [ovatr_code]).fetchall()
        tp_cols = [desc[0] for desc in con.description]

        template_path = os.path.join(settings.BASE_DIR, 'templates', 'Sample-Excel_Report.xlsx')
        if not os.path.exists(template_path): template_path = os.path.join(settings.MEDIA_ROOT, 'templates', 'Sample-Excel_Report.xlsx')
        
        wb = load_workbook(template_path)
        khmer_font = Font(name='Khmer OS Siemreap', size=11)
        khmer_font_bold = Font(name='Khmer OS Siemreap', size=11, bold=True)
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        align_middle = Alignment(vertical='center', wrap_text=False)
        align_center = Alignment(horizontal='center', vertical='center', wrap_text=False)
        align_left_middle = Alignment(horizontal='left', vertical='center', wrap_text=False)
        align_right_middle = Alignment(horizontal='right', vertical='center', wrap_text=False)
        bg_gray_header = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        bg_gray_summary = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        bg_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        def to_excel_date(date_val):
            if not date_val: return None
            for fmt in ('%d-%m-%Y', '%Y-%m-%d', '%d/%m/%Y', '%Y/%m/%d'):
                try: return datetime.strptime(str(date_val).strip(), fmt)
                except: continue
            return date_val

        def to_khmer_numeral(text):
            if text is None or text == "": return ""
            khmer_digits = "០១២៣៤៥៦៧៨៩"
            return "".join(khmer_digits[int(c)] if c.isdigit() else c for c in str(text))
            
        def clean_text(val):
            if pd.isna(val) or val is None: return ""
            s = str(val).strip()
            if s.lower() in ['nan', 'none', 'null']: return ""
            return re.sub(r'[\x00-\x08\x0b-\x0c\x0e-\x1f]', '', s)

        ws_info = next((wb[n] for n in wb.sheetnames if n.strip().lower() == 'company information'), None)
        if ws_info:
            business_activity_str = ""
            try:
                activities = json.loads(company_data.get('business_activities', '[]'))
                if isinstance(activities, list) and len(activities) > 0:
                    business_activity_str = f"{activities[0].get('name', '')} {activities[0].get('desc', '')}".strip()
            except Exception: pass

            bank_acc_num = bank_name = ""
            try:
                accounts = json.loads(company_data.get('enterprise_accounts', '[]'))
                if isinstance(accounts, list) and len(accounts) > 0:
                    bank_acc_num = str(accounts[0].get('number', ''))
                    bank_name = str(accounts[0].get('bank', ''))
            except Exception: pass

            company_mappings = [
                ('D2', company_data.get('company_name_kh', ''), 'text'), ('D3', company_data.get('company_name_en', ''), 'text'),
                ('D4', company_data.get('vatin', ''), 'text'), ('D5', business_activity_str, 'text'),
                ('D6', company_data.get('address_main', ''), 'text'), ('D7', company_data.get('i_request_date', ''), 'khmer_text'), 
                ('D10', company_data.get('phone', ''), 'text'), ('D11', bank_acc_num, 'text'), ('D12', bank_name, 'text'),
                ('D13', company_data.get('i_contact_person', ''), 'text'), ('D14', company_data.get('i_contact_position', ''), 'text'),
                ('H2', company_data.get('i_audit_timeline', ''), 'khmer_text'), ('H4', company_data.get('i_moc_number', ''), 'khmer_text'),
                ('H5', company_data.get('i_moc_date', ''), 'khmer_date'), ('H6', company_data.get('i_patent_date', ''), 'khmer_date'),        
                ('H7', company_data.get('i_vat_cert_date', ''), 'khmer_date'), ('H8', company_data.get('i_request_submission_date', ''), 'khmer_text'), 
                ('H9', company_data.get('i_amount_requested', ''), 'currency'), ('K6', company_data.get('i_patent_amount', ''), 'khmer_currency')
            ]
            
            for ref, val, val_type in company_mappings:
                cell = ws_info[ref]
                if val_type == 'date':
                    dt_val = to_excel_date(val)
                    if dt_val: cell.value, cell.number_format = dt_val, 'DD-MM-YYYY'
                    else: cell.value = val
                elif val_type == 'khmer_date':
                    dt_val = to_excel_date(val)
                    cell.value = to_khmer_numeral(dt_val.strftime('%d-%m-%Y') if dt_val else val)
                elif val_type == 'khmer_text': cell.value = to_khmer_numeral(val)
                elif val_type == 'currency':
                    cell.value, cell.number_format = clean_currency(val), '#,### "៛"'
                elif val_type == 'khmer_currency':
                    curr_val = clean_currency(val)
                    formatted_str = f"{int(curr_val):,}" if curr_val.is_integer() else f"{curr_val:,.2f}"
                    cell.value = f"{to_khmer_numeral(formatted_str)} ៛"
                else: cell.value = val
                
                cell.font, cell.alignment = khmer_font, Alignment(horizontal='left', vertical='center')

            auditors = [a.strip() for a in company_data.get('i_auditor_names', '').split(',')] if company_data.get('i_auditor_names', '') else []
            def ext_name(full_name):
                for t in ['កញ្ញា', 'លោកស្រី', 'លោក']:
                    if full_name.startswith(t): return full_name[len(t):].strip()
                return full_name

            ws_info['D8'].value, ws_info['E8'].value = auditors[0] if len(auditors)>0 else "", auditors[1] if len(auditors)>1 else ""
            ws_info['D9'].value, ws_info['E9'].value = ext_name(ws_info['D8'].value), ext_name(ws_info['E8'].value)
            for ref in ['D8', 'E8', 'D9', 'E9']: ws_info[ref].font = khmer_font

        ws1 = next((wb[n] for n in wb.sheetnames if n.strip().lower() == 'annex i-im state charge'), None)
        ws1_title = ws1.title if ws1 else 'Annex I-IM State Charge'
        ws1_sum_row = 9 + len(annex_i_rows)
        if ws1:
            start_row = 9
            if ws1.max_row >= start_row: ws1.delete_rows(start_row, ws1.max_row - start_row + 1)
            for i, row_data in enumerate(annex_i_rows):
                curr_row = start_row + i
                for col in range(1, 9): ws1.cell(row=curr_row, column=col).border, ws1.cell(row=curr_row, column=col).font, ws1.cell(row=curr_row, column=col).alignment = thin_border, khmer_font, align_middle
                ws1.cell(row=curr_row, column=1, value=i+1).alignment = align_center
                ws1.cell(row=curr_row, column=2, value=row_data[0]); ws1.cell(row=curr_row, column=3, value=row_data[1])
                dt_cell = ws1.cell(row=curr_row, column=4, value=to_excel_date(row_data[2])); dt_cell.alignment, dt_cell.number_format = align_center, 'DD-MM-YYYY'
                ws1.cell(row=curr_row, column=7, value=row_data[3]).number_format = '#,### "៛"'
            sum_row = start_row + len(annex_i_rows)
            ws1.merge_cells(start_row=sum_row, start_column=1, end_row=sum_row, end_column=6)
            ws1.cell(row=sum_row, column=1, value="សរុបអាករលើការនាំចូលជាបន្ទុករដ្ឋ").font, ws1.cell(row=sum_row, column=1).alignment = khmer_font_bold, align_center
            sum_cell = ws1.cell(row=sum_row, column=7, value=f"=SUM(G{start_row}:G{sum_row-1})"); sum_cell.font, sum_cell.number_format, sum_cell.alignment = khmer_font_bold, '#,### "៛"', align_right_middle
            for col in range(1, 9): ws1.cell(row=sum_row, column=col).fill, ws1.cell(row=sum_row, column=col).border = bg_gray_summary, thin_border

            sig_row = sum_row + 2
            ws1.merge_cells(start_row=sig_row, start_column=5, end_row=sig_row, end_column=8); ws1.cell(row=sig_row, column=5, value="រាជធានីភ្នំពេញ.ថ្ងៃទី          ខែ          ឆ្នាំ").font, ws1.cell(row=sig_row, column=5).alignment = khmer_font, align_center
            ws1.merge_cells(start_row=sig_row+1, start_column=5, end_row=sig_row+1, end_column=8); ws1.cell(row=sig_row+1, column=5, value="មន្ត្រីសវនកម្ម").font, ws1.cell(row=sig_row+1, column=5).alignment = khmer_font, align_center
            ws1.merge_cells(start_row=sig_row+3, start_column=5, end_row=sig_row+3, end_column=7); ws1.cell(row=sig_row+3, column=5, value="='Company information'!D9").font, ws1.cell(row=sig_row+3, column=5).alignment = khmer_font, align_center
            ws1.cell(row=sig_row+3, column=8, value="='Company information'!E9").font, ws1.cell(row=sig_row+3, column=8).alignment = khmer_font, align_center

        ws2 = next((wb[n] for n in wb.sheetnames if n.strip().lower() == 'annex ii-im non-state charge'), None)
        ws2_title = ws2.title if ws2 else 'Annex II-IM Non-State Charge'
        if ws2:
            red_font = Font(name='Khmer OS Siemreap', size=11, color='FF0000')
            ws2.cell(row=8, column=10, value="ខ្វះ").font = red_font
            ws2.cell(row=8, column=10).alignment = align_center
            ws2.cell(row=9, column=10, value="សម្គាល់").font = red_font
            ws2.cell(row=9, column=10).alignment = align_center

            start_row = 11
            if ws2.max_row >= start_row: ws2.delete_rows(start_row, ws2.max_row - start_row + 1)
            curr_row = start_row
            
            for i, row_data in enumerate(annex_ii_rows):
                for col in range(1, 12): ws2.cell(row=curr_row, column=col).border, ws2.cell(row=curr_row, column=col).font, ws2.cell(row=curr_row, column=col).alignment = thin_border, khmer_font, align_middle
                ws2.cell(row=curr_row, column=1, value=i+1).alignment = align_center
                ws2.cell(row=curr_row, column=2, value=row_data[0])
                ws2.cell(row=curr_row, column=3, value=row_data[2])
                dt_cell = ws2.cell(row=curr_row, column=4, value=to_excel_date(row_data[3]))
                dt_cell.alignment, dt_cell.number_format = align_center, 'DD-MM-YYYY'
                
                # Column G: Import Amount
                ws2.cell(row=curr_row, column=7, value=row_data[4]).number_format = '#,### "៛"'
                
                # Column I: Approve Amount (row_data[5])
                approve_amt = float(row_data[5]) if row_data[5] else 0.0
                ws2.cell(row=curr_row, column=9, value=approve_amt).number_format = '#,### "៛"'
                
                # Column J: Shortfall Formula (=G - I)
                ws2.cell(row=curr_row, column=10, value=f"=G{curr_row}-I{curr_row}").number_format = '#,### "៛"'
                
                # Column K: Note (row_data[6])
                ws2.cell(row=curr_row, column=11, value=clean_text(row_data[6]))
                curr_row += 1

            ws2.merge_cells(start_row=curr_row, start_column=1, end_row=curr_row, end_column=11)
            rc_header = ws2.cell(row=curr_row, column=1, value="II. អាករលើតម្លៃបន្ថែមតាមវិធីគិតអាករជំនួស(Reverse Charge)")
            rc_header.font, rc_header.alignment, rc_header.fill = khmer_font_bold, Alignment(horizontal='left', vertical='center', wrap_text=False), bg_gray_header
            for col in range(1, 12): ws2.cell(row=curr_row, column=col).border = thin_border
            curr_row += 1

            for i, row_data in enumerate(rc_rows):
                for col in range(1, 12): ws2.cell(row=curr_row, column=col).border, ws2.cell(row=curr_row, column=col).font, ws2.cell(row=curr_row, column=col).alignment = thin_border, khmer_font, align_middle
                ws2.cell(row=curr_row, column=1, value=i+1).alignment = align_center
                ws2.cell(row=curr_row, column=2, value=row_data[0])
                ws2.cell(row=curr_row, column=3, value=row_data[2])
                dt_cell = ws2.cell(row=curr_row, column=4, value=to_excel_date(row_data[3]))
                dt_cell.alignment, dt_cell.number_format = align_center, 'DD-MM-YYYY'
                
                # RC Import equivalent (G)
                ws2.cell(row=curr_row, column=7, value=row_data[3]).number_format = '#,### "៛"'
                ws2.cell(row=curr_row, column=8, value="អនុញ្ញាត (បានប្រកាស)").alignment = align_center
                
                # RC Approve Amount (I) defaults to matching Import
                ws2.cell(row=curr_row, column=9, value=f"=G{curr_row}").number_format = '#,### "៛"'
                
                # RC Shortfall (J)
                ws2.cell(row=curr_row, column=10, value=f"=G{curr_row}-I{curr_row}").number_format = '#,### "៛"'
                
                # RC Note (K)
                ws2.cell(row=curr_row, column=11, value="")
                curr_row += 1

            sum_row = curr_row
            ws2.merge_cells(start_row=sum_row, start_column=1, end_row=sum_row, end_column=6)
            ws2.cell(row=sum_row, column=1, value="សរុបអាករលើការនាំចូល ឬ អាករលើតម្លៃបន្ថែមតាមវិធីគិតអាករជំនួស(Reverse Charge)").font, ws2.cell(row=sum_row, column=1).alignment = khmer_font_bold, align_right_middle
            
            # G Total
            ws2.cell(row=sum_row, column=7, value=f"=SUM(G{start_row}:G{sum_row-1})").font, ws2.cell(row=sum_row, column=7).alignment, ws2.cell(row=sum_row, column=7).number_format = khmer_font_bold, align_right_middle, '#,### "៛"'
            ws2.cell(row=sum_row, column=8, value="សរុបទឺកប្រាក់អនុញ្ញាត").font, ws2.cell(row=sum_row, column=8).alignment = khmer_font_bold, align_right_middle
            
            # I Total (Approve)
            ws2.cell(row=sum_row, column=9, value=f"=SUM(I{start_row}:I{sum_row-1})").font, ws2.cell(row=sum_row, column=9).alignment, ws2.cell(row=sum_row, column=9).number_format = khmer_font_bold, align_right_middle, '#,### "៛"'
            
            # J Total (Shortfall)
            ws2.cell(row=sum_row, column=10, value=f"=SUM(J{start_row}:J{sum_row-1})").font, ws2.cell(row=sum_row, column=10).alignment, ws2.cell(row=sum_row, column=10).number_format = khmer_font_bold, align_right_middle, '#,### "៛"'
            
            # K Total (None - it's a string note field)
            ws2.cell(row=sum_row, column=11, value="")
            
            for col in range(1, 12): ws2.cell(row=sum_row, column=col).fill, ws2.cell(row=sum_row, column=col).border = bg_gray_summary, thin_border
            ws2_end_row = sum_row - 1 

            decl_row = sum_row + 2
            ws2.merge_cells(start_row=decl_row, start_column=1, end_row=decl_row+1, end_column=9)
            
            try:
                from openpyxl.cell.text import InlineFont
                from openpyxl.cell.rich_text import TextBlock, CellRichText
                
                bold_font = InlineFont(rFont='Khmer OS Siemreap', sz=11, b=True)
                normal_font = InlineFont(rFont='Khmer OS Siemreap', sz=11)
                
                rich_text = CellRichText(
                    TextBlock(bold_font, 'សេចក្តីធានាអះអាងរបស់មន្ត្រីសវនករទទួលបន្ទុក៖\n'),
                    TextBlock(normal_font, 'លទ្ធផលផ្ទៀងផ្ទាត់ឥណទានអាករ ចំពោះការនាំចូល ឬ/និង អាករលើតម្លៃបន្ថែមតាមវិធីគិតអាករជំនួស(Reverse Charge)ខាងលើពិតជាត្រឹមត្រូវតាមរបាយការណ៍លម្អិតដែលបានទាញទិន្នន័យពីអគ្គនាយកដ្ឋានគយ/ប្រព័ន្ធE-Filing ពិតប្រាកដមែន។')
                )
                ws2.cell(row=decl_row, column=1).value = rich_text
            except ImportError:
                ws2.cell(row=decl_row, column=1, value="សេចក្តីធានាអះអាងរបស់មន្ត្រីសវនករទទួលបន្ទុក៖\nលទ្ធផលផ្ទៀងផ្ទាត់ឥណទានអាករ ចំពោះការនាំចូល ឬ/និង អាករលើតម្លៃបន្ថែមតាមវិធីគិតអាករជំនួស(Reverse Charge)ខាងលើពិតជាត្រឹមត្រូវតាមរបាយការណ៍លម្អិតដែលបានទាញទិន្នន័យពីអគ្គនាយកដ្ឋានគយ/ប្រព័ន្ធE-Filing ពិតប្រាកដមែន។").font = khmer_font

            ws2.cell(row=decl_row, column=1).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

            sig_row = decl_row + 3
            ws2.merge_cells(start_row=sig_row, start_column=5, end_row=sig_row, end_column=9); ws2.cell(row=sig_row, column=5, value="រាជធានីភ្នំពេញ.ថ្ងៃទី          ខែ          ឆ្នាំ").font, ws2.cell(row=sig_row, column=5).alignment = khmer_font, align_center
            ws2.merge_cells(start_row=sig_row+1, start_column=5, end_row=sig_row+1, end_column=9); ws2.cell(row=sig_row+1, column=5, value="មន្ត្រីសវនកម្ម").font, ws2.cell(row=sig_row+1, column=5).alignment = khmer_font, align_center
            ws2.merge_cells(start_row=sig_row+3, start_column=5, end_row=sig_row+3, end_column=8); ws2.cell(row=sig_row+3, column=5, value="='Company information'!D9").font, ws2.cell(row=sig_row+3, column=5).alignment = khmer_font, align_center
            ws2.cell(row=sig_row+3, column=9, value="='Company information'!E9").font, ws2.cell(row=sig_row+3, column=9).alignment = khmer_font, align_center

        ws5_title = next((n for n in wb.sheetnames if n.strip().lower() == 'annex v-local sale'), 'Annex V-Local Sale')
        ws5_sum_row = 9 + len(annex_v_rows)

        # --- PART C.2: Annex III Local Purchase ---
        ws3 = next((wb[n] for n in wb.sheetnames if n.strip().lower() == 'annexiii-local pur'), None)
        if ws3:
            start_row = 9
            if ws3.max_row >= start_row: ws3.delete_rows(start_row, ws3.max_row - start_row + 1)
            
            for i, p_row in enumerate(annex_iii_local_purchases):
                curr_row = start_row + i
                
                p_inv_val = p_row[3] or ""
                p_inv_clean = clean_invoice_text(p_inv_val)
                
                ws3.cell(row=curr_row, column=1, value=i+1).alignment = align_center
                ws3.cell(row=curr_row, column=2, value=clean_text(p_row[0]))
                ws3.cell(row=curr_row, column=3, value=clean_text(p_row[1]))
                ws3.cell(row=curr_row, column=4, value=clean_text(p_row[2]))
                ws3.cell(row=curr_row, column=5, value=p_inv_val)
                
                raw_date = p_row[4]
                dt_val = ""
                if raw_date and str(raw_date).lower() not in ['nan', 'nat', 'none', '']:
                    try: dt_val = pd.to_datetime(raw_date).date()
                    except: dt_val = str(raw_date).split()[0]
                ws3.cell(row=curr_row, column=6, value=dt_val).number_format = 'DD-MM-YYYY'
                
                amt = float(p_row[5]) if p_row[5] else 0.0
                ws3.cell(row=curr_row, column=7, value=amt).number_format = '#,### "៛"'
                ws3.cell(row=curr_row, column=9, value=amt)
                
                status_formula = f'=IF(AND(S{curr_row}=TRUE, T{curr_row}=TRUE, U{curr_row}=TRUE), IF(W{curr_row}<-0.05, "អនុញ្ញាត (អ្នកផ្គត់ផ្គង់ប្រកាសខ្វះ)", "បានប្រកាស (អនុញ្ញាត)"), IF(AND(S{curr_row}=FALSE, T{curr_row}=FALSE, U{curr_row}=FALSE), "ព្យួរទុក (មិនមានទិន្នន័យ)", "ប្រកាសខុស (ព្យួរទុក)"))'
                ws3.cell(row=curr_row, column=10, value=status_formula).font = khmer_font
                
                user_status_val = p_row[7]
                if not user_status_val or str(user_status_val).strip().lower() in ['none', 'null', 'nan']:
                    user_status_val = ""
                ws3.cell(row=curr_row, column=11, value=user_status_val).font = khmer_font
                
                ws3.cell(row=curr_row, column=12, value=p_row[8] or "").font = khmer_font 
                ws3.cell(row=curr_row, column=13, value=None) 
                ws3.cell(row=curr_row, column=14, value=f"=IF(W{curr_row}<0,AH{curr_row},I{curr_row})") 
                ws3.cell(row=curr_row, column=15, value=f"=I{curr_row}-M{curr_row}") 
                ws3.cell(row=curr_row, column=16, value=None) 
                ws3.cell(row=curr_row, column=17, value=p_inv_clean) 
                
                d_row = dec_map.get(p_inv_clean)
                d_inv_val = ""
                ag_val = 0.0
                
                if d_row:
                    d_inv_val = clean_text(d_row[1] if d_row else "")
                    ag_val = float(d_row[10] if d_row[10] else 0)
                    
                    raw_d_date = d_row[0] if d_row else ""
                    dt_d_val = ""
                    if raw_d_date and str(raw_d_date).lower() not in ['nan', 'nat', 'none', '']:
                        try: dt_d_val = pd.to_datetime(raw_d_date).date()
                        except: dt_d_val = str(raw_d_date).split()[0]
                    ws3.cell(row=curr_row, column=24, value=dt_d_val).number_format = 'DD-MM-YYYY'

                    ws3.cell(row=curr_row, column=25, value=d_inv_val)
                    ws3.cell(row=curr_row, column=26, value=clean_text(d_row[2] if d_row else ""))
                    ws3.cell(row=curr_row, column=27, value=clean_text(d_row[3] if d_row else ""))
                    ws3.cell(row=curr_row, column=28, value=clean_text(d_row[4] if d_row else ""))
                    ws3.cell(row=curr_row, column=29, value=clean_text(d_row[5] if d_row else ""))
                    
                    for idx, col_num in enumerate(range(30, 43)): 
                        ws3.cell(row=curr_row, column=col_num, value=float(d_row[6 + idx]) if d_row[6 + idx] else 0.0)
                        
                    ws3.cell(row=curr_row, column=43, value=clean_text(d_row[19] if d_row else ""))
                    ws3.cell(row=curr_row, column=44, value=clean_text(d_row[20] if d_row else ""))
                    ws3.cell(row=curr_row, column=45, value=clean_text(d_row[21] if d_row else ""))

                ws3.cell(row=curr_row, column=18, value=clean_invoice_text(d_inv_val)) 
                ws3.cell(row=curr_row, column=19, value=f"=Q{curr_row}=R{curr_row}") 
                ws3.cell(row=curr_row, column=20, value=f"=AND(MONTH(F{curr_row})=MONTH(X{curr_row}), YEAR(F{curr_row})=YEAR(X{curr_row}))")
                
                tin_formula = f'=AND(AB{curr_row}<>"", \'Company information\'!D$4<>"", RIGHT(SUBSTITUTE(AB{curr_row},"-",""),9)=RIGHT(SUBSTITUTE(\'Company information\'!D$4,"-",""),9))'
                ws3.cell(row=curr_row, column=21, value=tin_formula) 
                
                ws3.cell(row=curr_row, column=22, value=f"=AH{curr_row}-I{curr_row}") 
                ws3.cell(row=curr_row, column=23, value=None) 

                format_cols = [9, 13, 14, 15, 23] + list(range(30, 43))
                for col_idx in format_cols:
                    ws3.cell(row=curr_row, column=col_idx).number_format = '#,###0'

                for col in range(1, 46):
                    cell = ws3.cell(row=curr_row, column=col)
                    cell.border, cell.font = thin_border, khmer_font
                    if col not in [1, 6, 24]: cell.alignment = align_middle
                    else: cell.alignment = align_center

            end_data_row = start_row + len(annex_iii_local_purchases) - 1
            if end_data_row < start_row: end_data_row = start_row

            sum_row = end_data_row + 2
            total_cell = ws3.cell(row=sum_row, column=1, value="Total")
            total_cell.font, total_cell.alignment = khmer_font_bold, align_right_middle
            total_cell.number_format = '#,### "៛"'
            
            for col_letter, col_idx in [('I', 9), ('M', 13), ('N', 14)]:
                sum_cell = ws3.cell(row=sum_row, column=col_idx, value=f"=SUM({col_letter}{start_row}:{col_letter}{end_data_row})")
                sum_cell.font, sum_cell.number_format, sum_cell.alignment = khmer_font_bold, '#,### "៛"', align_right_middle
            
            for col in range(1, 16): 
                cell = ws3.cell(row=sum_row, column=col); cell.fill, cell.border = bg_gray_summary, thin_border

            sum_table_start = sum_row + 2
            ws3.cell(row=sum_table_start, column=1, value="=\"តារាងសង្ខេបប្រាក់អាករដែលអាចស្នើសុំបង្វិលសង \" & 'Company information'!D2").font = khmer_font_bold
            
            h_row = sum_table_start + 1
            headers = ["ចំនួន.វិ", "បរិយាយ", "ចំនួនទឹកប្រាក់", "តាង", "ផ្សេងៗ"]
            for col_idx, h_text in enumerate(headers, 1):
                cell = ws3.cell(row=h_row, column=col_idx, value=h_text)
                cell.font, cell.border, cell.alignment = khmer_font_bold, thin_border, align_center

            d_row1 = h_row + 1; ws3.cell(row=d_row1, column=2, value="ចំនួនប្រាក់អាករលើការនាំចូល").font = khmer_font; ws3.cell(row=d_row1, column=3, value=f"='{ws1_title}'!G{ws1_sum_row}").number_format = '#,### "៛"'
            d_row2 = d_row1 + 1; ws3.cell(row=d_row2, column=2, value="ចំនួនប្រាក់អាករលើធាតុចូលទិញក្នុងស្រុក").font = khmer_font; ws3.cell(row=d_row2, column=3, value=f"=I{sum_row}").number_format = '#,### "៛"'
            d_row3 = d_row2 + 1; ws3.cell(row=d_row3, column=2, value="ចំនួនប្រាក់អាករលើធាតុចូលសរុប").font = khmer_font; ws3.cell(row=d_row3, column=3, value=f"=C{d_row1}+C{d_row2}").number_format = '#,### "៛"'
            d_row4 = d_row3 + 1; ws3.cell(row=d_row4, column=2, value="ចំនួនប្រាក់អាករលើធាតុចេញលក់ក្នុងស្រុក").font = khmer_font; ws3.cell(row=d_row4, column=3, value=f"='{ws5_title}'!G{ws5_sum_row}").number_format = '#,### "៛"'
            d_row5 = d_row4 + 1; ws3.cell(row=d_row5, column=2, value="ចំនួនប្រាក់អាករលើធាតុចេញសរុប").font = khmer_font; ws3.cell(row=d_row5, column=3, value=f"=C{d_row4}").number_format = '#,### "៛"'
            d_row6 = d_row5 + 1; ws3.cell(row=d_row6, column=2, value="ចំនួនប្រាក់អាករដែលអាចធ្វើការផ្ទៀងផ្ទាត់").font = khmer_font_bold; ws3.cell(row=d_row6, column=3, value=f"=C{d_row3}-C{d_row5}").number_format = '#,### "៛"'; ws3.cell(row=d_row6, column=3).font = khmer_font_bold
            
            d_row7 = d_row6 + 1; ws3.cell(row=d_row7, column=2, value="ចំនួនប្រាក់អាករស្នើសុំតាមប្រព័ន្ធ E-VAT").font = khmer_font_bold; ws3.cell(row=d_row7, column=3, value="='Company information'!H9").number_format = '#,### "៛"'; ws3.cell(row=d_row7, column=3).font = khmer_font_bold; ws3.cell(row=d_row7, column=4, value="ក").font = khmer_font_bold; ws3.cell(row=d_row7, column=4).alignment = align_center
            
            ws3.column_dimensions['B'].width = 40

            current_sum_row = d_row7 + 1
            khmer_alphabet = ['ខ', 'គ', 'ឃ', 'ង', 'ច', 'ឆ', 'ជ', 'ឈ', 'ញ', 'ដ', 'ឋ', 'ឌ', 'ឍ', 'ណ']
            
            visible_chars = []
            visible_rows_for_calc = []
            alphabet_index = 0
            
            for status in status_configs:
                raw_stat_name = str(status[0])
                stat_summary = str(status[1])
                stat_action = str(status[2])
                
                eval_count_query = """
                    SELECT COUNT(*) FROM purchase 
                    WHERE ovatr = ? AND purchase > 0 
                    AND (
                        user_status = ? 
                        OR ( (user_status IS NULL OR user_status = '') AND sys_status = ? )
                    )
                """
                eval_res = con.execute(eval_count_query, [ovatr_code, raw_stat_name, raw_stat_name]).fetchone()
                eval_count = int(eval_res[0]) if eval_res and eval_res[0] else 0

                safe_stat_name = raw_stat_name.replace('"', '""')
                
                count_formula = f'=COUNTIFS($K$9:$K${end_data_row}, "{safe_stat_name}") + COUNTIFS($K$9:$K${end_data_row}, "", $J$9:$J${end_data_row}, "{safe_stat_name}")'
                ws3.cell(row=current_sum_row, column=1, value=count_formula).alignment = align_center
                ws3.cell(row=current_sum_row, column=2, value=stat_summary).font = khmer_font
                
                sum_formula = f'=SUMIFS($G$9:$G${end_data_row}, $K$9:$K${end_data_row}, "{safe_stat_name}") + SUMIFS($G$9:$G${end_data_row}, $K$9:$K${end_data_row}, "", $J$9:$J${end_data_row}, "{safe_stat_name}")'
                ws3.cell(row=current_sum_row, column=3, value=sum_formula).number_format = '#,### "៛"'
                
                ws3.cell(row=current_sum_row, column=5, value=stat_action).font = khmer_font

                if eval_count > 0:
                    kh_char = khmer_alphabet[alphabet_index] if alphabet_index < len(khmer_alphabet) else str(alphabet_index)
                    ws3.cell(row=current_sum_row, column=4, value=kh_char).font = khmer_font_bold
                    ws3.cell(row=current_sum_row, column=4).alignment = align_center
                    
                    visible_chars.append(kh_char)
                    visible_rows_for_calc.append(current_sum_row)
                    alphabet_index += 1
                else:
                    ws3.cell(row=current_sum_row, column=4, value="") 
                    ws3.row_dimensions[current_sum_row].hidden = True 
                
                current_sum_row += 1

            d_row_final = current_sum_row
            ws3.cell(row=d_row_final, column=2, value="លម្អៀងបា្រក់អាករជាមួយប្រព័ន្ធ E-VAT").font = khmer_font
            ws3.cell(row=d_row_final, column=3, value=f"=C{d_row7}-C{d_row6}").number_format = '#,### "៛"'
            
            final_char = khmer_alphabet[alphabet_index] if alphabet_index < len(khmer_alphabet) else "ចុង"
            visible_chars.append(final_char)
            visible_rows_for_calc.append(d_row_final)
            
            ws3.cell(row=d_row_final, column=4, value=final_char).font = khmer_font_bold; ws3.cell(row=d_row_final, column=4).alignment = align_center
            ws3.cell(row=d_row_final, column=5, value="ព្យួរទុក").font = khmer_font

            d_row_total = d_row_final + 1
            ws3.cell(row=d_row_total, column=2, value="សរុបប្រាក់អាករគួរបង្វិលសងជូនក្រុមហ៊ុន").font = khmer_font_bold
            
            sum_formula = f"=C{d_row7}"
            for r in visible_rows_for_calc: 
                sum_formula += f"-C{r}"
            ws3.cell(row=d_row_total, column=3, value=sum_formula).number_format = '#,### "៛"'; ws3.cell(row=d_row_total, column=3).font = khmer_font_bold
            
            total_formula_text = f"សរុប=ក-{'-'.join(visible_chars)}"
            ws3.cell(row=d_row_total, column=4, value=total_formula_text).font = khmer_font_bold; ws3.cell(row=d_row_total, column=4).alignment = align_center

            ws3.merge_cells(start_row=d_row1, start_column=1, end_row=d_row7, end_column=1)
            for r in range(d_row1, d_row_total + 1):
                for c in range(1, 6): ws3.cell(row=r, column=c).border = thin_border

            # --- SIGNATURE AND DECLARATION SECTION ---
            sig_start_row = d_row_total + 2
            ws3.merge_cells(start_row=sig_start_row, start_column=1, end_row=sig_start_row+2, end_column=5)
            
            try:
                from openpyxl.cell.text import InlineFont
                from openpyxl.cell.rich_text import TextBlock, CellRichText
                
                bold_font = InlineFont(rFont='Khmer OS Siemreap', sz=11, b=True)
                normal_font = InlineFont(rFont='Khmer OS Siemreap', sz=11)
                
                rich_text = CellRichText(
                    TextBlock(bold_font, 'សេចក្តីធានាអះអាងរបស់មន្ត្រីសវនករទទួលបន្ទុក៖\n'),
                    TextBlock(normal_font, 'លទ្ធផលនៃការផ្ទៀងផ្ទាត់វិក្កយបត្រអាករ (Invoice Cross-check) ខាងលើ ពិតជាត្រឹមត្រូវតាមការប្រកាសរបស់អ្នកផ្គត់ផ្គង់ពិតប្រាកដមែន។')
                )
                ws3.cell(row=sig_start_row, column=1).value = rich_text
            except ImportError:
                ws3.cell(row=sig_start_row, column=1, value="សេចក្តីធានាអះអាងរបស់មន្ត្រីសវនករទទួលបន្ទុក៖\nលទ្ធផលនៃការផ្ទៀងផ្ទាត់វិក្កយបត្រអាករ (Invoice Cross-check) ខាងលើ ពិតជាត្រឹមត្រូវតាមការប្រកាសរបស់អ្នកផ្គត់ផ្គង់ពិតប្រាកដមែន។").font = khmer_font

            ws3.cell(row=sig_start_row, column=1).alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            ws3.row_dimensions[sig_start_row].height = 50

            row_step2 = sig_start_row + 3
            ws3.merge_cells(start_row=row_step2, start_column=1, end_row=row_step2, end_column=5)
            ws3.cell(row=row_step2, column=1, value="រាជធានីភ្នំពេញ.ថ្ងៃទី          ខែ                    ឆ្នាំ២០២").font = Font(name='Khmer OS Siemreap', size=12)
            ws3.cell(row=row_step2, column=1).alignment = align_center

            row_step3 = sig_start_row + 4
            ws3.merge_cells(start_row=row_step3, start_column=1, end_row=row_step3, end_column=5)
            ws3.cell(row=row_step3, column=1, value="មន្ត្រីសវនកម្ម").font = khmer_font
            ws3.cell(row=row_step3, column=1).alignment = align_center
            
            ws3.merge_cells(start_row=row_step3, start_column=6, end_row=row_step3, end_column=11)
            ws3.cell(row=row_step3, column=6, value="បានឃើញ និងឯកភាព").font = khmer_font
            ws3.cell(row=row_step3, column=6).alignment = align_center

            row_step4 = sig_start_row + 5
            ws3.merge_cells(start_row=row_step4, start_column=6, end_row=row_step4, end_column=11)
            ws3.cell(row=row_step4, column=6, value="រាជធានីភ្នំពេញ.ថ្ងៃទី          ខែ                    ឆ្នាំ២០២").font = khmer_font
            ws3.cell(row=row_step4, column=6).alignment = align_center

            row_step5 = sig_start_row + 6
            ws3.merge_cells(start_row=row_step5, start_column=6, end_row=row_step5, end_column=11)
            ws3.cell(row=row_step5, column=6, value="ប្រធានការិយាល័យ").font = khmer_font
            ws3.cell(row=row_step5, column=6).alignment = align_center

            row_step6 = sig_start_row + 7
            ws3.merge_cells(start_row=row_step6, start_column=1, end_row=row_step6, end_column=2)
            ws3.cell(row=row_step6, column=1, value="='Company information'!D9").font = khmer_font
            ws3.cell(row=row_step6, column=1).alignment = align_center

            ws3.merge_cells(start_row=row_step6, start_column=3, end_row=row_step6, end_column=5)
            ws3.cell(row=row_step6, column=3, value="='Company information'!E9").font = khmer_font
            ws3.cell(row=row_step6, column=3).alignment = align_center

        ws4 = next((wb[n] for n in wb.sheetnames if n.strip().lower() == 'annex iv-ex'), None)
        if ws4:
            start_row = 9
            if ws4.max_row >= start_row: ws4.delete_rows(start_row, ws4.max_row - start_row + 1)
            for i, row_data in enumerate(annex_iv_rows):
                curr_row = start_row + i
                for col in range(1, 6): cell = ws4.cell(row=curr_row, column=col); cell.border = thin_border; cell.font = khmer_font; cell.alignment = align_middle
                ws4.cell(row=curr_row, column=1, value=i+1).alignment = align_center; ws4.cell(row=curr_row, column=2, value=row_data[0]); ws4.cell(row=curr_row, column=3, value=row_data[1])
                dt_cell = ws4.cell(row=curr_row, column=4, value=to_excel_date(row_data[2])); dt_cell.alignment = align_center; dt_cell.number_format = 'DD-MM-YYYY'
                ws4.cell(row=curr_row, column=5, value=row_data[3]).number_format = '#,### "៛"'
            sum_row = start_row + len(annex_iv_rows)
            ws4.merge_cells(start_row=sum_row, start_column=1, end_row=sum_row, end_column=4); ws4.cell(row=sum_row, column=1, value="សរុបការនាំចេញ").font = khmer_font_bold; ws4.cell(row=sum_row, column=1).alignment = align_center
            sum_cell = ws4.cell(row=sum_row, column=5, value=f"=SUM(E{start_row}:E{sum_row-1})"); sum_cell.font = khmer_font_bold; sum_cell.number_format = '#,### "៛"'; sum_cell.alignment = align_center
            for col in range(1, 6): cell = ws4.cell(row=sum_row, column=col); cell.fill = bg_gray_summary; cell.border = thin_border

            sig_row = sum_row + 2
            ws4.merge_cells(start_row=sig_row, start_column=4, end_row=sig_row, end_column=5)
            ws4.cell(row=sig_row, column=4, value="រាជធានីភ្នំពេញ.ថ្ងៃទី           ខែ           ឆ្នាំ").font = khmer_font
            ws4.cell(row=sig_row, column=4).alignment = align_center
            ws4.merge_cells(start_row=sig_row+1, start_column=4, end_row=sig_row+1, end_column=5)
            ws4.cell(row=sig_row+1, column=4, value="មន្ត្រីសវនកម្ម").font = khmer_font
            ws4.cell(row=sig_row+1, column=4).alignment = align_center
            ws4.cell(row=sig_row+3, column=4, value="='Company information'!D9").font = khmer_font
            ws4.cell(row=sig_row+3, column=4).alignment = align_center
            ws4.cell(row=sig_row+3, column=5, value="='Company information'!E9").font = khmer_font
            ws4.cell(row=sig_row+3, column=5).alignment = align_center

        ws5 = next((wb[n] for n in wb.sheetnames if n.strip().lower() == 'annex v-local sale'), None)
        if ws5:
            start_row = 9
            if ws5.max_row >= start_row: ws5.delete_rows(start_row, ws5.max_row - start_row + 1)
            for i, row_data in enumerate(annex_v_rows):
                curr_row = start_row + i
                for col in range(1, 9): cell = ws5.cell(row=curr_row, column=col); cell.border = thin_border; cell.font = khmer_font; cell.alignment = align_middle
                ws5.cell(row=curr_row, column=1, value=i+1).alignment = align_center; ws5.cell(row=curr_row, column=2, value=row_data[0]); ws5.cell(row=curr_row, column=3, value=row_data[1])
                dt = ws5.cell(row=curr_row, column=4, value=to_excel_date(row_data[2])); dt.alignment = align_center; dt.number_format = 'DD-MM-YYYY'
                ws5.cell(row=curr_row, column=7, value=row_data[3]).number_format = '#,### "៛"'
            sum_row = start_row + len(annex_v_rows)
            ws5.merge_cells(start_row=sum_row, start_column=1, end_row=sum_row, end_column=6); ws5.cell(row=sum_row, column=1, value="សរុបការលក់ក្នុងស្រុក").font = khmer_font_bold; ws5.cell(row=sum_row, column=1).alignment = align_center
            sum_cell = ws5.cell(row=sum_row, column=7, value=f"=SUM(G{start_row}:G{sum_row-1})"); sum_cell.font = khmer_font_bold; sum_cell.number_format = '#,### "៛"'; sum_cell.alignment = align_center
            for col in range(1, 9): cell = ws5.cell(row=sum_row, column=col); cell.fill = bg_gray_summary; cell.border = thin_border

            sig_row = sum_row + 2
            ws5.merge_cells(start_row=sig_row, start_column=7, end_row=sig_row, end_column=8)
            ws5.cell(row=sig_row, column=7, value="រាជធានីភ្នំពេញ.ថ្ងៃទី           ខែ           ឆ្នាំ").font = khmer_font
            ws5.cell(row=sig_row, column=7).alignment = align_center
            ws5.merge_cells(start_row=sig_row+1, start_column=7, end_row=sig_row+1, end_column=8)
            ws5.cell(row=sig_row+1, column=7, value="មន្ត្រីសវនកម្ម").font = khmer_font
            ws5.cell(row=sig_row+1, column=7).alignment = align_center
            ws5.cell(row=sig_row+3, column=7, value="='Company information'!D9").font = khmer_font
            ws5.cell(row=sig_row+3, column=7).alignment = align_center
            ws5.cell(row=sig_row+3, column=8, value="='Company information'!E9").font = khmer_font
            ws5.cell(row=sig_row+3, column=8).alignment = align_center

        ws_tp = next((wb[n] for n in wb.sheetnames if n.strip().lower() == 'taxpaid'), None)
        if ws_tp and taxpaid_raw:
            month_keys = ['jan', 'feb', 'mar', 'apr', 'may', 'jun', 'jul', 'aug', 'sep', 'oct', 'nov', 'dec']
            khmer_months = {
                'មករា': 1, 'កុម្ភៈ': 2, 'មីនា': 3, 'មេសា': 4,
                'ឧសភា': 5, 'មិថុនា': 6, 'កក្កដា': 7, 'សីហា': 8,
                'កញ្ញា': 9, 'តុលា': 10, 'វិច្ឆិកា': 11, 'ធ្នូ': 12
            }
            month_cols = {1: 'jan', 2: 'feb', 3: 'mar', 4: 'apr', 5: 'may', 6: 'jun', 7: 'jul', 8: 'aug', 9: 'sep', 10: 'oct', 11: 'nov', 12: 'dec'}
            
            start_m, start_y, end_m, end_y = None, None, None, None
            req_date_str = str(company_data.get('i_request_date', '')).strip()
            years_found = re.findall(r'\b(20\d{2})\b', req_date_str)
            months_found = []
            for k_month in khmer_months.keys():
                occurrences = [m.start() for m in re.finditer(k_month, req_date_str)]
                for idx in occurrences: months_found.append((idx, khmer_months[k_month]))
            months_found.sort(key=lambda x: x[0])
            
            if len(years_found) >= 1:
                start_y = int(years_found[0])
                end_y = int(years_found[-1])
            if len(months_found) >= 1:
                start_m = months_found[0][1]
                end_m = months_found[-1][1]
            
            processed_taxpaid = []
            for r in taxpaid_raw:
                rd = dict(zip(tp_cols, r))
                try: t_year = int(rd.get('tax_year', 0))
                except: t_year = 0
                
                if start_y and end_y:
                    if t_year < start_y or t_year > end_y: continue 
                    if t_year:
                        for m_num, m_col in month_cols.items():
                            is_before_start = (t_year == start_y and start_m is not None and m_num < start_m)
                            is_after_end = (t_year == end_y and end_m is not None and m_num > end_m)
                            if is_before_start or is_after_end: rd[m_col] = 0
                processed_taxpaid.append(rd)

            grouped_data = {}
            years = sorted(list(set(rd.get('tax_year') for rd in processed_taxpaid)))
            header_row, data_start_row = 5, 6
            if ws_tp.max_row >= header_row: ws_tp.delete_rows(header_row, ws_tp.max_row - header_row + 1)

            ws_tp.cell(row=header_row, column=2, value="ល.រ").font = khmer_font; ws_tp.cell(row=header_row, column=2).alignment = align_center
            ws_tp.cell(row=header_row, column=3, value="ប្រភេទពន្ធ").font = khmer_font; ws_tp.cell(row=header_row, column=3).alignment = align_right_middle
            ws_tp.cell(row=header_row, column=4, value="ចំនួនទឹកប្រាក់ពន្ធ").font = khmer_font; ws_tp.cell(row=header_row, column=4).alignment = align_right_middle
            
            header_map = []
            for yr in years:
                for m in month_keys: header_map.append((f"{m.capitalize()}-{yr}", m, yr))
            for idx, (display, _, _) in enumerate(header_map):
                cell = ws_tp.cell(row=header_row, column=5 + idx, value=display)
                cell.font = khmer_font; cell.alignment = align_right_middle
            
            for col in range(2, 5 + len(header_map)):
                cell = ws_tp.cell(row=header_row, column=col); cell.fill = bg_yellow; cell.border = thin_border

            for rd in processed_taxpaid:
                desc, yr = rd.get('description', 'Unknown'), rd.get('tax_year')
                if desc not in grouped_data: grouped_data[desc] = {}
                for m in month_keys: grouped_data[desc][f"{m}-{yr}"] = rd.get(m, 0)

            for i, (desc, months_dict) in enumerate(grouped_data.items()):
                curr_row = data_start_row + i
                c_no = ws_tp.cell(row=curr_row, column=2, value=i+1); c_no.font = khmer_font; c_no.border = thin_border; c_no.alignment = align_center
                c_desc = ws_tp.cell(row=curr_row, column=3, value=desc); c_desc.font = khmer_font; c_desc.border = thin_border; c_desc.alignment = align_right_middle
                for m_idx, (display_key, m_key, yr) in enumerate(header_map):
                    val = months_dict.get(f"{m_key}-{yr}", 0)
                    cell = ws_tp.cell(row=curr_row, column=5 + m_idx, value=val); cell.font = khmer_font; cell.border = thin_border; cell.alignment = align_right_middle
                    cell.number_format = '#,### "៛"' if val != 0 else '#,###0'
                lc = openpyxl.utils.get_column_letter(4 + len(header_map))
                c_sum = ws_tp.cell(row=curr_row, column=4, value=f"=SUM(E{curr_row}:{lc}{curr_row})")
                c_sum.font = khmer_font_bold; c_sum.border = thin_border; c_sum.alignment = align_right_middle; c_sum.number_format = '#,### "៛"'

            final_data_row = data_start_row + len(grouped_data) - 1
            if final_data_row < data_start_row: final_data_row = data_start_row 

            sum_row = final_data_row + 1
            ws_tp.cell(row=sum_row, column=3, value="សរុបទឹកប្រាក់ពន្ធបានបង់ចូលរដ្ឋ").font = khmer_font_bold; ws_tp.cell(row=sum_row, column=3).alignment = align_right_middle
            v_sum = ws_tp.cell(row=sum_row, column=4, value=f"=SUM(D{data_start_row}:D{final_data_row})")
            v_sum.font = khmer_font_bold; v_sum.alignment = align_right_middle; v_sum.number_format = '#,### "៛"'
            for col in range(2, 5 + len(header_map)): ws_tp.cell(row=sum_row, column=col).border = thin_border; ws_tp.cell(row=sum_row, column=col).fill = bg_gray_summary

        save_dir = os.path.join(settings.MEDIA_ROOT, 'reports'); os.makedirs(save_dir, exist_ok=True)
        fname = f"Audit_Result_{ovatr_code}.xlsx"; full_path = os.path.join(save_dir, fname); wb.save(full_path)
        return FileResponse(open(full_path, 'rb'), as_attachment=True, filename=fname)
    finally:
        con.close()
        
def download_word_report(request):
    ovatr_code = request.GET.get('ovatr_code')
    if not ovatr_code:
        return JsonResponse({'status': 'error', 'message': 'Missing OVATR code'}, status=400)

    con = get_db_connection()
    try:
        from docxtpl import DocxTemplate, RichText
        import re

        # 1. Fetch Company Info
        row = con.execute("SELECT * FROM company_info WHERE ovatr = ?", [ovatr_code]).fetchone()
        if not row:
            return JsonResponse({'status': 'error', 'message': 'Company info not found'}, status=404)
        
        cols = [desc[0] for desc in con.description]
        company_data = dict(zip(cols, row))

        # --- Helper Formatting Functions ---
        def to_khmer_numeral(text):
            if text is None or text == "": return ""
            khmer_digits = "០១២៣៤៥៦៧៨៩"
            return "".join(khmer_digits[int(c)] if c.isdigit() else c for c in str(text))

        def khmer_currency(val, hide_zero=False, include_symbol=True):
            try:
                clean_val = str(val).replace(',', '').replace('៛', '').strip()
                v = float(clean_val) if clean_val else 0.0
                v_rounded = round(v)
                
                if v_rounded == 0: 
                    return "" if hide_zero else ("0 ៛" if include_symbol else "0")
                    
                formatted = f"{v_rounded:,}"
                if include_symbol:
                    formatted += " ៛"
                return formatted
            except:
                return "" if hide_zero else ("0 ៛" if include_symbol else "0")

        def format_khmer_date(date_val):
            if not date_val: return ""
            for fmt in ('%d-%m-%Y', '%Y-%m-%d', '%d/%m/%Y', '%Y/%m/%d'):
                try: 
                    dt = datetime.strptime(str(date_val).strip(), fmt)
                    return to_khmer_numeral(dt.strftime('%d-%m-%Y'))
                except: continue
            return to_khmer_numeral(str(date_val))
            
        def parse_khmer_submission_date(text):
            if not text: return ""
            khmer_to_arabic = str.maketrans('០១២៣៤៥៦៧៨៩', '0123456789')
            khmer_months = {
                'មករា': '01', 'កុម្ភៈ': '02', 'មីនា': '03', 'មេសា': '04',
                'ឧសភា': '05', 'មិថុនា': '06', 'កក្កដា': '07', 'សីហា': '08',
                'កញ្ញា': '09', 'តុលា': '10', 'វិច្ឆិកា': '11', 'ធ្នូ': '12'
            }
            text_clean = str(text).replace('\u200b', '')
            match = re.search(r'ថ្ងៃទី\s*([០-៩0-9]+)\s*ខែ\s*([ក-៿]+)\s*ឆ្នាំ\s*([០-៩0-9]+)', text_clean)
            if match:
                day = match.group(1).translate(khmer_to_arabic).zfill(2)
                month_str = match.group(2).strip()
                year = match.group(3).translate(khmer_to_arabic)
                month = '01'
                for k_m, num in khmer_months.items():
                    if k_m in month_str:
                        month = num
                        break
                return f"{day}/{month}/{year}"
            
            for fmt in ('%d-%m-%Y', '%Y-%m-%d', '%d/%m/%Y', '%Y/%m/%d'):
                try: 
                    dt = datetime.strptime(str(text_clean).strip(), fmt)
                    return dt.strftime('%d/%m/%Y')
                except: continue
            return str(text_clean)

        def clean_invoice_text(val):
            if pd.isna(val) or val is None: return ""
            return re.sub(r'[^A-Z0-9]', '', str(val).upper())

        def to_excel_date(date_val):
            if not date_val: return None
            for fmt in ('%d-%m-%Y', '%Y-%m-%d', '%d/%m/%Y', '%Y/%m/%d'):
                try: return datetime.strptime(str(date_val).strip(), fmt)
                except: continue
            return None
            
        # --- TIN CLEANUP HELPER ---
        def get_last_9_digits(val):
            if pd.isna(val) or val is None: return ""
            # Strip everything except numbers (removes hyphens, letters, etc.)
            digits = re.sub(r'\D', '', str(val))
            # Return strictly the last 9 numbers
            return digits[-9:] if len(digits) >= 9 else digits

        business_activity_str = ""
        try:
            activities = json.loads(company_data.get('business_activities', '[]'))
            if activities: business_activity_str = f"{activities[0].get('name', '')} {activities[0].get('desc', '')}".strip()
        except: pass

        bank_acc_num, bank_name = "", ""
        try:
            accounts = json.loads(company_data.get('enterprise_accounts', '[]'))
            if accounts:
                bank_acc_num = str(accounts[0].get('number', ''))
                bank_name = str(accounts[0].get('bank', ''))
        except: pass

        # ======================================================================
        # 2. NATIVE PYTHON SUMMARY CALCULATOR (Using DB Data)
        # ======================================================================
        
        sum_ws1_raw = con.execute("SELECT SUM(import_state_charge) FROM purchase WHERE ovatr = ? AND import_state_charge <> 0", [ovatr_code]).fetchone()
        sum_ws1 = float(sum_ws1_raw[0]) if sum_ws1_raw and sum_ws1_raw[0] else 0.0
        
        sum_ws5_raw = con.execute("SELECT SUM(vat_local_sale) FROM sale WHERE ovatr = ? AND vat_local_sale <> 0", [ovatr_code]).fetchone()
        sum_ws5 = float(sum_ws5_raw[0]) if sum_ws5_raw and sum_ws5_raw[0] else 0.0

        try: status_configs = con.execute("SELECT name, summary, action FROM user_status_config").fetchall()
        except: status_configs = []
        
        status_sums = {str(stat[0]): 0.0 for stat in status_configs}

        annex_iii_local_purchases = con.execute("SELECT description, supplier_name, supplier_tin, invoice_no, date, purchase, status, user_status, comment FROM purchase WHERE ovatr = ? AND purchase > 0 ORDER BY CAST(no AS INTEGER) ASC", [ovatr_code]).fetchall()
        
        annex_iii_raw_decs = con.execute("""
            SELECT d.date, d.invoice_number, d.tax_registration_id, d.vat_local_sale, p.invoice_no
            FROM purchase p
            LEFT JOIN tax_declaration d ON regexp_replace(upper(d.invoice_number), '[^A-Z0-9]', '', 'g') = regexp_replace(upper(p.invoice_no), '[^A-Z0-9]', '', 'g')
            WHERE p.ovatr = ?
        """, [ovatr_code]).fetchall()
        
        dec_map = {clean_invoice_text(d[4]): d for d in annex_iii_raw_decs if clean_invoice_text(d[4]) and d[1]}
        
        # --- GET CLEAN 9-DIGIT COMPANY TIN FOR MATCHING ---
        user_vatin_9 = get_last_9_digits(company_data.get('vatin', ''))

        sum_ws3 = 0.0
        for p in annex_iii_local_purchases:
            p_inv_val = p[3] or ""
            p_inv_clean = clean_invoice_text(p_inv_val)
            p_date = to_excel_date(p[4])
            p_vat = float(p[5]) if p[5] else 0.0
            sum_ws3 += p_vat
            
            d_row = dec_map.get(p_inv_clean)
            
            d_inv_clean = clean_invoice_text(d_row[1] if d_row else "")
            S_match = (p_inv_clean == d_inv_clean) and bool(p_inv_clean)
            
            d_date = to_excel_date(d_row[0] if d_row else "")
            T_match = (p_date.month == d_date.month and p_date.year == d_date.year) if (p_date and d_date) else False
            
            # --- FIXED: CROSSCHECK DB TIN (Only compare last 9 digits of supplier declared TIN vs Company TIN) ---
            d_tin = d_row[2] if d_row else ""
            d_tin_9 = get_last_9_digits(d_tin)
            
            U_match = (d_tin_9 == user_vatin_9) and bool(d_tin_9) and bool(user_vatin_9)
            # -----------------------------------------------------------------------------------------------------
            
            d_vat = float(d_row[3]) if d_row and d_row[3] else 0.0
            W_diff = d_vat - p_vat
            
            if S_match and T_match and U_match:
                sys_status = "អនុញ្ញាត (អ្នកផ្គត់ផ្គង់ប្រកាសខ្វះ)" if W_diff < -0.05 else "បានប្រកាស (អនុញ្ញាត)"
            else:
                if not S_match and not T_match and not U_match:
                    sys_status = "ព្យួរទុក (មិនមានទិន្នន័យ)"
                else:
                    sys_status = "ប្រកាសខុស (ព្យួរទុក)"
                    
            user_status = p[7]
            final_status = user_status.strip() if (user_status and str(user_status).strip().lower() not in ['none', 'null', 'nan', '']) else sys_status
            
            if final_status in status_sums:
                status_sums[final_status] += p_vat

        total_in = sum_ws1 + sum_ws3
        total_out = sum_ws5
        verifiable_vat = total_in - total_out
        
        req_str = str(company_data.get('i_amount_requested', '0')).replace(',', '').strip()
        try: 
            extracted_nums = re.findall(r"[-+]?\d*\.\d+|\d+", req_str)
            requested_vat = float(extracted_nums[0]) if extracted_nums else 0.0
        except: requested_vat = 0.0

        excel_summary = []
        suspense_amount = 0.0
        
        def add_summary_row(desc, amount, other='', is_bold=False, indent=False):
            nonlocal suspense_amount
            if str(other).strip() == 'ព្យួរទុក':
                suspense_amount += float(amount)
                
            fmt_amt = khmer_currency(amount, hide_zero=True, include_symbol=True)
            if fmt_amt != "":
                desc_text = "\t    " + desc if indent else desc
                if is_bold:
                    excel_summary.append({
                        'description': RichText(desc_text, bold=True, size=22, font='Khmer OS Siemreap'),
                        'total_amount': RichText(fmt_amt, bold=True, size=22, font='Khmer OS Siemreap'),
                        'other': RichText(other, bold=True, size=22, font='Khmer OS Siemreap') if other else ""
                    })
                else:
                    excel_summary.append({
                        'description': desc_text,
                        'total_amount': fmt_amt,
                        'other': other
                    })

        add_summary_row('ចំនួនប្រាក់អាករលើការនាំចូល', sum_ws1)
        add_summary_row('ចំនួនប្រាក់អាករលើធាតុចូលទិញក្នុងស្រុក', sum_ws3)
        add_summary_row('ចំនួនប្រាក់អាករលើធាតុចូលសរុប', total_in)
        add_summary_row('ចំនួនប្រាក់អាករលើធាតុចេញលក់ក្នុងស្រុក', sum_ws5)
        add_summary_row('ចំនួនប្រាក់អាករលើធាតុចេញសរុប', total_out)
        
        add_summary_row('ចំនួនប្រាក់អាករដែលអាចធ្វើការផ្ទៀងផ្ទាត់', verifiable_vat, is_bold=True)
        add_summary_row('ចំនួនប្រាក់អាករស្នើសុំតាមប្រព័ន្ធ E-VAT', requested_vat, is_bold=True)
        
        total_deductions = 0.0
        suspended_reasons_list = [] 
        
        for stat in status_configs:
            raw_stat_name = str(stat[0])
            stat_summary = str(stat[1])
            stat_action = str(stat[2])
            
            eval_sum = status_sums.get(raw_stat_name, 0.0)
            add_summary_row(stat_summary, eval_sum, stat_action, indent=True)
            total_deductions += eval_sum

            if stat_action.strip() == 'ព្យួរទុក' and round(eval_sum) > 0:
                match = re.search(r'\((.*?)\)', raw_stat_name)
                extracted_reason = match.group(1).strip() if match else raw_stat_name
                if extracted_reason not in suspended_reasons_list:
                    suspended_reasons_list.append(extracted_reason)

        diff_vat = requested_vat - verifiable_vat
        add_summary_row('លម្អៀងបា្រក់អាករជាមួយប្រព័ន្ធ E-VAT', diff_vat, 'ព្យួរទុក', indent=True)
        total_deductions += diff_vat
        
        final_refundable = requested_vat - total_deductions
        
        excel_summary.append({
            'description': RichText('\t    សរុបប្រាក់អាករគួរបង្វិលសងជូនក្រុមហ៊ុន', bold=True, size=22, font='Khmer OS Siemreap'), 
            'total_amount': RichText(khmer_currency(final_refundable, hide_zero=False, include_symbol=True), bold=True, size=22, font='Khmer OS Siemreap'), 
            'other': ''
        })
        
        refund_amount = requested_vat - suspense_amount

        if len(suspended_reasons_list) > 1:
            suspended_reason_text = " និង ".join([", ".join(suspended_reasons_list[:-1]), suspended_reasons_list[-1]])
        elif len(suspended_reasons_list) == 1:
            suspended_reason_text = suspended_reasons_list[0]
        else:
            suspended_reason_text = "-"

        # ======================================================================
        # 3. EXTRA AGGREGATION QUERIES FOR CONTEXT
        # ======================================================================
        def safe_query_float(query, params=[]):
            try:
                res = con.execute(query, params).fetchone()
                if res and res[0] is not None:
                    clean_val = str(res[0]).replace(',', '').replace('៛', '').strip()
                    return float(clean_val)
                return 0.0
            except: return 0.0

        def safe_query_int(query, params=[]):
            try:
                res = con.execute(query, params).fetchone()
                return int(res[0]) if res and res[0] else 0
            except: return 0

        import_state_charge = safe_query_float('SELECT SUM(import_state_charge) FROM purchase WHERE ovatr = ? AND import_state_charge <> 0', [ovatr_code])
        count_import_state_charge = safe_query_int('SELECT COUNT(*) FROM purchase WHERE ovatr = ? AND import_state_charge <> 0', [ovatr_code])

        rc_vat_sum = safe_query_float('SELECT SUM(vat) FROM reverse_charge WHERE ovatr = ? AND vat <> 0', [ovatr_code])
        rc_count = safe_query_int('SELECT COUNT(*) FROM reverse_charge WHERE ovatr = ? AND vat <> 0', [ovatr_code])

        import_non_state_charge_base = safe_query_float('SELECT SUM("import") FROM purchase WHERE ovatr = ? AND "import" <> 0', [ovatr_code])
        count_import_non_state_charge_base = safe_query_int('SELECT COUNT(*) FROM purchase WHERE ovatr = ? AND "import" <> 0', [ovatr_code])
        
        import_non_state_charge = import_non_state_charge_base + rc_vat_sum
        count_import_non_state_charge = count_import_non_state_charge_base + rc_count
        
        import_include_vat_base = safe_query_float('SELECT SUM(exclude_vat) FROM purchase WHERE ovatr = ? AND "import" <> 0', [ovatr_code])
        rc_base_sum = rc_vat_sum * 10
        import_include_vat = import_include_vat_base + rc_base_sum

        vat_local_sale = safe_query_float('SELECT SUM(vat_local_sale) FROM sale WHERE ovatr = ? AND vat_local_sale <> 0', [ovatr_code])
        count_vat_local_sale = safe_query_int('SELECT COUNT(*) FROM sale WHERE ovatr = ? AND vat_local_sale <> 0', [ovatr_code])

        export_val = safe_query_float('SELECT SUM(vat_export) FROM sale WHERE ovatr = ? AND vat_export <> 0', [ovatr_code])
        count_export = safe_query_int('SELECT COUNT(*) FROM sale WHERE ovatr = ? AND vat_export <> 0', [ovatr_code])

        purchase_include_vat = safe_query_float('SELECT SUM(exclude_vat) FROM purchase WHERE ovatr = ? AND purchase <> 0', [ovatr_code])
        export_include_vat = safe_query_float('SELECT SUM(vat_export) FROM sale WHERE ovatr = ? AND vat_export <> 0', [ovatr_code])
        sale_include_vat = safe_query_float('SELECT SUM(vat_local_sale) FROM sale WHERE ovatr = ? AND vat_local_sale <> 0', [ovatr_code])

        purchase_val = safe_query_float('SELECT SUM(purchase) FROM purchase WHERE ovatr = ? AND purchase <> 0', [ovatr_code])
        
        total_purchase_include_vat = import_include_vat + purchase_include_vat
        total_purchase_vat = import_non_state_charge + purchase_val 
        total_sale_include_vat = export_include_vat + sale_include_vat
        total_sale_vat_val = vat_local_sale
        
        total_verify_vat = total_purchase_vat - total_sale_vat_val

        # ======================================================================
        # 4. TAX PAID LIST 
        # ======================================================================
        tax_rows = con.execute("""
            SELECT description, SUM(total) 
            FROM tax_paid 
            WHERE ovatr = ? 
            GROUP BY description 
            HAVING SUM(total) > 0
        """, [ovatr_code]).fetchall()

        tax_list = []
        grand_total_tax = 0.0
        for i, r in enumerate(tax_rows):
            desc = r[0]
            amt = r[1]
            grand_total_tax += amt
            tax_list.append({
                'no': to_khmer_numeral(str(i + 1)),
                'description': desc,
                'amount': khmer_currency(amt, include_symbol=False)
            })

        # ======================================================================
        # 5. Build Word Context 
        # ======================================================================
        context = {
            'company_name_kh': company_data.get('company_name_kh', ''),
            'company_name_en': company_data.get('company_name_en', ''),
            'vatin': company_data.get('vatin', ''),
            'business_activity': business_activity_str,
            'address_main': company_data.get('address_main', ''),
            'request_period': to_khmer_numeral(company_data.get('i_request_date', '')),
            'auditor_names': company_data.get('i_auditor_names', ''),
            'audit_timeline': to_khmer_numeral(company_data.get('i_audit_timeline', '')),
            'bank_account_no': bank_acc_num,
            'bank_name': bank_name,
            'contact_person': company_data.get('i_contact_person', ''),
            'contact_position': company_data.get('i_contact_position', ''),
            
            'moc_number': to_khmer_numeral(company_data.get('i_moc_number', '')),
            'moc_date': format_khmer_date(company_data.get('i_moc_date', '')),
            'patent_date': format_khmer_date(company_data.get('i_patent_date', '')),
            'vat_cert_date': format_khmer_date(company_data.get('i_vat_cert_date', '')),
            
            'request_submission_date': to_khmer_numeral(company_data.get('i_request_submission_date', '')),
            'request_submission_date_formatted': parse_khmer_submission_date(str(company_data.get('i_request_submission_date', ''))),
            
            'patent_amount': khmer_currency(company_data.get('i_patent_amount', ''), hide_zero=False, include_symbol=False),
            'amount_requested': khmer_currency(company_data.get('i_amount_requested', ''), hide_zero=False, include_symbol=False),
            'suspense_amount': khmer_currency(suspense_amount, include_symbol=False),
            'refund_amount': khmer_currency(refund_amount, include_symbol=False),
            'import_state_charge': khmer_currency(import_state_charge, include_symbol=False),
            'import_non_state_charge': khmer_currency(import_non_state_charge, include_symbol=False),
            'vat_local_sale': khmer_currency(vat_local_sale, include_symbol=False),
            'export': khmer_currency(export_val, include_symbol=False),
            'import_include_vat': khmer_currency(import_include_vat, include_symbol=False),
            'purchase_include_vat': khmer_currency(purchase_include_vat, include_symbol=False),
            'export_include_vat': khmer_currency(export_include_vat, include_symbol=False),
            'sale_include_vat': khmer_currency(sale_include_vat, include_symbol=False),
            'purchase': khmer_currency(purchase_val, include_symbol=False),
            'total_purchase_include_vat': khmer_currency(total_purchase_include_vat, include_symbol=False),
            'total_purchase_vat': khmer_currency(total_purchase_vat, include_symbol=False),
            'total_sale_include_vat': khmer_currency(total_sale_include_vat, include_symbol=False),
            'total_sale_vat': khmer_currency(total_sale_vat_val, include_symbol=False),
            'total_verify_vat': khmer_currency(total_verify_vat, include_symbol=False),

            'suspended_reason': suspended_reason_text,
            
            'count_import_state_charge': str(count_import_state_charge),
            'count_import_non_state_charge': str(count_import_non_state_charge),
            'count_vat_local_sale': str(count_vat_local_sale),
            'count_export': str(count_export),

            'excel_summary': excel_summary,
            'tax_list': tax_list,
            'tax_total_amount': khmer_currency(grand_total_tax, include_symbol=False)
        }

        # 6. Generate Word Document
        template_path = os.path.join(settings.BASE_DIR, 'templates', 'Sample-Word_Report.docx')
        if not os.path.exists(template_path):
            template_path = os.path.join(settings.MEDIA_ROOT, 'templates', 'Sample-Word_Report.docx')

        doc = DocxTemplate(template_path)
        doc.render(context)

        file_stream = io.BytesIO()
        doc.save(file_stream)
        file_stream.seek(0)

        filename = f"Audit_Report_{ovatr_code}.docx"
        response = HttpResponse(
            file_stream.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    except Exception as e:
        import traceback
        print(traceback.format_exc())
        return JsonResponse({'status': 'error', 'message': str(e)}, status=500)
    finally:
        con.close()
        