import os
import duckdb
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from copy import copy
from django.conf import settings

# --- STYLING CONSTANTS ---
FONT_KHMER = Font(name='Khmer OS Siemreap', size=10)
BORDER_THIN = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)

class ReportGenerator:
    def __init__(self, ovatr_code):
        self.ovatr = ovatr_code
        self.db_path = os.path.join(settings.BASE_DIR, 'datawarehouse.duckdb')
        self.template_path = os.path.join(settings.BASE_DIR, 'templates', 'Sample-Excel_Report.xlsx')
        
        # Fallback if specific template doesn't exist, use static
        if not os.path.exists(self.template_path):
             self.template_path = os.path.join(settings.BASE_DIR, 'core', 'templates', 'static', 'Sample-Excel_Report.xlsx')

    def _get_connection(self):
        return duckdb.connect(self.db_path)

    def generate(self):
        if not os.path.exists(self.template_path):
            raise FileNotFoundError("Report template not found.")

        wb = load_workbook(self.template_path)
        con = self._get_connection()

        try:
            # 1. Company Info
            self._fill_company_info(wb, con)

            # 2. Tax Paid
            self._fill_taxpaid(wb, con)

            # 3. Purchases (Annex I, II, III)
            self._fill_purchases(wb, con)

            # 4. Sales (Annex IV, V)
            self._fill_sales(wb, con)

            # 5. Reverse Charge
            self._fill_reverse_charge(wb, con)

            # Save to temp
            output_dir = os.path.join(settings.MEDIA_ROOT, 'reports')
            os.makedirs(output_dir, exist_ok=True)
            filename = f"Tax_Declaration_Report_{self.ovatr}.xlsx"
            output_path = os.path.join(output_dir, filename)
            wb.save(output_path)
            
            return output_path

        finally:
            con.close()

    def _fill_sheet(self, ws, data, start_row=8, col_mapping=None):
        """
        Generic helper to fill a sheet with data.
        col_mapping: dict { db_col_index: excel_col_index } (1-based)
        """
        if not data:
            return

        # 1. Capture style from the first data row (template row)
        styles = {}
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(row=start_row, column=c)
            styles[c] = {
                'font': copy(cell.font),
                'border': copy(cell.border),
                'alignment': copy(cell.alignment),
                'number_format': cell.number_format
            }

        # 2. Clear existing rows from start_row onwards
        if ws.max_row >= start_row:
             ws.delete_rows(start_row, ws.max_row - start_row + 1)

        # 3. Write Data
        for i, row_data in enumerate(data):
            r = start_row + i
            
            # Write Sequence No (Column A usually)
            ws.cell(row=r, column=1, value=i+1)

            # Write mapped columns
            if col_mapping:
                for db_idx, xl_idx in col_mapping.items():
                    val = row_data[db_idx]
                    ws.cell(row=r, column=xl_idx, value=val)
            else:
                # Default mapping: direct match skipping col 1 (sequence)
                for c_idx, val in enumerate(row_data):
                    ws.cell(row=r, column=c_idx + 2, value=val)

            # Apply Styles
            for c in range(1, ws.max_column + 1):
                if c in styles:
                    cell = ws.cell(row=r, column=c)
                    s = styles[c]
                    cell.font = copy(s['font'])
                    cell.border = copy(s['border'])
                    cell.alignment = copy(s['alignment'])
                    cell.number_format = s['number_format']

    def _fill_company_info(self, wb, con):
        if 'Company information' not in wb.sheetnames: return
        ws = wb['Company information']
        
        row = con.execute("SELECT * FROM company_info WHERE ovatr = ?", [self.ovatr]).fetchone()
        if not row: return

        # Fetch column names to map properly
        cols = [d[0] for d in con.description]
        data = dict(zip(cols, row))

        # Mapping logic (This depends heavily on your specific Excel template cell locations)
        # Assuming a standard layout based on input file analysis
        # You might need to adjust row/col indices based on the actual Template file
        
        # Example Mapping:
        # ws['C4'] = data.get('company_name_kh', '')
        # ws['C5'] = data.get('company_name_en', '')
        # ws['H6'] = data.get('vatin', '')
        pass # Placeholder: Implementing flexible mapping requires seeing the exact template grid

    def _fill_taxpaid(self, wb, con):
        if 'Taxpaid' not in wb.sheetnames: return
        ws = wb['Taxpaid']
        
        data = con.execute("""
            SELECT description, jan, feb, mar, apr, may, jun, jul, aug, sep, oct, nov, dec, total 
            FROM tax_paid WHERE ovatr = ?
        """, [self.ovatr]).fetchall()
        
        # Assuming Data starts at Row 5
        # DB Cols 0-13 -> Excel Cols C-P
        mapping = {
            0: 3, 1: 4, 2: 5, 3: 6, 4: 7, 5: 8, 6: 9, 
            7: 10, 8: 11, 9: 12, 10: 13, 11: 14, 12: 15, 13: 16
        }
        self._fill_sheet(ws, data, start_row=5, col_mapping=mapping)

    def _fill_purchases(self, wb, con):
        # 1. Annex III - Local Purchases
        if 'AnnexIII-Local Pur' in wb.sheetnames:
            ws = wb['AnnexIII-Local Pur']
            # Fetch local purchases
            # Schema: description, supplier_name, supplier_tin, invoice_no, date, purchase, vat
            # We calculate VAT = total / 1.1 * 0.1 if strictly standard, but let's use stored values
            data = con.execute("""
                SELECT description, supplier_name, supplier_tin, invoice_no, date, purchase 
                FROM purchase 
                WHERE ovatr = ? AND purchase > 0
                ORDER BY CAST(no AS INTEGER)
            """, [self.ovatr]).fetchall()
            
            # Mapping: Desc(0)->B, Supp(1)->C, TIN(2)->D, Inv(3)->E, Date(4)->F, Amount(5)->G
            mapping = {0: 2, 1: 3, 2: 4, 3: 5, 4: 6, 5: 7}
            self._fill_sheet(ws, data, start_row=8, col_mapping=mapping)
            
            # Add Formulas for VAT (10%)
            for r in range(8, 8 + len(data)):
                ws[f'I{r}'] = f"=G{r}*0.1" # Assuming 10% VAT
                ws[f'H{r}'] = 0 # Non-creditable placeholder

        # 2. Annex I - Imports (State Charge)
        if 'Annex I-IM State Charge' in wb.sheetnames:
            ws = wb['Annex I-IM State Charge']
            data = con.execute("""
                SELECT description, supplier_name, invoice_no, date, purchase_state_charge
                FROM purchase 
                WHERE ovatr = ? AND purchase_state_charge > 0
                ORDER BY CAST(no AS INTEGER)
            """, [self.ovatr]).fetchall()
            # Adjust mapping based on specific Annex I columns
            mapping = {0: 2, 1: 3, 2: 4, 3: 5, 4: 6}
            self._fill_sheet(ws, data, start_row=8, col_mapping=mapping)

    def _fill_sales(self, wb, con):
        # Annex V - Local Sale
        if 'Annex V-Local Sale' in wb.sheetnames:
            ws = wb['Annex V-Local Sale']
            # Fetch Sales
            data = con.execute("""
                SELECT description, buyer_name, tax_registration_id, invoice_no, date, total_invoice_amount, 
                       amount_exclude_vat, vat_local_sale
                FROM sale 
                WHERE ovatr = ? AND vat_local_sale > 0
                ORDER BY CAST(no AS INTEGER)
            """, [self.ovatr]).fetchall()
            
            # Mapping: Desc(0)->B, Buyer(1)->C, TIN(2)->D, Inv(3)->E, Date(4)->F, Total(5)->G, Base(6)->H, VAT(7)->I
            mapping = {0: 2, 1: 3, 2: 4, 3: 5, 4: 6, 5: 7, 6: 8, 7: 9}
            self._fill_sheet(ws, data, start_row=8, col_mapping=mapping)

    def _fill_reverse_charge(self, wb, con):
        # Often mapped to a specific sheet or Annex
        # Logic similar to above
        pass