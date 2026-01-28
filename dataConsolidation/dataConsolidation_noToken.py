import asyncio
from copy import copy
import os
import logging
import re
import shutil
import subprocess
import sys
import time
from telegram import Update, BotCommand
from telegram.ext import ApplicationBuilder, MessageHandler, filters, ContextTypes, CommandHandler
import patoolib
import sqlite3
import pandas as pd
import traceback
import openpyxl
from datetime import datetime

# Configuration
BASE_DIR = r'C:\Users\TiiTii\Desktop\Q_Wrkr'
DOWNLOAD_DIR = os.path.join(BASE_DIR, 'Bot Downloads')
EXTRACT_DIR = os.path.join(BASE_DIR, 'Bot Extract')
ARCHIVE_DIR = os.path.join(BASE_DIR, 'Archive')
DATABASE_PATH = os.path.join(BASE_DIR, 'dataWarehouse.sqlite3')

# Ensure directories exist
os.makedirs(DOWNLOAD_DIR, exist_ok=True)
os.makedirs(ARCHIVE_DIR, exist_ok=True)

# Add new TINs here as strings inside the list
TINS_SPLIT_BY_YEAR = [
    "L001-100044638",
    # "K002-123456789",  <-- You can add more here later
]

class SaleDataImporter:

    def __init__(self, share_drive_path, db_path=DATABASE_PATH):
        """
        Initialize SaleDataImporter
        :param share_drive_path: Path to directory containing Excel files
        :param db_path: Path to SQLite database
        """
        self.share_drive_path = share_drive_path
        self.db_path = db_path

        # Configure logging
        logging.basicConfig(
            level=logging.DEBUG,
            format='%(asctime)s - %(levelname)s: %(message)s',
            handlers=[
                logging.FileHandler('import_db.log'),
                logging.StreamHandler()
            ]
        )

    def check_table_exists(self, table_name):
        """
        Check if a table exists in the database
        :param table_name: Name of the table to check
        :return: Boolean indicating table existence
        """
        try:
            with sqlite3.connect(self.db_path) as conn:
                cursor = conn.cursor()
                cursor.execute("""
                    SELECT name FROM sqlite_master
                    WHERE type='table' AND name=?
                """, (table_name,))
                return cursor.fetchone() is not None
        except sqlite3.Error as e:
            logging.error(f"Error checking table existence: {e}")
            return False

    def process_excel_file(self, file_path, username="unknown"):
        """
        Process individual Excel file with comprehensive data mapping and optimized database operations
        using executemany for batch inserts
        :param file_path: Path to Excel file
        :param username: Username of the uploader
        :return: Number of rows imported
        """
        try:
            # Extract TIN from filename
            file_name = os.path.basename(file_path)
            tin_match = re.match(r'^([LKB]\d{3}-\d{9})_', file_name)
            file_tin = tin_match.group(1) if tin_match else ""

            # Read Excel file, skipping first 3 rows
            df = pd.read_excel(file_path, skiprows=3, dtype={2: str})

            # Current timestamp as string in ISO format
            current_timestamp = pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')

            # Comprehensive column mapping
            processed_df = pd.DataFrame({
                # Date column - format as dd-mm-yyyy string
                'date': pd.to_datetime(df.iloc[:,1], format='%d-%m-%Y', dayfirst=True).dt.strftime('%d-%m-%Y'),
                # Invoice and other basic information
                'invoice_number': df.iloc[:, 2].astype(str),
                'credit_notification_letter_number': df.iloc[:, 3].astype(str),
                'buyer_type': df.iloc[:, 4],
                'tax_registration_id': df.iloc[:, 5],
                'buyer_name': df.iloc[:, 6],
                'total_invoice_amount': pd.to_numeric(df.iloc[:, 7], errors='coerce'),
                # VAT and tax-related columns
                'amount_exclude_vat': pd.to_numeric(df.iloc[:, 8], errors='coerce'),
                'non_vat_sales': pd.to_numeric(df.iloc[:, 9], errors='coerce'),
                'vat_zero_rate': pd.to_numeric(df.iloc[:, 10], errors='coerce'),
                'vat_local_sale': pd.to_numeric(df.iloc[:, 11], errors='coerce'),
                'vat_export': pd.to_numeric(df.iloc[:, 12], errors='coerce'),
                'vat_local_sale_state_burden': pd.to_numeric(df.iloc[:, 13], errors='coerce'),
                'vat_withheld_by_national_treasury': pd.to_numeric(df.iloc[:, 14], errors='coerce'),
                # Additional tax-related columns
                'plt': pd.to_numeric(df.iloc[:, 15], errors='coerce'),
                'special_tax_on_goods': pd.to_numeric(df.iloc[:, 16], errors='coerce'),
                'special_tax_on_services': pd.to_numeric(df.iloc[:, 17], errors='coerce'),
                'accommodation_tax': pd.to_numeric(df.iloc[:, 18], errors='coerce'),
                'income_tax_redemption_rate': pd.to_numeric(df.iloc[:, 19], errors='coerce'),
                # Notes and description
                'notes': df.iloc[:, 20].fillna(''),
                'description': df.iloc[:, 21].fillna(''),
                'tax_declaration_status': df.iloc[:, 22].fillna(''),
                # Add telegram username
                'telegram_username': username,
                # Metadata - use string instead of Timestamp object
                'import_timestamp': current_timestamp,
                'file_tin': file_tin
            })

            # Replace NaN with 0 for numeric columns
            numeric_columns = [
                'amount_exclude_vat', 'non_vat_sales', 'vat_zero_rate',
                'vat_local_sale', 'vat_export', 'vat_local_sale_state_burden',
                'vat_withheld_by_national_treasury', 'plt',
                'special_tax_on_goods', 'special_tax_on_services',
                'accommodation_tax', 'income_tax_redemption_rate'
            ]
            processed_df[numeric_columns] = processed_df[numeric_columns].fillna(0)

            # Group by invoice number and take the first occurrence
            unique_df = processed_df.groupby('invoice_number').first().reset_index()

            # Connect to SQLite with optimized configuration
            with sqlite3.connect(self.db_path) as conn:
                # Optimization 4: Configure SQLite for performance
                conn.execute('PRAGMA journal_mode = WAL')
                conn.execute('PRAGMA synchronous = NORMAL')
                conn.execute('PRAGMA cache_size = 10000')
                conn.execute('PRAGMA temp_store = MEMORY')
                conn.execute('PRAGMA mmap_size = 30000000000')

                # Optimization 1: Use transactions
                conn.execute('BEGIN TRANSACTION')

                # Optimization 2: Batch existence checks
                # Collect all invoice numbers, tax IDs, and VAT values for checking
                all_invoice_numbers = unique_df['invoice_number'].tolist()

                # Check existence in batches (SQLite has limits on parameter count)
                batch_size = 500
                new_records = []

                for i in range(0, len(unique_df), batch_size):
                    batch_df = unique_df.iloc[i:i+batch_size]
                    batch_invoice_numbers = batch_df['invoice_number'].tolist()

                    # Create placeholders for SQL query
                    placeholders = ','.join(['?'] * len(batch_invoice_numbers))

                    # Get existing invoice numbers with tax_registration_id and vat_local_sale
                    cursor = conn.cursor()
                    cursor.execute(f"""
                        SELECT invoice_number, tax_registration_id, vat_local_sale
                        FROM tax_declaration
                        WHERE invoice_number IN ({placeholders})
                    """, batch_invoice_numbers)

                    # Create a set of existing records for fast lookup
                    existing_records = {
                        (inv, tin, vat): True
                        for inv, tin, vat in cursor.fetchall()
                    }

                    # Add non-existing records
                    for _, row in batch_df.iterrows():
                        key = (row['invoice_number'], row['tax_registration_id'], row['vat_local_sale'])
                        if key not in existing_records:
                            new_records.append(row)

                # Optimization 3: Use executemany for batch inserts
                if new_records:
                    new_df = pd.DataFrame(new_records)

                    # Get column names and prepare data for executemany
                    columns = new_df.columns
                    column_names = ', '.join(columns)
                    placeholders = ', '.join(['?'] * len(columns))

                    # Convert DataFrame to list of tuples
                    # Ensure all values are of types supported by SQLite
                    values = []
                    for _, row in new_df.iterrows():
                        # Convert each value to a SQLite-compatible type
                        row_values = []
                        for val in row:
                            if isinstance(val, pd.Timestamp):
                                # Convert Timestamp to string
                                row_values.append(val.strftime('%Y-%m-%d %H:%M:%S'))
                            else:
                                row_values.append(val)
                        values.append(tuple(row_values))

                    # Prepare insert query
                    insert_query = f"INSERT INTO tax_declaration ({column_names}) VALUES ({placeholders})"

                    # Execute batch insert
                    cursor = conn.cursor()
                    cursor.executemany(insert_query, values)

                    # Commit the transaction
                    conn.commit()

                    new_records_count = len(new_records)
                    logging.info(f"Imported {new_records_count} new unique invoice rows from {file_path}")
                    return new_records_count
                else:
                    logging.info(f"No new unique invoice rows to import from {file_path}")
                    return 0

        except Exception as e:
            logging.error(f"Error processing {file_path}: {e}")
            logging.error(f"Detailed error: {traceback.format_exc()}")
            return 0

class TelegramSaleDataBot:

    def __init__(self, token):
        """
        Initialize Telegram Bot
        :param token: Telegram Bot Token
        """

        # Add a flag to track if the bot is active
        self.is_active = True

        self.production_chat_id = PRODUCTION_GROUP_ID

        # Configure logging
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(name)s - %(levelname)s: %(message)s',
            handlers=[
                logging.FileHandler('telegram_bot.log', encoding='utf-8'),
                logging.StreamHandler(stream=sys.stdout)
            ]
        )
        # Reduce httpx logging level to suppress HTTP request logs
        logging.getLogger('httpx').setLevel(logging.WARNING)

        # Reduce other verbose libraries
        logging.getLogger('telegram').setLevel(logging.WARNING)
        logging.getLogger('urllib3').setLevel(logging.WARNING)
        logging.getLogger('asyncio').setLevel(logging.WARNING)

        self.logger = logging.getLogger(__name__)
        self.token = token
        self.is_active = True  # Add the active flag

        # Initialize database when bot starts
        self.initialize_database()

        self.admin_commands = [
            BotCommand("start", "Start the bot"),
            BotCommand("stop", "Stop the bot"),
            BotCommand("restart", "Restart the bot"),
            BotCommand("status", "Check bot status"),
            BotCommand("logs", "Get recent error logs"),
            BotCommand("retry", "Retry processing a failed file")
        ]

    def convert_to_khmer_number(self, number):
        """
        Convert Arabic numerals to Khmer numerals

        :param number: Number to convert (can be int or str)
        :return: Khmer numeral representation
        """
        khmer_digits = {
            '0': '០',
            '1': '១',
            '2': '២',
            '3': '៣',
            '4': '៤',
            '5': '៥',
            '6': '៦',
            '7': '៧',
            '8': '៨',
            '9': '៩'
        }
        return ''.join(khmer_digits.get(digit, digit) for digit in str(number))

    def extract_branch_info(self, original_title):
        """
        Extract branch information from the title
        Prioritize ទីស្នាក់ការកណ្តាល over other parenthetical information

        :param original_title: Original title string
        :return: Extracted branch information or empty string
        """
        # First, look for ទីស្នាក់ការកណ្តាល
        location_match = re.search(r'\(([^)]*ទីស្នាក់ការកណ្តាល[^)]*)\)', original_title)
        if location_match:
            return location_match.group(1).strip()

        # If no ទីស្នាក់ការកណ្តាល, look for សាខា
        branch_match = re.search(r'\(([^)]*សាខា[^)]*)\)', original_title)
        if branch_match:
            return branch_match.group(1).strip()

        return ''

    def clean_company_name(self, original_title):
        """
        Clean company name by removing specific parenthetical information

        :param original_title: Original title string
        :return: Cleaned company name
        """
        # First, extract the part after 'របស់'
        company_match = re.search(r'របស់\s*([^(]+(?:\s*\([^)]*\))?)', original_title)

        if company_match:
            company_name = company_match.group(1).strip()

            # Remove everything from '(សាខា' or '(ទីស្នាក់ការកណ្តាល' to the end
            # Handle variations like (សាខា...), (សាខាក្រុមហ៊ុន...), (ទីស្នាក់ការកណ្តាល...)
            company_name = re.sub(r'\s*\((?:សាខា[^)]*|ទីស្នាក់ការកណ្តាល[^)]*)\)', '', company_name).strip()

            return company_name

        # Fallback if no match found
        return original_title.split('របស់')[-1].strip()

    def initialize_database(self):
        """
        Initialize database with all necessary tables and indexes
        """
        try:
            # Check if database file exists
            if not os.path.exists(DATABASE_PATH):
                self.logger.info(f"Database not found. Creating new database at {DATABASE_PATH}")
                # Create a new database connection, which will create the file
                conn = sqlite3.connect(DATABASE_PATH)
                conn.close()
            # Ensure tables are created
            with sqlite3.connect(DATABASE_PATH) as conn:
                cursor = conn.cursor()
                # Create tax_declaration table
                cursor.execute('''
                CREATE TABLE IF NOT EXISTS tax_declaration (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    date DATE NOT NULL,
                    invoice_number TEXT NOT NULL,
                    credit_notification_letter_number TEXT,
                    buyer_type TEXT,
                    tax_registration_id TEXT,
                    buyer_name TEXT,
                    total_invoice_amount DECIMAL(15,2),
                    amount_exclude_vat DECIMAL(15,2),
                    non_vat_sales DECIMAL(15,2),
                    vat_zero_rate DECIMAL(15,2),
                    vat_local_sale DECIMAL(15,2),
                    vat_export DECIMAL(15,2),
                    vat_local_sale_state_burden DECIMAL(15,2),
                    vat_withheld_by_national_treasury DECIMAL(15,2),
                    plt DECIMAL(15,2),
                    special_tax_on_goods DECIMAL(15,2),
                    special_tax_on_services DECIMAL(15,2),
                    accommodation_tax DECIMAL(15,2),
                    income_tax_redemption_rate DECIMAL(15,2),
                    notes TEXT,
                    description TEXT,
                    tax_declaration_status TEXT,
                    telegram_username TEXT,
                    import_timestamp DATETIME,
                    file_tin TEXT
                )
                ''')
                # Create import_sessions table
                cursor.execute('''
                CREATE TABLE IF NOT EXISTS import_sessions (
                    session_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    start_timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
                    end_timestamp DATETIME,
                    total_files_processed INTEGER DEFAULT 0,
                    total_rows_imported INTEGER DEFAULT 0,
                    status TEXT CHECK(status IN ('Started', 'Completed', 'Failed'))
                )
                ''')
                # Create indexes for performance
                index_queries = [
                    'CREATE INDEX IF NOT EXISTS idx_sale_date ON tax_declaration(date)',
                    'CREATE INDEX IF NOT EXISTS idx_invoice_number ON tax_declaration(invoice_number)',
                    'CREATE INDEX IF NOT EXISTS idx_tax_registration_id ON tax_declaration(tax_registration_id)',
                    'CREATE INDEX IF NOT EXISTS idx_buyer_type ON tax_declaration(buyer_type)',
                    'CREATE INDEX IF NOT EXISTS idx_tax_declaration_status ON tax_declaration(tax_declaration_status)'
                ]
                for query in index_queries:
                    try:
                        cursor.execute(query)
                    except sqlite3.OperationalError as e:
                        self.logger.warning(f"Index creation warning: {e}")
                # Commit changes
                conn.commit()
        except Exception as e:
            self.logger.error(f"Database initialization failed: {e}")

    def is_compressed_file(self, filename):
        """
        Check if file is a supported compressed file
        :param filename: Name of the file
        :return: Boolean indicating if file is compressed
        """
        return any(filename.lower().endswith(ext) for ext in ['.zip', '.rar'])

    def extract_archive(self, file_path):
        """
        Extract compressed archive using patool as the primary method
        :param file_path: Path to compressed file
        :return: Path to extracted directory
        """
        try:
            filename = os.path.basename(file_path)
            extract_to = os.path.join(EXTRACT_DIR, os.path.splitext(filename)[0])
            os.makedirs(extract_to, exist_ok=True)

            # Use patool directly as the primary extraction method
            self.logger.info(f"Extracting {filename} using patool")
            patoolib.extract_archive(file_path, outdir=extract_to)
            self.logger.info(f"Successfully extracted {filename} to {extract_to}")

            return extract_to
        except Exception as e:
            self.logger.error(f"Extraction error for {file_path}: {e}")
            # If patool fails, provide a more specific error message
            if "Cannot find working tool" in str(e):
                self.logger.error("Extraction failed: Required extraction tools not found. Please install WinRAR, 7-Zip, or other archive utilities.")
            return None

    def get_month_name(self, month_num):
        months = {
            1: 'Jan', 2: 'Feb', 3: 'Mar', 4: 'Apr', 5: 'May', 6: 'Jun',
            7: 'Jul', 8: 'Aug', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dec'
        }
        return months.get(int(month_num), str(month_num))
    
    def format_month_list(self, month_nums):
        """
        Convert list of month numbers [1, 2, 3, 6] to "Jan-Mar, Jun"
        """
        if not month_nums: return ""
        
        # Sort and remove duplicates
        nums = sorted(list(set(month_nums)))
        
        ranges = []
        range_start = nums[0]
        prev = nums[0]
        
        for curr in nums[1:]:
            if curr == prev + 1:
                prev = curr
            else:
                # End of a range
                if range_start == prev:
                    ranges.append(self.get_month_name(range_start))
                else:
                    ranges.append(f"{self.get_month_name(range_start)}-{self.get_month_name(prev)}")
                range_start = curr
                prev = curr
        
        # Add the last range
        if range_start == prev:
            ranges.append(self.get_month_name(range_start))
        else:
            ranges.append(f"{self.get_month_name(range_start)}-{self.get_month_name(prev)}")
            
        return " , ".join(ranges)

    def combine_excel_files_by_year(self, tin_folder_path):
        try:
            # Get all Excel files
            excel_files = []
            for f in os.listdir(tin_folder_path):
                if any([
                    f.startswith('~$'), f.startswith('._'), 'backup' in f.lower(),
                    not (f.endswith('.xlsx') or f.endswith('.xls')), not 'SALE' in f.upper()
                ]):
                    continue
                if f == f"{os.path.basename(tin_folder_path).replace('TEMP_', '')}.xlsx":
                    continue
                excel_files.append(f)
            if not excel_files: return {}

            # Parse filenames
            parsed_files = []
            for file in excel_files:
                match = re.search(r'_(\d{2})_(\d{4})(?:\s*\(\d+\))*\.', file)
                if match:
                    month, year = match.groups()
                    parsed_files.append({
                        'year': int(year),
                        'month': int(month),
                        'path': os.path.join(tin_folder_path, file)
                    })
            if not parsed_files: return {}

            # Sort chronologically
            parsed_files.sort(key=lambda x: (x['year'], x['month']))
            
            tin = os.path.basename(tin_folder_path).replace("TEMP_", "")
            combined_filename = f"{tin}.xlsx"
            combined_filepath = os.path.join(tin_folder_path, combined_filename)
            
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            
            min_year = parsed_files[0]['year']
            max_year = parsed_files[-1]['year']
            
            if tin in TINS_SPLIT_BY_YEAR:
                chunk_size = 1
                self.logger.info(f"Special TIN detected ({tin}): Splitting sheets by single year.")
            else:
                chunk_size = 4

            # Dictionary to track counts per year
            # Structure: { 2024: 150, 2025: 200 }
            year_counts = {} 
            
            current_start_year = min_year
            
            while current_start_year <= max_year:
                current_end_year = current_start_year + chunk_size - 1
                chunk_files = [f for f in parsed_files if current_start_year <= f['year'] <= current_end_year]
                
                if chunk_files:
                    chunk_start = chunk_files[0]
                    chunk_end = chunk_files[-1]
                    sheet_name = f"{chunk_start['month']:02d}-{chunk_start['year']} - {chunk_end['month']:02d}-{chunk_end['year']}"
                    
                    ws = wb.create_sheet(title=sheet_name)
                    
                    # We need to process files one by one to count rows per year accurately
                    # Group files in this chunk by year to count them
                    files_in_chunk_by_year = {}
                    for f in chunk_files:
                        y = f['year']
                        if y not in files_in_chunk_by_year: files_in_chunk_by_year[y] = []
                        files_in_chunk_by_year[y].append((str(f['month']), f['path']))
                    
                    # Fill sheet and count
                    # Note: _fill_sheet_with_all_data returns total rows for the whole chunk
                    # To get per-year counts, we need to rely on the fact that _fill_sheet_with_all_data
                    # processes all files. We will modify the logic slightly:
                    # We will pass ALL files to fill_sheet, but we need to know the count per year.
                    # Since modifying _fill_sheet is complex, let's estimate or count separately.
                    # BETTER APPROACH: Let's just count the rows in the dataframes before sending to fill_sheet.
                    
                    sorted_file_paths = [(str(f['month']), f['path']) for f in chunk_files]
                    self._fill_sheet_with_all_data(ws, tin, sorted_file_paths, chunk_start['year'])
                    
                    # Count rows per year for statistics
                    for f in chunk_files:
                        try:
                            df = pd.read_excel(f['path'], skiprows=3, dtype={2: str})
                            # Filter empty rows
                            valid_rows = df[df.iloc[:, 1].notna() | df.iloc[:, 0].notna()].shape[0]
                            
                            y = f['year']
                            if y not in year_counts: year_counts[y] = 0
                            year_counts[y] += valid_rows
                        except:
                            pass

                current_start_year += chunk_size
            
            wb.save(combined_filepath)
            
            # Archive files
            files_by_year = {}
            for f in parsed_files:
                y = str(f['year'])
                if y not in files_by_year: files_by_year[y] = []
                files_by_year[y].append((str(f['month']), f['path']))
            for year, files in files_by_year.items():
                self._archive_original_files(tin, year, files)
            
            # --- PREPARE DETAILED STATS ---
            # We need to return a list of objects, one for each year
            detailed_stats = []
            
            # Group months by year for description
            years_months_map = {}
            for f in parsed_files:
                y = f['year']
                if y not in years_months_map: years_months_map[y] = []
                years_months_map[y].append(f['month'])
            
            for year in sorted(years_months_map.keys()):
                months_str = self.format_month_list(years_months_map[year])
                count = year_counts.get(year, 0)
                
                detailed_stats.append({
                    "tin": tin,
                    "month_desc": months_str,
                    "year": year,
                    "count": count
                })
                
            return {
                "combined": combined_filepath,
                "stats": detailed_stats # Returns a LIST of dicts
            }
        except Exception as e:
            self.logger.error(f"Error combining Excel files: {e}")
            self.logger.error(traceback.format_exc())
            return {}

    def _fill_sheet_with_all_data(self, ws, tin, files, start_year):
        """
        Helper to fill the worksheet with data from ALL files provided.
        Returns the number of data rows added.
        """
        # Define styles
        khmer_font = openpyxl.styles.Font(name="Khmer OS Muol Light", size=12)
        khmer_font_normal = openpyxl.styles.Font(name="Khmer OS Siemreap", size=11)
        khmer_font_header = openpyxl.styles.Font(name="Khmer OS Siemreap", bold=True, size=11)
        thin_border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'), bottom=openpyxl.styles.Side(style='thin'))
        light_border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin', color="DDDDDD"), right=openpyxl.styles.Side(style='thin', color="DDDDDD"), top=openpyxl.styles.Side(style='thin', color="DDDDDD"), bottom=openpyxl.styles.Side(style='thin', color="DDDDDD"))
        header_fill = openpyxl.styles.PatternFill(start_color="B0DBBC", end_color="B0DBBC", fill_type="solid")
        alt_row_fill = openpyxl.styles.PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")
        
        # Get headers from first file
        first_file = files[0][1]
        wb_temp = openpyxl.load_workbook(first_file, read_only=True)
        ws_temp = wb_temp.active
        original_title = ws_temp.cell(row=1, column=1).value
        
        # Read header rows (Row 2 and 3 in original file)
        header_row1 = [cell.value for cell in ws_temp[2]]
        header_row2 = [cell.value for cell in ws_temp[3]]
        wb_temp.close()
        
        # Prepare Title
        clean_company_name = self.clean_company_name(original_title) if original_title else tin
        title = f"បញ្ជីទិន្នានុប្បវត្តិលក់ របស់ {clean_company_name}"
        
        # --- WRITE HEADERS ---
        ws.cell(row=1, column=1, value=title).font = khmer_font
        
        # 1. Write "Year" (Column A)
        ws.cell(row=2, column=1, value="ឆ្នាំ")
        ws.merge_cells('A2:A3') 
        
        # 2. Write "No." (Column B)
        ws.cell(row=2, column=2, value=header_row1[0]) 
        ws.merge_cells('B2:B3')

        # 3. Write remaining headers shifted by +1
        # We iterate through the original headers starting from index 1
        for i in range(1, len(header_row1)):
            val1 = header_row1[i]
            val2 = header_row2[i]
            
            target_col = i + 2
            
            ws.cell(row=2, column=target_col, value=val1)
            ws.cell(row=3, column=target_col, value=val2)

        # Write Branch Header at the end
        last_col = len(header_row1) + 2
        ws.cell(row=2, column=last_col, value="ស្នាក់ការ") 
        ws.merge_cells(f'{openpyxl.utils.get_column_letter(last_col)}2:{openpyxl.utils.get_column_letter(last_col)}3')

        # --- FORMAT HEADERS ---
        for row in range(2, 4):
            for col in range(1, last_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = khmer_font_header
                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=False)
                cell.fill = header_fill
                cell.border = thin_border

        # --- APPLY MERGES ---
        # Based on standard template + 1 column shift for Year
        
        # Horizontal Merges (Headers spanning multiple columns)
        # Buyer Info: Original E-G -> New F-H
        ws.merge_cells('F2:H2') 
        # Total Value: Original I-J -> New J-K
        ws.merge_cells('J2:K2') 
        # Non-Taxable: Original L-M -> New M-N
        ws.merge_cells('M2:N2') 
        # VAT: Original O-P -> New P-Q (Wait, check your pic)
        
        # Let's look at your Pic 2 specifically:
        # L: VAT Taxable (Merged L2:M2) -> New M:N
        # N: VAT State Burden (Merged N2:N3? No, usually single) -> New O
        # O: VAT Withheld (Merged O2:O3?) -> New P
        # P: Public Lighting (Merged P2:P3) -> New Q
        
        # Let's stick to the standard logic: Merge if the top cell has text and bottom is None/Same
        # Or hardcode based on the visual:
        
        # 1. Buyer Info (F-H)
        ws.merge_cells('F2:H2')
        # 2. Total Value (J-K)
        ws.merge_cells('J2:K2')
        # 3. Non-Taxable / Exempt (M-N) - Based on Pic 2 (L-M becomes M-N)
        ws.merge_cells('M2:N2') 
        
        # Vertical Merges (Single columns)
        # We need to merge top and bottom for columns that are single attributes
        # C: Date
        # D: Invoice
        # E: Credit Note
        # I: Description
        # L: Export (Original K -> New L)
        # O: VAT State Burden (Original N -> New O)
        # P: Public Lighting (Original O -> New P) <-- THIS IS THE MISSING ONE
        # Q: VAT Withheld (Original P -> New Q)
        # R: Special Tax (Original Q -> New R)
        # ... etc
        
        # List of columns to vertically merge (Shifted +1 from original A,B,C...)
        # Original: A, B, C, D, H, K, N, O, P, Q, R, S, T, U
        # New:      B, C, D, E, I, L, O, P, Q, R, S, T, U, V
        
        vertical_merge_cols = [
            'C', # Date
            'D', # Invoice
            'E', # Credit Note
            'I', # Description
            'L', # Export Value
            'O', # VAT State Burden
            'P', # Public Lighting Tax (The missing one)
            'Q', # VAT Withheld
            'R', # Special Tax Goods
            'S', # Special Tax Services
            'T', # Accommodation Tax
            'U', # Income Tax
            'V', # Other
            'W', # Other
            'X'  # Other
        ]
        
        for col in vertical_merge_cols:
            try:
                # Only merge if within bounds
                if openpyxl.utils.column_index_from_string(col) <= last_col:
                    ws.merge_cells(f'{col}2:{col}3')
            except: pass

        # --- COLLECT DATA ---
        all_data = []
        existing_entries = set()
        
        for _, file_path in files:
            try:
                match = re.search(r'_(\d{2})_(\d{4})', os.path.basename(file_path))
                file_year = int(match.group(2)) if match else 0
                wb_branch = openpyxl.load_workbook(file_path, read_only=True)
                branch_info = self.extract_branch_info(wb_branch.active.cell(row=1, column=1).value)
                wb_branch.close()
                df = pd.read_excel(file_path, skiprows=3, dtype={2: str})
                for _, row in df.iterrows():
                    if pd.isna(row.iloc[1]) and pd.isna(row.iloc[0]): continue
                    
                    key = (str(row.iloc[2]), str(row.iloc[1]), str(row.iloc[11]))
                    if key in existing_entries: continue
                    existing_entries.add(key)
                    row_data = row.tolist()
                    row_data.append(branch_info)
                    
                    all_data.append({
                        'year': file_year,
                        'data': row_data
                    })
            except Exception as e:
                self.logger.error(f"Error reading {file_path}: {e}")
        
        # --- WRITE DATA ---
        current_row_num = 4 
        previous_year = None
        row_counter = 1 
        for item in all_data:
            row_year = item['year']
            row_data = item['data']
            if previous_year is not None and row_year != previous_year:
                current_row_num += 1 
            
            previous_year = row_year
            # 1. Write YEAR in Column 1 (A)
            cell_year = ws.cell(row=current_row_num, column=1, value=row_year)
            cell_year.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            
            # 2. Write Row ID in Column 2 (B)
            ws.cell(row=current_row_num, column=2, value=row_counter) 
            
            # 3. Write Data Columns starting from Column 3 (C)
            for col_idx, value in enumerate(row_data[1:], 3):
                cell = ws.cell(row=current_row_num, column=col_idx)
                
                if col_idx == 3: # Date Column (C)
                    try:
                        if isinstance(value, str):
                            date_obj = datetime.strptime(value, '%d-%m-%Y')
                            cell.value = date_obj
                        elif isinstance(value, pd.Timestamp):
                            cell.value = value.to_pydatetime()
                        else:
                            cell.value = value
                        cell.number_format = 'dd-mm-yyyy'
                    except (ValueError, TypeError):
                        cell.value = value
                else:
                    cell.value = value
            
            # Formatting
            for col_idx in range(1, len(row_data) + 2):
                cell = ws.cell(row=current_row_num, column=col_idx)
                cell.font = khmer_font_normal
                cell.border = light_border
                
                if row_counter % 2 == 0: 
                    cell.fill = alt_row_fill
            
            current_row_num += 1
            row_counter += 1
            
        # --- COLUMN WIDTHS ---
        ws.column_dimensions['A'].width = 8  # Year
        ws.column_dimensions['B'].width = 6  # No
        ws.column_dimensions['C'].width = 12 # Date
        ws.column_dimensions['D'].width = 15 # Invoice
        
        # Auto-width for others
        for col in range(5, last_col + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18
            
        return len(all_data)

    def _process_year_files(self, tin, year, files, tin_folder_path):
        """
        Process files for a specific year
        """
        try:
            # Define fonts and styles
            khmer_font = openpyxl.styles.Font(name="Khmer OS Muol Light", size=12)
            khmer_font_normal = openpyxl.styles.Font(name="Khmer OS Siemreap", size=11)
            khmer_font_header = openpyxl.styles.Font(name="Khmer OS Siemreap", bold=True, size=11)

            # Define border styles
            thin_border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style='thin'),
                right=openpyxl.styles.Side(style='thin'),
                top=openpyxl.styles.Side(style='thin'),
                bottom=openpyxl.styles.Side(style='thin')
            )
            light_border = openpyxl.styles.Border(
                left=openpyxl.styles.Side(style='thin', color="DDDDDD"),
                right=openpyxl.styles.Side(style='thin', color="DDDDDD"),
                top=openpyxl.styles.Side(style='thin', color="DDDDDD"),
                bottom=openpyxl.styles.Side(style='thin', color="DDDDDD")
            )

            # Define fill styles
            header_fill = openpyxl.styles.PatternFill(start_color="B0DBBC", end_color="B0DBBC", fill_type="solid")
            alt_row_fill = openpyxl.styles.PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid")

            combined_filename = f"{tin}_SALE_{year}.xlsx"
            combined_filepath = os.path.join(tin_folder_path, combined_filename)

            # Check if a combined file already exists
            file_already_existed = os.path.exists(combined_filepath)

            # Collect all data first
            all_data = []
            branch_info_by_file = {}

            # Process the first file to get headers
            first_file = files[0][1]
            first_wb = None
            try:
                first_wb = openpyxl.load_workbook(first_file, read_only=True)
                first_ws = first_wb.active

                # Get title and headers
                original_title = first_ws.cell(row=1, column=1).value
                header_row1 = [cell.value for cell in first_ws[2]]
                header_row2 = [cell.value for cell in first_ws[3]]

                # Clean the company name
                clean_company_name = self.clean_company_name(original_title) if original_title else tin

                # Convert year to Khmer
                khmer_year = self.convert_to_khmer_number(year)

                # Create new title format
                title = f"បញ្ជីទិន្នានុប្បវត្តិលក់ ប្រចាំឆ្នាំ {khmer_year} របស់ {clean_company_name}"

            finally:
                # Ensure workbook is closed
                if first_wb:
                    first_wb.close()

            # Set to track existing entries
            existing_entries = set()

            # If file exists, load existing entries to avoid duplicates
            if file_already_existed:
                self.logger.info(f"Found existing combined file for {year}: {combined_filename}")
                try:
                    # Use pandas to efficiently read existing data
                    existing_df = pd.read_excel(combined_filepath, skiprows=3, dtype={2: str})

                    # Create unique keys from existing data
                    for _, row in existing_df.iterrows():
                        invoice_num = str(row.iloc[2]) if pd.notna(row.iloc[2]) else ''
                        date_val = str(row.iloc[1]) if pd.notna(row.iloc[1]) else ''
                        amount = str(row.iloc[11]) if pd.notna(row.iloc[11]) else ''
                        entry_key = (invoice_num, date_val, amount)
                        existing_entries.add(entry_key)

                    self.logger.info(f"Loaded {len(existing_entries)} existing entries from {combined_filename}")

                    # Also collect existing data
                    all_data.extend([
                        row.tolist() for row in existing_df.values
                    ])
                except Exception as e:
                    self.logger.error(f"Error reading existing file {combined_filename}: {e}")
                    # If error reading existing file, start fresh
                    file_already_existed = False

            # Process each file SEQUENTIALLY
            total_new_rows = 0
            for idx, (month, file_path) in enumerate(files):
                self.logger.info(f"Processing file {idx+1}/{len(files)}: {os.path.basename(file_path)}")
                try:
                    # Extract branch info
                    wb = None
                    try:
                        wb = openpyxl.load_workbook(file_path, read_only=True)
                        ws = wb.active
                        file_branch_info = self.extract_branch_info(ws.cell(row=1, column=1).value)
                        branch_info_by_file[file_path] = file_branch_info
                    finally:
                        # Ensure workbook is closed
                        if wb:
                            wb.close()

                    # Use pandas for faster data reading
                    df = pd.read_excel(file_path, skiprows=3, dtype={2: str})  # Skip header rows

                    # Process each row
                    new_rows = []
                    for _, row in df.iterrows():
                        # Skip empty rows
                        if pd.isna(row.iloc[1]) and pd.isna(row.iloc[0]):
                            continue

                        # Create unique key
                        invoice_num = str(row.iloc[2]) if pd.notna(row.iloc[2]) else ''
                        date_val = str(row.iloc[1]) if pd.notna(row.iloc[1]) else ''
                        amount = str(row.iloc[11]) if pd.notna(row.iloc[11]) else ''
                        entry_key = (invoice_num, date_val, amount)

                        # Skip if already exists
                        if entry_key in existing_entries:
                            continue

                        # Add to existing entries
                        existing_entries.add(entry_key)

                        # Add branch info to row
                        row_data = row.tolist()
                        row_data.append(file_branch_info)

                        # Add to new rows WITHOUT row number
                        new_rows.append(row_data)
                        total_new_rows += 1

                    # Add to all data
                    all_data.extend(new_rows)
                    self.logger.info(f"Added {len(new_rows)} rows from {os.path.basename(file_path)}")
                except Exception as e:
                    self.logger.error(f"Error processing file {file_path}: {e}")

            # If no new data, return early
            if total_new_rows == 0 and file_already_existed:
                self.logger.info(f"No new data to add to {combined_filename}")

                # Now that we've processed all files, archive them
                self._archive_original_files(tin, year, files)

                return combined_filepath


            # Create a new workbook
            wb = openpyxl.Workbook()
            ws = wb.active

            # Add title and headers
            ws.cell(row=1, column=1, value=title)
            ws.cell(row=1, column=1).font = khmer_font

            # Add headers
            for col_idx, value in enumerate(header_row1, 1):
                ws.cell(row=2, column=col_idx, value=value)

            for col_idx, value in enumerate(header_row2, 1):
                ws.cell(row=3, column=col_idx, value=value)

            # Add "ស្នាក់ការ" header
            ws.cell(row=2, column=len(header_row1) + 1, value="ស្នាក់ការ")

            # Format headers
            for row in range(2, 4):
                for col in range(1, len(header_row1) + 2):
                    cell = ws.cell(row=row, column=col)
                    cell.font = khmer_font_header
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=False)
                    cell.fill = header_fill
                    cell.border = thin_border

            # Merge cells as needed
            ws.merge_cells('E2:G2')
            ws.merge_cells('I2:J2')
            ws.merge_cells('L2:M2')

            # List of columns to merge
            columns_to_merge = [
                'A', 'B', 'C', 'D', 'H', 'K',
                'O', 'P', 'Q', 'R', 'S', 'T',
                'U', 'V', 'W', 'X'
            ]

            # Merge specified columns
            for col in columns_to_merge:
                try:
                    col_idx = openpyxl.utils.column_index_from_string(col)
                    if col_idx <= ws.max_column:
                        ws.merge_cells(f'{col}2:{col}3')
                except:
                    # Skip if column doesn't exist or can't be merged
                    pass

            # Write data in chunks to reduce memory usage
            chunk_size = 1000
            for i in range(0, len(all_data), chunk_size):
                chunk = all_data[i:i+chunk_size]

                for row_idx, row_data in enumerate(chunk, i + 4):  # Start at row 4 (after headers)

                    row_number = row_idx - 3

                    # Write row number
                    ws.cell(row=row_idx, column=1, value=row_data[0])

                    # Write data
                    for col_idx, value in enumerate(row_data[1:], 2):
                        ws.cell(row=row_idx, column=col_idx, value=value)

                    # Write row number in column A
                    ws.cell(row=row_idx, column=1, value=row_number)

                    # Apply formatting
                    for col_idx in range(1, len(row_data) + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.font = khmer_font_normal
                        cell.border = light_border

                        # Alternate row colors
                        if row_data[0] % 2 == 0:
                            cell.fill = alt_row_fill

            # Set column widths
            ws.column_dimensions['A'].width = 6  # Narrow width for column A

            # Set width for the branch column
            branch_col = openpyxl.utils.get_column_letter(len(header_row1) + 1)
            ws.column_dimensions[branch_col].width = 21.29  # Wider for Khmer text

            # Auto-adjust other column widths
            for col in range(2, len(header_row1) + 1):
                col_letter = openpyxl.utils.get_column_letter(col)
                # Set reasonable default widths based on content type
                if col in [2, 3]:  # Date and invoice columns
                    ws.column_dimensions[col_letter].width = 21.29
                elif col in [7, 8, 9, 10, 11, 12]:  # Numeric columns
                    ws.column_dimensions[col_letter].width = 21.29
                else:
                    ws.column_dimensions[col_letter].width = 21.29

            # Check if we need to split into multiple sheets (over 500K rows)
            MAX_ROWS_PER_SHEET = 500000
            if len(all_data) > MAX_ROWS_PER_SHEET:
                self.logger.info(f"File has {len(all_data)} rows, splitting into multiple sheets")

                # Calculate number of sheets needed
                num_sheets = (len(all_data) + MAX_ROWS_PER_SHEET - 1) // MAX_ROWS_PER_SHEET

                # Create additional sheets
                for sheet_idx in range(2, num_sheets + 1):
                    sheet_name = f"Sheet{sheet_idx}"
                    wb.create_sheet(sheet_name)
                    new_sheet = wb[sheet_name]

                    # Copy header rows to new sheet
                    for row_idx in range(1, 4):
                        for col_idx in range(1, ws.max_column + 1):
                            new_sheet.cell(row=row_idx, column=col_idx,
                                        value=ws.cell(row=row_idx, column=col_idx).value)
                            # Copy formatting
                            source_cell = ws.cell(row=row_idx, column=col_idx)
                            target_cell = new_sheet.cell(row=row_idx, column=col_idx)

                            # Copy font, alignment, fill, border
                            if row_idx == 1:  # Title row
                                target_cell.font = khmer_font
                                target_cell.alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center')
                            else:  # Header rows 2-3
                                target_cell.font = khmer_font_header
                                target_cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)
                                target_cell.fill = header_fill
                                target_cell.border = thin_border

                    # Copy merged cell ranges for headers
                    for merged_range in ws.merged_cells.ranges:
                        # Only copy merged ranges in the header rows (1-3)
                        if merged_range.min_row <= 3 and merged_range.max_row <= 3:
                            new_sheet.merge_cells(str(merged_range))

                    # Copy column dimensions
                    for col_letter, dimension in ws.column_dimensions.items():
                        new_sheet.column_dimensions[col_letter].width = dimension.width

                # Write data to appropriate sheets
                for idx, row_data in enumerate(all_data):
                    sheet_idx = idx // MAX_ROWS_PER_SHEET + 1
                    row_in_sheet = idx % MAX_ROWS_PER_SHEET + 4  # +4 for header rows

                    # Get the correct sheet
                    if sheet_idx == 1:
                        sheet = ws
                    else:
                        sheet = wb[f"Sheet{sheet_idx}"]

                    # Update row counter (column A)
                    row_data[0] = (idx % MAX_ROWS_PER_SHEET) + 1

                    # Write row data to sheet
                    for col_idx, value in enumerate(row_data):
                        sheet.cell(row=row_in_sheet, column=col_idx+1, value=value)

                        # Apply formatting
                        cell = sheet.cell(row=row_in_sheet, column=col_idx+1)
                        cell.font = khmer_font_normal
                        cell.border = light_border

                        if row_data[0] % 2 == 0:  # Use the row counter for alternating colors
                            cell.fill = alt_row_fill

            # Save the workbook
            wb.save(combined_filepath)

            # Log appropriate message
            if file_already_existed:
                self.logger.info(f"Updated existing combined file: {combined_filename} with {total_new_rows} new rows")
            else:
                self.logger.info(f"Created new combined file: {combined_filename}")

            # IMPORTANT: Only archive files AFTER the combined file is saved
            self._archive_original_files(tin, year, files)

            return combined_filepath

        except Exception as e:
            self.logger.error(f"Error processing year {year}: {e}")
            self.logger.error(traceback.format_exc())
            return None

    def _archive_original_files(self, tin, year, files):
        """
        Archive original files to year-specific folders
        """
        try:
            # Create archive folder
            tin_archive_folder = os.path.join(ARCHIVE_DIR, tin, year)
            os.makedirs(tin_archive_folder, exist_ok=True)

            # Process files SEQUENTIALLY to avoid conflicts
            for _, file_path in files:
                try:
                    self._archive_single_file(file_path, tin_archive_folder)
                except Exception as e:
                    self.logger.error(f"Error archiving file {file_path}: {e}")
                    # Continue with other files even if one fails
                    continue

        except Exception as e:
            self.logger.error(f"Error archiving files: {e}")

    def _archive_single_file(self, file_path, archive_folder):
        """
        Archive a single file with version control
        """
        try:
            # Check if file exists before trying to archive it
            if not os.path.exists(file_path):
                self.logger.warning(f"File not found for archiving: {file_path}")
                return

            filename = os.path.basename(file_path)
            base_name, extension = os.path.splitext(filename)

            # Check if file with same name already exists in archive
            archive_path = os.path.join(archive_folder, filename)
            if os.path.exists(archive_path):
                # Compare file content to see if they're different
                existing_file_hash = self.get_file_hash(archive_path)
                new_file_hash = self.get_file_hash(file_path)

                if existing_file_hash != new_file_hash:
                    # Files are different - create a versioned filename
                    # Get current timestamp for versioning
                    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                    versioned_filename = f"{base_name}_v{timestamp}{extension}"
                    archive_path = os.path.join(archive_folder, versioned_filename)
                    self.logger.info(f"File with same name but different content detected. Archiving as: {versioned_filename}")
                else:
                    # Files are identical - no need to keep duplicate
                    self.logger.info(f"Identical file already exists in archive: {filename}")
                    # Delete the duplicate file since we already have it
                    os.remove(file_path)
                    return

            # Add retry logic for file operations
            max_retries = 3
            retry_delay = 1  # seconds

            for attempt in range(max_retries):
                try:
                    # Move file to archive
                    shutil.move(file_path, archive_path)
                    self.logger.info(f"Moved {filename} to central archive folder: {archive_folder}")
                    return
                except PermissionError as e:
                    if attempt < max_retries - 1:
                        self.logger.warning(f"Permission error moving file {filename}, retrying in {retry_delay} seconds: {e}")
                        time.sleep(retry_delay)
                        retry_delay *= 2  # Exponential backoff
                    else:
                        # On final attempt, try to copy instead of move
                        try:
                            shutil.copy2(file_path, archive_path)
                            self.logger.info(f"Copied {filename} to central archive folder: {archive_folder}")
                            # Try to remove original after successful copy
                            try:
                                os.remove(file_path)
                                self.logger.info(f"Removed original file after copying: {filename}")
                            except:
                                self.logger.warning(f"Could not remove original file after copying: {filename}")
                            return
                        except Exception as copy_error:
                            raise copy_error
                except Exception as e:
                    if attempt < max_retries - 1:
                        self.logger.warning(f"Error moving file {filename}, retrying in {retry_delay} seconds: {e}")
                        time.sleep(retry_delay)
                        retry_delay *= 2  # Exponential backoff
                    else:
                        raise e

        except Exception as e:
            self.logger.error(f"Error archiving file {file_path}: {e}")
            # Re-raise to be handled by the caller
            raise

    def organize_excel_files(self, extract_dir):
        """
        Organize Excel files by TIN and combine them directly in the Bot Extract directory.
        Creates ONE file per TIN containing ALL years in ONE sheet.
        Generates a summary report at the end.
        """
        try:
            bot_extracts_dir = os.path.dirname(extract_dir)
            
            # Find all Excel files
            excel_files = []
            for root, dirs, files in os.walk(extract_dir):
                excel_files.extend([
                    os.path.join(root, f) for f in files
                    if f.endswith(('.xlsx', '.xls')) and 'SALE' in f.upper()
                ])
            # Group by TIN
            files_by_tin = {}
            for file_path in excel_files:
                file_name = os.path.basename(file_path)
                tin = file_name.split('_')[0]
                if tin not in files_by_tin: files_by_tin[tin] = []
                files_by_tin[tin].append(file_path)
            
            processed_tins = []
            summary_stats = []  # List to hold stats for the report
            
            for tin, file_paths in files_by_tin.items():
                # Create temp dir
                temp_tin_dir = os.path.join(extract_dir, "TEMP_" + tin)
                os.makedirs(temp_tin_dir, exist_ok=True)
                
                # Move files to temp dir
                for src_path in file_paths:
                    shutil.move(src_path, os.path.join(temp_tin_dir, os.path.basename(src_path)))
                
                # Combine files
                result = self.combine_excel_files_by_year(temp_tin_dir)
                
                if result and "combined" in result:
                    processed_tins.append(tin)
                    
                    # Collect stats if available
                    if "stats" in result:
                        summary_stats.append(result["stats"])
                    
                    combined_path = result["combined"]
                    if os.path.exists(combined_path):
                        final_filename = os.path.basename(combined_path)
                        final_filename = final_filename.replace("TEMP_", "")
                        final_dest = os.path.join(bot_extracts_dir, final_filename)
                        
                        if os.path.exists(final_dest):
                            os.remove(final_dest)
                            
                        shutil.move(combined_path, final_dest)
                        self.logger.info(f"Moved master file to: {final_dest}")
            
            # Generate Summary Report
            if summary_stats:
                self.generate_summary_report(summary_stats, bot_extracts_dir)

            # Cleanup
            try:
                if os.path.exists(extract_dir):
                    shutil.rmtree(extract_dir)
            except Exception as e:
                self.logger.error(f"Error removing extracted directory: {e}")
            return processed_tins
            
        except Exception as e:
            self.logger.error(f"Error organizing Excel files: {e}")
            traceback.print_exc()
            return []

    def get_file_hash(self, file_path):
        """
        Calculate a hash of file contents to detect if files are different
        :param file_path: Path to the file
        :return: Hash string representing file contents
        """
        import hashlib

        # Use a buffer to handle large files efficiently
        hash_md5 = hashlib.md5()
        with open(file_path, "rb") as f:
            # Read file in chunks of 4K
            for chunk in iter(lambda: f.read(4096), b""):
                hash_md5.update(chunk)
        return hash_md5.hexdigest()

    def generate_summary_report(self, summary_data_list, output_dir):
        """
        Update the master summary Excel file.
        Columns: No | TIN | Month | Year | Total Transactions
        Merges No and TIN columns.
        Alternates background colors by TIN group.
        """
        if not summary_data_list:
            return None

        try:
            filename = "Import_Summary.xlsx"
            filepath = os.path.join(output_dir, filename)
            
            # 1. Load existing data if available
            existing_data = []
            if os.path.exists(filepath):
                try:
                    existing_df = pd.read_excel(filepath, sheet_name='Summary')
                    
                    # --- FIX FOR NAN / GHOST ROWS ---
                    # 1. Convert TIN to string to handle mixed types
                    existing_df['TIN'] = existing_df['TIN'].astype(str)
                    
                    # 2. Filter out specific bad values
                    # Remove 'nan', 'GRAND TOTAL', 'Total TINs', and empty strings
                    existing_df = existing_df[
                        (existing_df['TIN'] != 'nan') & 
                        (existing_df['TIN'] != 'GRAND TOTAL') & 
                        (~existing_df['TIN'].str.contains('Total TINs', na=False)) &
                        (existing_df['TIN'].str.strip() != '')
                    ]
                    
                    # Convert to list of dicts
                    existing_data = existing_df.to_dict('records')
                except Exception as e:
                    self.logger.warning(f"Error reading existing summary: {e}")
                    # If reading fails, start fresh to avoid corruption
                    existing_data = []

            # 2. Process new data (Flatten list of lists)
            new_flat_data = []
            for item in summary_data_list:
                if isinstance(item, list):
                    new_flat_data.extend(item)
                else:
                    new_flat_data.append(item)
            
            # 3. Merge New Data into Existing Data
            # Get list of TINs being updated
            updated_tins = set(str(item['tin']) for item in new_flat_data)
            
            # Keep rows where TIN is NOT in the updated list
            # Ensure we compare strings
            final_data = [row for row in existing_data if str(row.get('TIN')) not in updated_tins]
            
            # Add new rows
            for item in new_flat_data:
                final_data.append({
                    'TIN': str(item['tin']), # Ensure string
                    'Month': item['month_desc'],
                    'Year': item['year'],
                    'Total Transactions': item['count']
                })
            
            # 4. Create DataFrame
            df = pd.DataFrame(final_data)
            if df.empty: return None
            
            # Remove 'No' column if it exists
            if 'No' in df.columns:
                df = df.drop(columns=['No'])

            # Sort by TIN, then Year
            df['TIN'] = df['TIN'].astype(str)
            df = df.sort_values(by=['TIN', 'Year'])
            
            # Add 'No' column placeholder
            df.insert(0, 'No', 0) 
            
            # --- CALCULATE TOTALS ---
            total_transactions = df['Total Transactions'].sum()
            unique_tins_count = df['TIN'].nunique()

            # --- WRITE TO EXCEL ---
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Summary')
                
                workbook = writer.book
                worksheet = writer.sheets['Summary']
                
                # --- STYLES ---
                header_font = openpyxl.styles.Font(bold=True, color="FFFFFF")
                header_fill = openpyxl.styles.PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                
                thin_border = openpyxl.styles.Side(style='thin', color="000000")
                border = openpyxl.styles.Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
                
                center_align = openpyxl.styles.Alignment(horizontal='center', vertical='center')
                left_align = openpyxl.styles.Alignment(horizontal='left', vertical='center', wrap_text=True)
                
                fill_white = openpyxl.styles.PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                fill_grey = openpyxl.styles.PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                
                # Apply Header Style
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
                    cell.border = border

                # --- MERGING & FORMATTING LOGIC ---
                current_tin = None
                start_row = 2
                tin_counter = 1
                use_grey_fill = False 
                
                for row in range(2, worksheet.max_row + 2):
                    cell_value = worksheet.cell(row=row, column=2).value if row <= worksheet.max_row else None
                    
                    if cell_value != current_tin:
                        if current_tin is not None:
                            end_row = row - 1
                            
                            # Write 'No'
                            worksheet.cell(row=start_row, column=1).value = tin_counter
                            tin_counter += 1
                            
                            # Merge Cells
                            if end_row > start_row:
                                worksheet.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
                                worksheet.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)
                            
                            # Apply Color
                            current_fill = fill_grey if use_grey_fill else fill_white
                            
                            for r in range(start_row, end_row + 1):
                                for c in range(1, 6):
                                    cell = worksheet.cell(row=r, column=c)
                                    cell.fill = current_fill
                                    cell.border = border
                                    
                                    if c == 3: # Month
                                        cell.alignment = left_align
                                    else:
                                        cell.alignment = center_align
                                    
                                    if c == 5:
                                        cell.number_format = '#,##0'

                            use_grey_fill = not use_grey_fill
                        
                        current_tin = cell_value
                        start_row = row

                # --- ADD GRAND TOTAL ROW ---
                last_row = worksheet.max_row + 1
                
                worksheet.cell(row=last_row, column=1).value = "GRAND TOTAL"
                worksheet.cell(row=last_row, column=2).value = f"Total TINs: {unique_tins_count}"
                worksheet.cell(row=last_row, column=5).value = total_transactions
                
                total_font = openpyxl.styles.Font(bold=True)
                total_fill = openpyxl.styles.PatternFill(start_color="ACB9CA", end_color="ACB9CA", fill_type="solid")
                
                for col in range(1, 6):
                    cell = worksheet.cell(row=last_row, column=col)
                    cell.font = total_font
                    cell.fill = total_fill
                    cell.border = border
                    cell.alignment = center_align
                    if col == 5:
                        cell.number_format = '#,##0'

                # Column Widths
                worksheet.column_dimensions['A'].width = 15
                worksheet.column_dimensions['B'].width = 25
                worksheet.column_dimensions['C'].width = 35
                worksheet.column_dimensions['D'].width = 10
                worksheet.column_dimensions['E'].width = 20
                
            self.logger.info(f"Summary report updated: {filepath}")
            return filepath
            
        except Exception as e:
            self.logger.error(f"Error generating summary report: {e}")
            traceback.print_exc()
            return None

    async def handle_document(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """
        Handle incoming document in Telegram
        """
        try:
            # Determine document source (direct message, channel, or group)
            document = (update.message or update.channel_post).document
            
            # Check file size (document.file_size is in bytes)
            max_size_mb = 20
            max_size_bytes = max_size_mb * 1024 * 1024  # Convert MB to bytes

            if not self.is_active:
                # Only respond if the message is from a user (not a channel)
                if update.message:
                    await update.message.reply_text(
                        "⏹️ Bot is currently stopped. Please wait until an admin restarts it."
                    )
                return

            if document.file_size > max_size_bytes:
                await (update.message or update.channel_post).reply_text(
                    f"❌ File size exceeds {max_size_mb}MB limit. Please upload a smaller file or split your data into multiple files."
                )
                return

            # Get username of the sender
            if update.message:
                if update.message.chat.type in ['group', 'supergroup']:
                    username = f"group_{update.message.chat.id}_{update.message.from_user.username or update.message.from_user.id}"
                else:
                    username = update.message.from_user.username or f"user_{update.message.from_user.id}"
            elif update.channel_post:
                username = f"channel_{update.channel_post.chat.id}"
            else:
                username = "unknown"

            # Check if it's a compressed file
            if not self.is_compressed_file(document.file_name):
                await (update.message or update.channel_post).reply_text("❌ Oops! Please send a ZIP or RAR file.")
                return

            # Download file
            try:
                file = await context.bot.get_file(document.file_id)
                download_path = os.path.join(DOWNLOAD_DIR, document.file_name)
                await file.download_to_drive(download_path)
            except asyncio.TimeoutError:
                # Handle timeout during download
                timeout_message = (
                    f"⚠️ *Connection Timeout Error* ⚠️\n\n"
                    f"Failed to download file: _{document.file_name}_\n\n"
                    f"Please check your internet connection and try again later."
                )
                await (update.message or update.channel_post).reply_text(
                    timeout_message,
                    parse_mode='Markdown'
                )
                # Also notify production group
                await self.notify_production_group(context, timeout_message)
                return

            # Extract file
            extracted_path = self.extract_archive(download_path)
            
            if extracted_path:
                # Trigger import process
                importer = SaleDataImporter(extracted_path)
                
                # Find Excel files in TIN subfolders
                excel_files = [
                    os.path.join(root, f) for root, dirs, files in os.walk(extracted_path)
                    for f in files
                    if f.endswith(('.xlsx', '.xls')) and 'SALE' in f.upper()
                ]
                
                excel_files_count = len(excel_files)
                total_rows = 0
                
                for excel_file in excel_files:
                    # Pass username to the process_excel_file method
                    rows = importer.process_excel_file(excel_file, username)
                    total_rows += rows

                # Organize Excel files into TIN folders and combine by year
                # This now returns a list of TINs but puts files directly in Bot Extract
                organized_tins = self.organize_excel_files(extracted_path)

                # Delete the original compressed file after successful import
                try:
                    os.remove(download_path)
                    self.logger.info(f"Deleted compressed file: {download_path}")
                except Exception as e:
                    self.logger.error(f"Error deleting compressed file: {e}")

                # --- FIX STARTS HERE ---
                # Count combined files directly in EXTRACT_DIR
                combined_files_count = 0
                if os.path.exists(EXTRACT_DIR):
                    for f in os.listdir(EXTRACT_DIR):
                        if f.endswith('.xlsx') and 'SALE' in f:
                            # Check if this file belongs to one of the TINs we just processed
                            for tin in organized_tins:
                                if f.startswith(tin):
                                    combined_files_count += 1
                                    break
                # --- FIX ENDS HERE ---

                # Then modify the success message
                await (update.message or update.channel_post).reply_text(
                    f"✅ File: *{document.file_name}* processed successfully!\n"
                    f"📊 {excel_files_count} Excel files found and processed.\n"
                    f"📈 {total_rows} new rows imported to the database.\n"
                    f"📚 Files have been combined by year.\n"
                    f"📁 {combined_files_count} combined yearly files created in Bot Extract."
                )
            else:
                await (update.message or update.channel_post).reply_text("❌ Extraction failed! Please check your file and try again.")

        except Exception as e:
            self.logger.error(f"Error processing document: {e}")
            # Print traceback to console for debugging
            traceback.print_exc() 
            await (update.message or update.channel_post).reply_text(
                "🆘 Oops! Something went wrong. Please check your file or try again later. \n"
            )

    async def notify_production_group(self, context: ContextTypes.DEFAULT_TYPE, message: str):
        """
        Send a notification to the production group
        :param context: Telegram context
        :param message: Message to send
        """
        try:
            await context.bot.send_message(
                chat_id=self.production_chat_id,
                text=message
            )
        except Exception as e:
            self.logger.error(f"Failed to send notification to production group: {e}")

    async def is_admin(self, update: Update) -> bool:
        """
        Check if the user is an admin
        - Allows commands only from admin group or specific admin users
        """
        user = update.effective_user
        chat_id = update.effective_chat.id

        # Allow commands from admin group
        if chat_id == ADMIN_GROUP_ID:
            return True

        # Allow commands from admin users in any chat
        if user and user.username in ADMIN_USERS:
            return True

    async def cmd_start(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """
        Handle /start command
        """
        # Verify admin privileges (only from admin group or admin users)
        if not await self.is_admin(update):
            await context.bot.send_message(
                chat_id=update.effective_chat.id,
                text="⚠️ You don't have permission to use this command."
            )
            return

        try:
            # Ensure database is properly initialized
            self.initialize_database()

            # Prepare notification message
            admin_username = update.effective_user.username or "Unknown Admin"
            notification_message = f"🟢 Bot started."

            # Send confirmation to DEV group
            await context.bot.send_message(
                chat_id=update.effective_chat.id,
                text="✅ Bot is now running and ready to process files!"
            )

            # Notify production group
            await self.notify_production_group(context, notification_message)

        except Exception as e:
            error_message = f"❌ Error starting bot: {str(e)}"
            self.logger.error(error_message)
            await context.bot.send_message(
                chat_id=update.effective_chat.id,
                text=error_message
            )

    async def cmd_stop(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """
        Handle /stop command
        """
        # Verify admin privileges
        if not await self.is_admin(update):
            await context.bot.send_message(
                chat_id=update.effective_chat.id,
                text="⚠️ You don't have permission to use this command."
            )
            return

        try:
            # Set bot to inactive
            self.is_active = False

            # Prepare notification message
            admin_username = update.effective_user.username or "Unknown Admin"
            notification_message = f"⏹️ Bot stopped."

            # Send confirmation to DEV group
            await context.bot.send_message(
                chat_id=update.effective_chat.id,
                text="⏹️ Bot has been stopped. No new files will be processed."
            )

            # Notify production group
            await self.notify_production_group(context, notification_message)

        except Exception as e:
            error_message = f"❌ Error stopping bot: {str(e)}"
            self.logger.error(error_message)
            await context.bot.send_message(
                chat_id=update.effective_chat.id,
                text=error_message
            )

    async def cmd_restart(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """
        Handle /restart command
        """
        # Verify admin privileges
        if not await self.is_admin(update):
            await context.bot.send_message(
                chat_id=update.effective_chat.id,
                text="⚠️ You don't have permission to use this command."
            )
            return

        try:
            # Prepare notification message
            admin_username = update.effective_user.username or "Unknown Admin"
            notification_message = f"🔄 Bot Service Restarting..."

            # Log the restart attempt
            self.logger.info(f"Hard restart initiated by {admin_username}")

            # Send message to DEV group
            await context.bot.send_message(
                chat_id=update.effective_chat.id,
                text="🔄 Bot is restarting. This will take a moment..."
            )

            # Notify production group before restarting
            await self.notify_production_group(context, notification_message)

            # Create a restart flag file
            restart_flag_file = os.path.join(BASE_DIR, '.restart_flag')
            with open(restart_flag_file, 'w') as f:
                f.write('restart')

            # Get the path to the restart_bot.bat file
            script_dir = os.path.dirname(os.path.abspath(__file__))
            restart_batch = os.path.join(script_dir, "restart_bot.bat")

            # Execute the batch file
            subprocess.Popen(restart_batch, shell=True, creationflags=subprocess.CREATE_NEW_PROCESS_GROUP)

            # Wait a moment to ensure the message is sent
            await asyncio.sleep(2)

            # Exit the current process
            os._exit(0)  # Force exit without cleanup

        except Exception as e:
            error_message = f"❌ Error during hard restart: {str(e)}"
            self.logger.error(error_message)
            self.logger.error(traceback.format_exc())
            await context.bot.send_message(
                chat_id=update.effective_chat.id,
                text=f"{error_message}\n\nBot will continue running."
            )

    async def cmd_retry(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """
        Handle /retry command when replying to a file message
        """
        # Verify admin privileges
        if not await self.is_admin(update):
            await context.bot.send_message(
                chat_id=update.effective_chat.id,
                text="⚠️ You don't have permission to use this command."
            )
            return

        # Check if the command is a reply to a message
        if update.message.reply_to_message is None:
            await context.bot.send_message(
                chat_id=update.effective_chat.id,
                text="❌ Please reply to a file message with /retry command."
            )
            return

        # Check if the replied message contains a document
        replied_message = update.message.reply_to_message
        if not replied_message.document:
            await context.bot.send_message(
                chat_id=update.effective_chat.id,
                text="❌ The message you replied to doesn't contain a file."
            )
            return

        # Get document details
        document = replied_message.document
        file_name = document.file_name
        admin_username = update.effective_user.username or "Unknown Admin"

        # Notify about retry attempt
        await context.bot.send_message(
            chat_id=update.effective_chat.id,
            text=f"🔄 Retrying file: *{file_name}*"
        )

        try:
            # Download the file again
            file = await context.bot.get_file(document.file_id)
            download_path = os.path.join(DOWNLOAD_DIR, file_name)
            await file.download_to_drive(download_path)

            # Extract and process the file
            extracted_path = self.extract_archive(download_path)
            if extracted_path:
                importer = SaleDataImporter(extracted_path)
                excel_files = [
                    os.path.join(r, f) for r, d, fs in os.walk(extracted_path)
                    for f in fs if f.endswith(('.xlsx', '.xls')) and 'SALE' in f.upper()
                ]
                total_rows = 0
                for excel_file in excel_files:
                    rows = importer.process_excel_file(excel_file, "admin_retry")
                    total_rows += rows

                # Organize files
                organized_tins = self.organize_excel_files(extracted_path)

                # Clean up the downloaded file
                try:
                    os.remove(download_path)
                except Exception as e:
                    self.logger.error(f"Error deleting downloaded file: {e}")

                # Prepare success message
                # success_message = f"✅ File: *{file_name}* processed successfully!"

                # Send success message to the group
                # await context.bot.send_message(
                #     chat_id=update.effective_chat.id,
                #     text=success_message
                # )

                # # Notify production group if needed
                # await self.notify_production_group(context, success_message)

            else:
                # Prepare failure message
                failure_message = f"❌ Retry file: *{file_name}* failed. Please check your file and try again later."

                # Send failure message to the group
                await context.bot.send_message(
                    chat_id=update.effective_chat.id,
                    text=failure_message
                )

                # Notify production group if needed
                await self.notify_production_group(context, failure_message)

        except Exception as e:
            error_message = f"❌ Error processing file: {str(e)}"
            self.logger.error(error_message)
            self.logger.error(traceback.format_exc())

            await context.bot.send_message(
                chat_id=update.effective_chat.id,
                text=error_message
            )

    async def cmd_status(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /status command"""
        if not await self.is_admin(update):
            return

        # Get database stats
        try:
            with sqlite3.connect(DATABASE_PATH) as conn:
                cursor = conn.cursor()
                cursor.execute("SELECT COUNT(*) FROM tax_declaration")
                total_records = cursor.fetchone()[0]

                cursor.execute("SELECT COUNT(DISTINCT tax_registration_id) FROM tax_declaration")
                unique_tins = cursor.fetchone()[0]

                cursor.execute("SELECT MAX(import_timestamp) FROM tax_declaration")
                last_import = cursor.fetchone()[0]

            status_message = (
                "📊 **Bot Status Report**\n\n"
                f"🟢 **Bot Status:** Running\n"
                f"📁 **Database Records:** {total_records:,}\n"
                f"🏢 **Unique TINs:** {unique_tins}\n"
                f"⏱️ **Last Import:** {last_import}\n"
                f"💾 **Database Size:** {os.path.getsize(DATABASE_PATH) / (1024*1024):.2f} MB"
            )
            await update.message.reply_text(status_message)
        except Exception as e:
            await update.message.reply_text(f"❌ Error getting status: {str(e)}")

    async def cmd_logs(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Handle /logs command"""
        if not await self.is_admin(update):
            return

        # Get the last 20 lines from the log file
        try:
            with open('telegram_bot.log', 'r') as log_file:
                logs = log_file.readlines()
                last_logs = logs[-20:] if len(logs) > 20 else logs
                log_text = "📋 **Recent Logs:**\n\n" + "".join(last_logs)

                # If log is too long, send as a file
                if len(log_text) > 4000:
                    with open('recent_logs.txt', 'w') as temp_log:
                        temp_log.write("".join(last_logs))
                    await update.message.reply_document(
                        document=open('recent_logs.txt', 'rb'),
                        caption="📋 Recent logs (last 20 entries)"
                    )
                    os.remove('recent_logs.txt')
                else:
                    await update.message.reply_text(log_text)
        except Exception as e:
            await update.message.reply_text(f"❌ Error retrieving logs: {str(e)}")

    def run(self):
        """Start the Telegram bot with admin commands"""
        try:
            # Create the application
            app = ApplicationBuilder().token(self.token).build()

            # Add document handler
            app.add_handler(MessageHandler(
                filters.Document.ALL,
                self.handle_document
            ))

            # Add admin command handlers
            app.add_handler(CommandHandler("start", self.cmd_start))
            app.add_handler(CommandHandler("stop", self.cmd_stop))
            app.add_handler(CommandHandler("restart", self.cmd_restart))
            app.add_handler(CommandHandler("status", self.cmd_status))
            app.add_handler(CommandHandler("logs", self.cmd_logs))
            app.add_handler(CommandHandler("retry", self.cmd_retry))

            # Set up a post_init callback to set commands and send startup notification
            async def post_init(application):
                """Callback called after initialization"""
                # Set bot commands
                await application.bot.set_my_commands(self.admin_commands)

                # Check if this is a restart by looking for a restart flag file
                restart_flag_file = os.path.join(BASE_DIR, '.restart_flag')
                if os.path.exists(restart_flag_file):
                    try:
                        startup_message = "🤖 Bot Service Started."
                        await application.bot.send_message(
                            chat_id=self.production_chat_id,
                            text=startup_message
                        )
                        # Remove the flag file after sending the message
                        os.remove(restart_flag_file)
                    except Exception as notify_error:
                        self.logger.error(f"Failed to send startup notification: {notify_error}")

            # Register the post_init callback
            app.post_init = post_init

            # Start the bot
            self.logger.info("Bot is running...")
            app.run_polling()

        except Exception as e:
            self.logger.critical(f"Bot startup failed: {e}")
            import traceback
            traceback.print_exc()

def main():
    try:
        # Initialize and run bot
        bot = TelegramSaleDataBot(TELEGRAM_BOT_TOKEN)
        bot.run()
    except Exception as e:
        logging.error(f"Critical error in main process: {e}")

def col_idx_from_letter(column_letter):
    """Convert column letter to column index (1-based)"""
    result = 0
    for c in column_letter:
        result = result * 26 + (ord(c.upper()) - ord('A') + 1)
    return result

if __name__ == "__main__":
    main()
