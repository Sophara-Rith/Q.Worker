import os
import shutil
import logging
import re
import traceback
import uuid
import pandas as pd
import duckdb
import patoolib
import xlsxwriter
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from datetime import datetime
from django.conf import settings
from core.models import UserSettings

logger = logging.getLogger(__name__)

# --- CONFIGURATION ---
BASE_DIR = settings.BASE_DIR
APPDATA_DIR = os.path.join(os.environ.get('APPDATA', os.path.expanduser('~')), 'AuditCore PRO')
DB_PATH = os.path.join(APPDATA_DIR, 'dataWarehouse.duckdb')

# --- SPECIAL CONFIG ---
# TINs that require splitting by single year (Chunk Size = 1)
TINS_SPLIT_BY_YEAR = [
    "L001-100044638", "L001-100103383", "L001-100063098", "L001-104015799",
    "L001-107032791", "L001-100050115", "L001-100181805", "L001-100045138",
    "L001-100057551", "L001-100045766", "L001-100055338", "L001-100011888",
    "L001-901705421", "L001-107000385", "L001-100045707", "L001-100136036",
    "L001-103007369", "L001-100046460", "L001-100044956", "L001-100099718",
    "L001-901503317", "L001-100191681", "L001-100121322", "L001-100121845",
    "L001-100053564", "L001-901501877", "L001-100129250", "L001-100045995",
    "L001-100066011", "L001-100048706", "L001-107041723", "L001-901705389",
    "L001-107030519", "L001-100112145", "L001-106004107", "L001-100190324",
    "L001-100186432", "L001-104013273", "L001-100045782", "L001-100098630",
    "L001-105001619", "L001-100045162", "L001-107040190", "L001-100068901",
    "L001-100074669", "L001-901639143", "L001-100049648", "L001-107024462",
    "L001-150000308", "L001-100120679", "L001-100049915", "L001-901501119"
]

class ProgressTracker:
    _instances = {}

    @classmethod
    def update(cls, task_id, progress, status, details=""):
        cls._instances[task_id] = {
            'progress': progress,
            'status': status,
            'details': details
        }

    @classmethod
    def get(cls, task_id):
        return cls._instances.get(task_id, {'progress': 0, 'status': 'Pending', 'details': ''})

class ConsolidationService:
    def __init__(self, task_id, user):
        self.task_id = task_id
        self.user = user
        self.con = duckdb.connect(DB_PATH)
        
        # --- DYNAMIC OUTPUT DIRECTORY (From Core) ---
        try:
            user_settings = UserSettings.objects.get(user=user)
            self.output_dir = user_settings.default_output_dir
        except UserSettings.DoesNotExist:
            if os.path.exists("D:/"):
                self.output_dir = "D:/AuditCore_Output"
            else:
                self.output_dir = os.path.join(os.path.expanduser("~"), "AuditCore_Output")
        
        os.makedirs(self.output_dir, exist_ok=True)
        self.archive_dir = os.path.join(self.output_dir, 'Archive')
        
        self.init_db()

    def init_db(self):
        """Init DB with EXACT schema."""
        self.con.execute("""
            CREATE SEQUENCE IF NOT EXISTS seq_tax_id;
            CREATE TABLE IF NOT EXISTS tax_declaration (
                id INTEGER DEFAULT nextval('seq_tax_id'),
                date DATE,
                invoice_number VARCHAR,
                credit_notification_letter_number VARCHAR,
                buyer_type VARCHAR,
                tax_registration_id VARCHAR,
                buyer_name VARCHAR,
                total_invoice_amount DECIMAL(15,2),
                amount_exclude_vat DECIMAL(15,2),
                non_vat_sales DECIMAL(15,2),
                vat_zero_rate DECIMAL(15,2),
                vat_local_sale DECIMAL(15,2),
                vat_export DECIMAL(15,2),
                vat_local_sale_state_burden DECIMAL(15,2),
                vat_withheld_by_national_treasury DECIMAL(15,2),
                plt DECIMAL(15,2),
                special_tax_on_goods DECIMAL(15,2),
                special_tax_on_services DECIMAL(15,2),
                accommodation_tax DECIMAL(15,2),
                income_tax_redemption_rate DECIMAL(15,2),
                notes VARCHAR,
                description VARCHAR,
                tax_declaration_status VARCHAR,
                telegram_username VARCHAR,
                import_timestamp TIMESTAMP,
                file_tin VARCHAR,
                branch_name VARCHAR
            );
        """)
        try:
            self.con.execute("SELECT branch_name FROM tax_declaration LIMIT 1")
        except:
            logger.warning("Schema mismatch detected. Recreating table.")
            self.con.execute("DROP TABLE IF EXISTS tax_declaration")
            self.init_db()

    def log(self, percent, status, details, failed_file=None):
        ProgressTracker.update(self.task_id, percent, status, details)

    def extract_branch_info(self, original_title):
        if not original_title or not isinstance(original_title, str): return ""
        location_match = re.search(r'\(([^)]*ទីស្នាក់ការកណ្តាល[^)]*)\)', original_title)
        if location_match: return location_match.group(1).strip()
        branch_match = re.search(r'\(([^)]*សាខា[^)]*)\)', original_title)
        if branch_match: return branch_match.group(1).strip()
        return ''

    def clean_company_name(self, original_title):
        if not original_title: return ""
        company_match = re.search(r'របស់\s*([^(]+(?:\s*\([^)]*\))?)', original_title)
        if company_match:
            company_name = company_match.group(1).strip()
            company_name = re.sub(r'\s*\((?:សាខា[^)]*|ទីស្នាក់ការកណ្តាល[^)]*)\)', '', company_name).strip()
            return company_name
        parts = original_title.split('របស់')
        if len(parts) > 1: return parts[-1].strip()
        return original_title

    def get_month_name(self, month_num):
        months = {1: 'Jan', 2: 'Feb', 3: 'Mar', 4: 'Apr', 5: 'May', 6: 'Jun', 7: 'Jul', 8: 'Aug', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dec'}
        return months.get(int(month_num), str(month_num))

    def format_month_list(self, month_nums):
        if not month_nums: return ""
        nums = sorted(list(set(month_nums)))
        ranges = []
        range_start = nums[0]
        prev = nums[0]
        
        for curr in nums[1:]:
            if curr == prev + 1:
                prev = curr
            else:
                if range_start == prev: ranges.append(self.get_month_name(range_start))
                else: ranges.append(f"{self.get_month_name(range_start)}-{self.get_month_name(prev)}")
                range_start = curr
                prev = curr
        
        if range_start == prev: ranges.append(self.get_month_name(range_start))
        else: ranges.append(f"{self.get_month_name(range_start)}-{self.get_month_name(prev)}")
        return " , ".join(ranges)

    def process(self, file_paths):
        current_file = "Unknown"
        dropped_logs = []
        
        try:
            self.log(5, "Initializing", "Preparing files...")
            extracted_files = []

            # 1. EXTRACT
            for f_path in file_paths:
                f_name = os.path.basename(f_path)
                current_file = f_name
                if f_path.lower().endswith(('.rar', '.zip')):
                    self.log(10, "Extracting", f"Extracting {f_name}...")
                    extract_to = os.path.join(BASE_DIR, 'temp_extract', str(uuid.uuid4()))
                    os.makedirs(extract_to, exist_ok=True)
                    try:
                        patoolib.extract_archive(f_path, outdir=extract_to, verbosity=-1)
                        for root, _, files in os.walk(extract_to):
                            for file in files:
                                if file.lower().endswith(('.xlsx', '.xls')) and not file.startswith('~$'):
                                    extracted_files.append(os.path.join(root, file))
                    except Exception as e:
                         logger.error(f"Extraction failed: {e}")
                elif f_path.lower().endswith(('.xlsx', '.xls')):
                    extracted_files.append(f_path)

            # 2. IMPORT
            total = len(extracted_files)
            if total == 0:
                self.log(100, "Error", "No valid Excel files found.", failed_file=current_file)
                return

            self.log(20, "Importing", f"Found {total} files. Importing...")
            
            tins_processed = set()
            tin_company_map = {} 
            failed_files = []
            
            for i, file in enumerate(extracted_files):
                fname = os.path.basename(file)
                current_file = fname
                
                percent_before = 20 + int((i / total) * 30)
                self.log(percent_before, "Importing", f"Reading {fname} ({i+1} of {total})")
                
                match_full = re.search(r'([A-Za-z0-9]{4}-\d{9})', fname)
                if match_full:
                    tin = match_full.group(1).upper()
                else:
                    match_short = re.search(r'(?<!\d)(\d{9})(?!\d)', fname)
                    tin = match_short.group(1) if match_short else "UNKNOWN"
                
                tins_processed.add(tin)
                branch_info = ""
                
                try:
                    df = pd.read_excel(file, skiprows=3, dtype=str)
                    
                    if len(df.columns) < 10:
                        raise ValueError("File structure does not match expected format.")
                        
                    title_df = pd.read_excel(file, nrows=1, header=None)
                    title_val = title_df.iloc[0, 0] if not title_df.empty else ""
                    
                    if title_val and isinstance(title_val, str):
                        if tin not in tin_company_map:
                            cleaned_name = self.clean_company_name(title_val)
                            tin_company_map[tin] = cleaned_name
                        branch_info = self.extract_branch_info(title_val)
                    
                    # Pad columns to ensure Column0 to Column22 always exist
                    while len(df.columns) < 23:
                        df[f'TempCol{len(df.columns)}'] = None
                    df.columns = [f'Column{j}' for j in range(len(df.columns))]
                    
                    safe_branch = branch_info.replace("'", "''")
                    
                    # Track legitimately dropped rows
                    missing_invoice_mask = df['Column2'].isna() | (df['Column2'].astype(str).str.strip() == '') | (df['Column2'].astype(str).str.lower().str.strip() == 'nan')
                    dropped_count = missing_invoice_mask.sum()
                    
                    if dropped_count > 0:
                        dropped_logs.append(f"📄 {fname} | Dropped: {dropped_count} rows | Reason: Missing Invoice Number")
                    
                    self.con.register('df_view', df)

                    core_fields_str = """
                        date, invoice_number, credit_notification_letter_number, buyer_type,
                        tax_registration_id, buyer_name, total_invoice_amount, 
                        amount_exclude_vat, non_vat_sales, vat_zero_rate, vat_local_sale,
                        vat_export, vat_local_sale_state_burden, vat_withheld_by_national_treasury,
                        plt, special_tax_on_goods, special_tax_on_services, accommodation_tax,
                        income_tax_redemption_rate, notes, description, tax_declaration_status
                    """

                    # -------------------------------------------------------------------
                    # USING YOUR TRUSTED EXCEPT LOGIC FOR PERFECT ITEMIZATION SUPPORT
                    # (Combined with advanced data cleansing to save numeric dropping)
                    # -------------------------------------------------------------------
                    self.con.execute(f"""
                        INSERT INTO tax_declaration 
                        (
                            {core_fields_str},
                            file_tin, import_timestamp, telegram_username, branch_name
                        )
                        SELECT 
                            *,
                            '{tin}', now(), '{self.user.username}', '{safe_branch}'
                        FROM (
                            SELECT 
                                COALESCE(
                                    TRY_STRPTIME(CAST(Column1 AS VARCHAR), '%d-%m-%Y')::DATE, 
                                    TRY_STRPTIME(CAST(Column1 AS VARCHAR), '%d/%m/%Y')::DATE,
                                    TRY_STRPTIME(CAST(Column1 AS VARCHAR), '%m/%d/%Y')::DATE,
                                    TRY_STRPTIME(CAST(Column1 AS VARCHAR), '%Y-%m-%d')::DATE,
                                    TRY_CAST(Column1 AS DATE)
                                ) as date,
                                REGEXP_REPLACE(CAST(Column2 AS VARCHAR), '\.0$', '') as invoice_number, 
                                CAST(Column3 AS VARCHAR) as credit_notification_letter_number, 
                                CAST(Column4 AS VARCHAR) as buyer_type,
                                CAST(Column5 AS VARCHAR) as tax_registration_id, 
                                CAST(Column6 AS VARCHAR) as buyer_name, 
                                TRY_CAST(REPLACE(CAST(Column7 AS VARCHAR), ',', '') AS DECIMAL(15,2)) as total_invoice_amount,
                                TRY_CAST(REPLACE(CAST(Column8 AS VARCHAR), ',', '') AS DECIMAL(15,2)) as amount_exclude_vat,
                                TRY_CAST(REPLACE(CAST(Column9 AS VARCHAR), ',', '') AS DECIMAL(15,2)) as non_vat_sales,
                                TRY_CAST(REPLACE(CAST(Column10 AS VARCHAR), ',', '') AS DECIMAL(15,2)) as vat_zero_rate,
                                TRY_CAST(REPLACE(CAST(Column11 AS VARCHAR), ',', '') AS DECIMAL(15,2)) as vat_local_sale,
                                TRY_CAST(REPLACE(CAST(Column12 AS VARCHAR), ',', '') AS DECIMAL(15,2)) as vat_export,
                                TRY_CAST(REPLACE(CAST(Column13 AS VARCHAR), ',', '') AS DECIMAL(15,2)) as vat_local_sale_state_burden,
                                TRY_CAST(REPLACE(CAST(Column14 AS VARCHAR), ',', '') AS DECIMAL(15,2)) as vat_withheld_by_national_treasury,
                                TRY_CAST(REPLACE(CAST(Column15 AS VARCHAR), ',', '') AS DECIMAL(15,2)) as plt,
                                TRY_CAST(REPLACE(CAST(Column16 AS VARCHAR), ',', '') AS DECIMAL(15,2)) as special_tax_on_goods,
                                TRY_CAST(REPLACE(CAST(Column17 AS VARCHAR), ',', '') AS DECIMAL(15,2)) as special_tax_on_services,
                                TRY_CAST(REPLACE(CAST(Column18 AS VARCHAR), ',', '') AS DECIMAL(15,2)) as accommodation_tax,
                                TRY_CAST(REPLACE(CAST(Column19 AS VARCHAR), ',', '') AS DECIMAL(15,2)) as income_tax_redemption_rate,
                                CAST(Column20 AS VARCHAR) as notes, 
                                CAST(Column21 AS VARCHAR) as description, 
                                CAST(Column22 AS VARCHAR) as tax_declaration_status
                            FROM df_view
                            WHERE Column2 IS NOT NULL 
                              AND TRIM(CAST(Column2 AS VARCHAR)) != ''
                              AND LOWER(TRIM(CAST(Column2 AS VARCHAR))) NOT IN ('nan', 'none', 'null')
                              
                            EXCEPT ALL
                            
                            SELECT {core_fields_str}
                            FROM tax_declaration
                            WHERE file_tin = '{tin}'
                        ) as unique_rows
                    """)
                    self.con.unregister('df_view')
                    
                    percent_after = 20 + int(((i + 1) / total) * 30)
                    self.log(percent_after, "Importing", f"Ingested {fname}")
                    
                except Exception as e:
                    error_msg = str(e)
                    logger.error(f"Failed to import {fname}: {error_msg}")
                    failed_files.append(f"❌ {fname} (Reason: {error_msg})")
                    continue

            # 3. CONSOLIDATE & EXPORT
            self.log(50, "Consolidating", "Generating Master Files...")
            all_summary_stats = []
            
            total_tins = len(tins_processed)
            if total_tins == 0:
                self.log(100, "Failed", "No valid data could be extracted.", failed_file=current_file)
                return

            tin_weight = 40.0 / total_tins

            for idx, tin in enumerate(tins_processed):
                tin_base_pct = 50 + (idx * tin_weight)
                
                self.log(int(tin_base_pct + (tin_weight * 0.1)), "Consolidating", f"Querying database for {tin} ({idx + 1}/{total_tins})...")
                
                df_all = self.con.sql(f"""
                    SELECT * FROM tax_declaration 
                    WHERE file_tin = '{tin}'
                    ORDER BY date ASC
                """).to_df()
                
                if df_all.empty:
                    self.log(int(tin_base_pct + tin_weight), "Consolidating", f"Skipped {tin} (No data)")
                    continue

                self.log(int(tin_base_pct + (tin_weight * 0.2)), "Consolidating", f"Analyzing statistics for {tin}...")
                
                df_all['year'] = df_all['date'].dt.year
                df_all['month'] = df_all['date'].dt.month
                
                stats_grouped = df_all.groupby('year').agg(
                    count=('id', 'count'),
                    months=('month', lambda x: list(x.unique()))
                ).reset_index()

                for _, row in stats_grouped.iterrows():
                    all_summary_stats.append({
                        "tin": tin,
                        "month_desc": self.format_month_list(row['months']),
                        "year": row['year'],
                        "count": row['count']
                    })

                self.log(int(tin_base_pct + (tin_weight * 0.3)), "Consolidating", f"Building Excel workbook for {tin}...")
                
                company_name = tin_company_map.get(tin, tin)
                min_year = int(df_all['year'].min())
                max_year = int(df_all['year'].max())
                total_years_span = max(1, max_year - min_year + 1)

                chunk_size = 1 if tin in TINS_SPLIT_BY_YEAR else 4
                
                master_file = os.path.join(self.output_dir, f"{tin}.xlsx")
                
                # USING HIGH SPEED XLSXWRITER FOR EXPORT
                workbook = xlsxwriter.Workbook(master_file, {'nan_inf_to_errors': True})
                
                current_start = min_year
                has_sheets = False
                
                while current_start <= max_year:
                    progress_ratio = (current_start - min_year) / total_years_span
                    chunk_pct = tin_base_pct + (tin_weight * 0.3) + (tin_weight * 0.6 * progress_ratio)
                    
                    current_end = current_start + chunk_size - 1
                    df_chunk = df_all[(df_all['year'] >= current_start) & (df_all['year'] <= current_end)]
                    
                    if not df_chunk.empty:
                        has_sheets = True
                        c_min_date = df_chunk['date'].min()
                        c_max_date = df_chunk['date'].max()
                        sheet_name = f"{c_min_date.strftime('%m-%Y')} - {c_max_date.strftime('%m-%Y')}"[:31]
                        
                        self.log(int(chunk_pct), "Consolidating", f"Writing sheet '{sheet_name}' for {tin}...")
                        
                        ws = workbook.add_worksheet(name=sheet_name)
                        self.write_sheet_data_fast(workbook, ws, df_chunk, company_name)
                    
                    current_start += chunk_size
                
                if has_sheets:
                    workbook.close()

                self.log(int(tin_base_pct + tin_weight), "Consolidating", f"Saved {tin}.xlsx")

            # 4. GENERATE SUMMARY REPORT
            self.log(90, "Finalizing", "Updating Summary Report...")
            self.generate_summary_report(all_summary_stats)

            shutil.rmtree(os.path.join(BASE_DIR, 'temp_extract'), ignore_errors=True)
            
            if failed_files:
                skipped_list = "\n".join(failed_files)
                warning_msg = f"Processed successfully, but skipped {len(failed_files)} file(s):\n\n{skipped_list}"
                self.log(100, "Completed with Warnings", warning_msg)
            else:
                self.log(100, "Completed", f"Saved to {self.output_dir}")
        
        except Exception as e:
            logger.error(traceback.format_exc())
            self.log(100, "Failed", str(e), failed_file=current_file)
        finally:
            self.con.close()
            
            if dropped_logs:
                print("\n" + "="*70)
                print("🚨 PANDAS DATA CONSOLIDATION - DROPPED ROWS SUMMARY 🚨")
                print("="*70)
                for log in dropped_logs:
                    print(log)
                print("="*70 + "\n")

    def write_sheet_data_fast(self, workbook, ws, df, company_name):
        """Populates a specific worksheet with formatted data using high-speed xlsxwriter."""
        
        # --- STYLES ---
        font_title = workbook.add_format({'font_name': 'Khmer OS Muol Light', 'size': 12, 'align': 'center', 'valign': 'vcenter'})
        font_header = workbook.add_format({'font_name': 'Khmer OS Siemreap', 'size': 11, 'bold': True, 'align': 'center', 'valign': 'vcenter', 'border': 1, 'bg_color': '#B0DBBC', 'text_wrap': True})
        
        format_dict = {
            'normal': workbook.add_format({'font_name': 'Khmer OS Siemreap', 'size': 11, 'border': 1, 'valign': 'vcenter'}),
            'alt': workbook.add_format({'font_name': 'Khmer OS Siemreap', 'size': 11, 'border': 1, 'valign': 'vcenter', 'bg_color': '#F5F5F5'}),
            'date_normal': workbook.add_format({'font_name': 'Khmer OS Siemreap', 'size': 11, 'border': 1, 'valign': 'vcenter', 'num_format': 'dd-mm-yyyy', 'align': 'center'}),
            'date_alt': workbook.add_format({'font_name': 'Khmer OS Siemreap', 'size': 11, 'border': 1, 'valign': 'vcenter', 'bg_color': '#F5F5F5', 'num_format': 'dd-mm-yyyy', 'align': 'center'}),
            'num_normal': workbook.add_format({'font_name': 'Khmer OS Siemreap', 'size': 11, 'border': 1, 'valign': 'vcenter', 'num_format': '#,##0'}),
            'num_alt': workbook.add_format({'font_name': 'Khmer OS Siemreap', 'size': 11, 'border': 1, 'valign': 'vcenter', 'bg_color': '#F5F5F5', 'num_format': '#,##0'}),
        }

        # --- 1. TITLE (Row 1) ---
        title = f"បញ្ជីទិន្នានុប្បវត្តិលក់ របស់ {company_name}"
        ws.merge_range('A1:Y1', title, font_title)
        ws.set_row(0, 30)

        # --- 2. HEADERS (Row 2 & 3) ---
        headers = {
            0: ("ឆ្នាំ", ""), 1: ("ល.រ", ""), 2: ("កាលបរិច្ឆេទ", ""), 3: ("លេខវិក្កយបត្រ ឬប្រតិវេទគយ", ""), 
            4: ("លេខលិខិតជូនដំណឹងឥណទាន", ""), 5: ("អ្នកទិញ", "ប្រភេទ"), 6: ("", "លេខសម្គាល់ចុះបញ្ជីពន្ធដា"), 
            7: ("", "ឈ្មោះ"), 8: ("តម្លៃសរុបលើវិក្កយបត្រ", ""), 9: ("តម្លៃ​​មិន​រួមអតប និងមិនជាប់​អតប​", "តម្លៃមិនរួមអតប"), 
            10: ("", "ការលក់មិនជាប់អតប"), 11: ("អាករលើតម្លៃបន្ថែមអត្រា ០% ", ""), 12: ("អាករលើតម្លៃបន្ថែម", "ការលក់ក្នុងស្រុក"), 
            13: ("", "អតបលើការនាំចេញ"), 14: ("អាករលើតម្លៃបន្ថែម(បន្ទុករដ្ឋ)", ""), 15: ("អតប កាត់ទុកដោយរតនាគារជាតិ", ""), 
            16: ("អាករបំភ្លឺសាធារណៈ", ""), 17: ("អាករពិសេសលើទំនិញមួយចំនួន", ""), 18: ("អាករពិសេសលើសេវាមួយចំនួន", ""), 
            19: ("អាករលើការស្នាក់នៅ", ""), 20: ("អត្រាប្រាក់រំដោះពន្ធលើប្រាក់ចំណូល", ""), 21: ("កំណត់សម្គាល់", ""), 
            22: ("បរិយាយ", ""), 23: ("ស្ថានភាពប្រកាសពន្ធ", ""), 24: ("ស្នាក់ការ", "")
        }

        for col, (p_text, c_text) in headers.items():
            ws.write(1, col, p_text, font_header)
            ws.write(2, col, c_text, font_header)

        # --- 3. APPLY MERGES ---
        ws.merge_range('F2:H2', "អ្នកទិញ", font_header)
        ws.merge_range('J2:K2', "តម្លៃ​​មិន​រួមអតប និងមិនជាប់​អតប​", font_header)
        ws.merge_range('M2:N2', "អាករលើតម្លៃបន្ថែម", font_header)

        vertical_cols = [0, 1, 2, 3, 4, 8, 11, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24]
        for col in vertical_cols:
            ws.merge_range(1, col, 2, col, headers[col][0], font_header)

        # --- 4. DATA WRITING ---
        df = df.sort_values('date')
        current_row = 3
        counter = 1
        previous_year = None

        for _, row in df.iterrows():
            year = row['date'].year if pd.notnull(row['date']) else ""
            if previous_year is not None and year != previous_year:
                current_row += 1 
            previous_year = year

            is_alt = (counter % 2 == 0)
            fmt_base = format_dict['alt'] if is_alt else format_dict['normal']
            fmt_date = format_dict['date_alt'] if is_alt else format_dict['date_normal']
            fmt_num = format_dict['num_alt'] if is_alt else format_dict['num_normal']

            vals = [
                year, counter, row['date'], row['invoice_number'], row['credit_notification_letter_number'],
                row['buyer_type'], row['tax_registration_id'], row['buyer_name'], row['total_invoice_amount'],
                row['amount_exclude_vat'], row['non_vat_sales'], row['vat_zero_rate'], row['vat_local_sale'],
                row['vat_export'], row['vat_local_sale_state_burden'], row['vat_withheld_by_national_treasury'],
                row['plt'], row['special_tax_on_goods'], row['special_tax_on_services'], row['accommodation_tax'],
                row['income_tax_redemption_rate'], row['notes'], row['description'], row['tax_declaration_status'], 
                row['branch_name']
            ]

            for i, val in enumerate(vals):
                if pd.isna(val):
                    val = ""
                    
                if i == 2 and val:
                    ws.write_datetime(current_row, i, val, fmt_date)
                elif 8 <= i <= 20 and val != "":
                    try:
                        ws.write_number(current_row, i, float(val), fmt_num)
                    except ValueError:
                        ws.write(current_row, i, val, fmt_base)
                else:
                    ws.write(current_row, i, val, fmt_base)

            current_row += 1
            counter += 1

        # --- 5. WIDTHS ---
        widths = {
            0: 8, 1: 6, 2: 14, 3: 18, 4: 15, 5: 15, 6: 18, 7: 30, 
            8: 20, 9: 18, 10: 18, 11: 18, 12: 18, 13: 18, 14: 18, 
            15: 18, 16: 18, 17: 18, 18: 18, 19: 18, 20: 15, 21: 20, 22: 25, 23: 20, 24: 25
        }
        for col, w in widths.items():
            ws.set_column(col, col, w)

    def generate_summary_report(self, summary_data_list):
        """
        Update the master summary Excel file.
        Uses openpyxl engine so it seamlessly updates your existing file format.
        """
        if not summary_data_list:
            return None

        try:
            filename = "0.Import_Summary.xlsx"
            filepath = os.path.join(self.output_dir, filename)
            
            existing_data = []
            if os.path.exists(filepath):
                try:
                    existing_df = pd.read_excel(filepath, sheet_name='Summary')
                    existing_df['TIN'] = existing_df['TIN'].astype(str)
                    existing_df = existing_df[
                        (existing_df['TIN'] != 'nan') & 
                        (existing_df['TIN'] != 'GRAND TOTAL') & 
                        (~existing_df['TIN'].str.contains('Total TINs', na=False)) &
                        (existing_df['TIN'].str.strip() != '')
                    ]
                    existing_data = existing_df.to_dict('records')
                except Exception as e:
                    logger.warning(f"Error reading existing summary: {e}")
                    existing_data = []

            updated_tins = set(str(item['tin']) for item in summary_data_list)
            final_data = [row for row in existing_data if str(row.get('TIN')) not in updated_tins]
            
            for item in summary_data_list:
                final_data.append({
                    'TIN': str(item['tin']),
                    'Month': item['month_desc'],
                    'Year': item['year'],
                    'Total Transactions': item['count']
                })
            
            df = pd.DataFrame(final_data)
            if df.empty: return None
            
            if 'No' in df.columns: df = df.drop(columns=['No'])

            df['TIN'] = df['TIN'].astype(str)
            df = df.sort_values(by=['TIN', 'Year'])
            df.insert(0, 'No', 0) 
            
            total_transactions = df['Total Transactions'].sum()
            unique_tins_count = df['TIN'].nunique()

            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Summary')
                
                worksheet = writer.sheets['Summary']
                
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                thin_border = Side(style='thin', color="000000")
                border = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
                center_align = Alignment(horizontal='center', vertical='center')
                left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
                fill_white = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
                fill_grey = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                
                for cell in worksheet[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
                    cell.border = border

                current_tin = None
                start_row = 2
                tin_counter = 1
                use_grey_fill = False 
                
                for row in range(2, worksheet.max_row + 2):
                    cell_value = worksheet.cell(row=row, column=2).value if row <= worksheet.max_row else None
                    
                    if cell_value != current_tin:
                        if current_tin is not None:
                            end_row = row - 1
                            worksheet.cell(row=start_row, column=1).value = tin_counter
                            tin_counter += 1
                            
                            if end_row > start_row:
                                worksheet.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
                                worksheet.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)
                            
                            current_fill = fill_grey if use_grey_fill else fill_white
                            for r in range(start_row, end_row + 1):
                                for c in range(1, 6):
                                    cell = worksheet.cell(row=r, column=c)
                                    cell.fill = current_fill
                                    cell.border = border
                                    if c == 3: cell.alignment = left_align
                                    else: cell.alignment = center_align
                                    if c == 5: cell.number_format = '#,##0'

                            use_grey_fill = not use_grey_fill
                        
                        current_tin = cell_value
                        start_row = row

                last_row = worksheet.max_row + 1
                worksheet.cell(row=last_row, column=1).value = "GRAND TOTAL"
                worksheet.cell(row=last_row, column=2).value = f"Total TINs: {unique_tins_count}"
                worksheet.cell(row=last_row, column=5).value = total_transactions
                
                total_font = Font(bold=True)
                total_fill = PatternFill(start_color="ACB9CA", end_color="ACB9CA", fill_type="solid")
                
                for col in range(1, 6):
                    cell = worksheet.cell(row=last_row, column=col)
                    cell.font = total_font
                    cell.fill = total_fill
                    cell.border = border
                    cell.alignment = center_align
                    if col == 5: cell.number_format = '#,##0'

                worksheet.column_dimensions['A'].width = 15
                worksheet.column_dimensions['B'].width = 25
                worksheet.column_dimensions['C'].width = 35
                worksheet.column_dimensions['D'].width = 10
                worksheet.column_dimensions['E'].width = 20
                
            return filepath
            
        except Exception as e:
            logger.error(f"Error generating summary report: {e}")
            return None

def run_task(task_id, file_paths, user):
    service = ConsolidationService(task_id, user)
    service.process(file_paths)